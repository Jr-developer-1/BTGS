from rest_framework import generics, viewsets, status, serializers # type: ignore
from rest_framework.exceptions import ValidationError # type: ignore
from django.core.exceptions import PermissionDenied # type: ignore
from rest_framework.response import Response # type: ignore
from rest_framework.decorators import action, permission_classes as permission_classes_decorator # type: ignore
from .models import ( # type: ignore
    Trip, Expense, TravelClaim, TravelAdvance, TripOdometer, Dispute, PolicyDocument, BulkActivityBatch, JobReport,
    TravelModeMaster, BookingTypeMaster, OperatorMaster, TravelClassMaster, VehicleMaster, ProviderMaster,
    TicketStatusMaster, QuotaTypeMaster,
    LocalTravelModeMaster, LocalProviderMaster, LocalSubTypeMaster,
    StayTypeMaster, RoomTypeMaster, StayBookingTypeMaster, StayBookingSourceMaster,
    MealCategoryMaster, MealTypeMaster, MealSourceMaster, MealProviderMaster,
    IncidentalTypeMaster, CustomMasterDefinition, CustomMasterValue, MasterModule, TripTracking,
    HistoricalTripStop, FinanceWorkflowStep
)
from travel_masters.models import FuelRateMaster, Cadre, EligibilityRule
from .serializers import ( # type: ignore
    TripSerializer, ExpenseSerializer, TravelClaimSerializer, TravelAdvanceSerializer,
    TripOdometerSerializer, DisputeSerializer, PolicyDocumentSerializer, PolicyDocumentDetailSerializer, BulkActivityBatchSerializer, JobReportSerializer,
    TravelModeMasterSerializer, BookingTypeMasterSerializer, OperatorMasterSerializer, TravelClassMasterSerializer,
    VehicleMasterSerializer, ProviderMasterSerializer, TicketStatusMasterSerializer, QuotaTypeMasterSerializer,
    LocalTravelModeMasterSerializer, LocalProviderMasterSerializer, LocalSubTypeMasterSerializer,
    StayTypeMasterSerializer, RoomTypeMasterSerializer, StayBookingTypeMasterSerializer, StayBookingSourceMasterSerializer,
    MealCategoryMasterSerializer, MealTypeMasterSerializer, MealSourceMasterSerializer, MealProviderMasterSerializer,
    IncidentalTypeMasterSerializer, CustomMasterDefinitionSerializer, CustomMasterValueSerializer, MasterModuleSerializer,
    TripTrackingSerializer, HistoricalTripStopSerializer
)
import io # type: ignore
import json # type: ignore
import pandas as pd # type: ignore
from django.http import HttpResponse # type: ignore
from rest_framework.permissions import AllowAny # type: ignore
from django.db.models import Q # type: ignore
import base64 # type: ignore
import binascii # type: ignore
from api_management.utils import encrypt_key, decrypt_key # type: ignore
from django.utils import timezone # type: ignore
from rest_framework.views import APIView # type: ignore
from django.db.models import Sum, Count, Q # type: ignore
from core.models import User # type: ignore
from notifications.models import Notification # type: ignore
from core.permissions import IsCustomAuthenticated # type: ignore
import requests # type: ignore
import datetime # type: ignore


# ──────────────────────────────────────────────────────────────────────────────
# Claim Report – submitted vs approved per employee & per approver
# ──────────────────────────────────────────────────────────────────────────────
class ClaimReportView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request):
        try:
            user = getattr(request, 'custom_user', None)
            if not user:
                return Response({"error": "Unauthorized"}, status=401)

            # ── Position-code-based access gate ──────────────────────────────
            # Collect every position code/id the user holds from their profile
            from .models import HRPositionConfig, FinanceWorkflowStep

            user_pos_codes = set()

            # 1. active_position_id stored on the user object
            if getattr(user, 'active_position_id', None):
                user_pos_codes.add(str(user.active_position_id).strip())

            # 2. All positions from get_available_positions (id + code + name)
            try:
                for pos in (user.get_available_positions() or []):
                    for key in ('id', 'code', 'name'):
                        val = str(pos.get(key) or '').strip()
                        if val:
                            user_pos_codes.add(val)
            except Exception:
                pass

            # 3. API data (position.code is what admin enters in config UI)
            try:
                api_data = user._get_api_data() if hasattr(user, '_get_api_data') else None
                if api_data:
                    for key in ('position',):
                        p = api_data.get(key)
                        if p:
                            for f in ('id', 'code', 'name'):
                                v = str(p.get(f) or '').strip()
                                if v:
                                    user_pos_codes.add(v)
                    for p in api_data.get('positions_details', []):
                        for f in ('id', 'code', 'name'):
                            v = str(p.get(f) or '').strip()
                            if v:
                                user_pos_codes.add(v)
            except Exception:
                pass

            # Admin check (role-based, same as _is_admin)
            role_name = (user.role.name if user.role else '').lower()
            is_admin_user = role_name in ['admin', 'it-admin', 'superuser']

            # HR check — position code must be in HRPositionConfig with can_view_reports=True
            is_hr_user = (
                HRPositionConfig.objects.filter(
                    position_id__in=user_pos_codes, is_active=True, can_view_reports=True
                ).exists()
                if user_pos_codes else False
            )

            # Finance check — position code must be in FinanceWorkflowStep (or direct user assignment) with can_view_reports=True
            is_finance_user = (
                FinanceWorkflowStep.objects.filter(
                    models.Q(position_id__in=user_pos_codes) | models.Q(user=user),
                    is_active=True,
                    can_view_reports=True
                ).exists()
            )

            if not (is_admin_user or is_hr_user or is_finance_user):
                return Response({"error": "Access restricted to HR, Finance, or Admin."}, status=403)

            # ── Query filters ─────────────────────────────────────────────────
            from_date_str = request.query_params.get('from_date')
            to_date_str   = request.query_params.get('to_date')
            emp_filter    = request.query_params.get('employee')

            qs = TravelClaim.objects.select_related(
                'trip', 'trip__user', 'processed_by', 'final_executive', 'sent_by_executive'
            ).prefetch_related('trip__expenses')

            if from_date_str:
                try:
                    from django.utils.timezone import make_aware
                    dt = datetime.datetime.strptime(from_date_str, '%Y-%m-%d')
                    dt_aware = make_aware(datetime.datetime.combine(dt.date(), datetime.time.min))
                    qs = qs.filter(created_at__gte=dt_aware)
                except ValueError:
                    pass

            if to_date_str:
                try:
                    from django.utils.timezone import make_aware
                    dt = datetime.datetime.strptime(to_date_str, '%Y-%m-%d')
                    dt_aware = make_aware(datetime.datetime.combine(dt.date(), datetime.time.max))
                    qs = qs.filter(created_at__lte=dt_aware)
                except ValueError:
                    pass

            if emp_filter:
                qs = qs.filter(
                    Q(user_name__icontains=emp_filter) |
                    Q(trip__user__name__icontains=emp_filter) |
                    Q(trip__user__employee_id__icontains=emp_filter)   # fixed: employee_id not employee_code
                )

            # ── Per-employee summary ──────────────────────────────────────────
            employee_map = {}
            claims_list = list(qs)  # evaluate once, reuse for both loops

            for claim in claims_list:
                emp_user  = claim.trip.user if claim.trip else None
                emp_name  = claim.user_name or (emp_user.name if emp_user else 'Unknown')
                emp_code  = (getattr(emp_user, 'employee_id', None) if emp_user else None) or emp_name
                dept_val  = claim.user_department or (getattr(emp_user, 'department', '') if emp_user else '') or ''
                desig_val = claim.user_designation or (getattr(emp_user, 'designation', '') if emp_user else '') or ''

                if emp_code not in employee_map:
                    employee_map[emp_code] = {
                        'employee_name': emp_name,
                        'employee_code': emp_code,
                        'department':    dept_val,
                        'designation':   desig_val,
                        'total_submitted': 0, 'total_approved': 0,
                        'total_over_eligibility': 0.0,
                        'total_claims': 0, 'approved_count': 0,
                        'pending_count': 0, 'rejected_count': 0,
                        'claims': [],           # individual claim drill-down
                    }

                submitted_amt = float(claim.total_amount or 0)
                approved_amt  = float(
                    claim.executive_approved_amount or claim.hr_approved_amount or claim.approved_amount or 0
                )

                # Calculate over eligibility for this claim
                over_eligibility_amt = 0.0
                if claim.trip:
                    from travel_masters.eligibility import compute_allowance_for_claim
                    has_none_allowed = False
                    for exp in claim.trip.expenses.all():
                        if not exp.is_deleted and exp.allowed_amount is None:
                            has_none_allowed = True
                            break
                    if has_none_allowed:
                        try:
                            allowance_res = compute_allowance_for_claim(claim)
                            allowance_map = {ea['expense_id']: ea['allowed_amount'] for ea in allowance_res.get('expense_allowances', [])}
                        except Exception:
                            allowance_map = {}
                    else:
                        allowance_map = {}

                    for exp in claim.trip.expenses.all():
                        if exp.is_deleted:
                            continue
                        claimed = float(exp.amount or 0)
                        allowed = allowance_map.get(exp.id)
                        if allowed is None:
                            allowed = float(exp.allowed_amount) if exp.allowed_amount is not None else claimed
                        diff = claimed - allowed
                        if diff > 0:
                            over_eligibility_amt += diff

                employee_map[emp_code]['total_submitted'] += submitted_amt
                employee_map[emp_code]['total_approved']  += approved_amt
                employee_map[emp_code]['total_over_eligibility'] += over_eligibility_amt
                employee_map[emp_code]['total_claims'] += 1

                s = (claim.status or '').lower()
                if s in ['paid', 'completed', 'transferred', 'partially_completed']:
                    employee_map[emp_code]['approved_count'] += 1
                elif s in ['rejected', 'rejected by finance', 'rejected_by_head']:
                    employee_map[emp_code]['rejected_count'] += 1
                else:
                    employee_map[emp_code]['pending_count'] += 1

                # Individual claim row for drill-down
                employee_map[emp_code]['claims'].append({
                    'claim_id':    claim.id,
                    'trip_id':     claim.trip.trip_id if claim.trip else None,
                    'source':      claim.trip.source if claim.trip else '',
                    'destination': claim.trip.destination if claim.trip else '',
                    'start_date':  claim.trip.start_date.strftime('%d-%m-%Y') if claim.trip and claim.trip.start_date else '',
                    'submitted':   round(submitted_amt, 2),
                    'approved':    round(approved_amt,  2),
                    'over_eligibility': round(over_eligibility_amt, 2),
                    'status':      claim.status or '',
                    'created_at':  claim.created_at.strftime('%d-%m-%Y') if claim.created_at else '',
                })

            employee_data = sorted(employee_map.values(), key=lambda x: x['total_submitted'], reverse=True)


            # ── Per-approver summary ──────────────────────────────────────────
            approver_map = {}
            from travel.models import HRIntimation
            from core.models import AuditLog

            hr_intimations = HRIntimation.objects.filter(
                claim__in=claims_list,
                is_read=True
            ).select_related('hr_user', 'claim')
            
            claim_hr_approvers = {}
            for hint in hr_intimations:
                if hint.hr_user:
                    claim_hr_approvers.setdefault(hint.claim_id, []).append(hint.hr_user)

            # Query all AuditLog entries for these claims to find anyone who took action in the middle
            claim_ids = [str(c.id) for c in claims_list]
            audit_logs = AuditLog.objects.filter(
                model_name='TravelClaim',
                object_id__in=claim_ids
            ).select_related('user')

            claim_audit_users = {}
            for log in audit_logs:
                if log.user and log.object_id:
                    try:
                        c_id = int(log.object_id)
                        claim_audit_users.setdefault(c_id, []).append(log.user)
                    except ValueError:
                        pass

            for claim in claims_list:
                claim_approvers = set()
                
                # 1. HR auditor(s) who marked it as read/approved
                for u in claim_hr_approvers.get(claim.id, []):
                    claim_approvers.add(u)

                # 2. Anyone who took an action in the middle (from AuditLog)
                for u in claim_audit_users.get(claim.id, []):
                    claim_approvers.add(u)
                    
                # 3. Finance / final processing approvers
                for u in [claim.processed_by, claim.final_executive, claim.sent_by_executive]:
                    if u:
                        claim_approvers.add(u)
                        
                for approver_obj in claim_approvers:
                    # Filter: Only include HR, Finance, or Admin users
                    is_hr_user = _is_hr(approver_obj)
                    is_fin_user = _is_finance_executive(approver_obj) or _is_finance_head(approver_obj)
                    is_admin_user = hasattr(approver_obj, 'role') and approver_obj.role.name.lower() in ['admin', 'it-admin', 'superuser']
                    
                    dept_lower = (getattr(approver_obj, 'department', '') or '').lower()
                    desig_lower = (getattr(approver_obj, 'designation', '') or '').lower()
                    role_lower = (getattr(approver_obj, 'active_role', '') or '').lower()
                    
                    is_text_hr = 'hr' in dept_lower or 'human resource' in dept_lower or 'hr' in desig_lower or 'hr' in role_lower
                    is_text_fin = 'finance' in dept_lower or 'accounts' in dept_lower or 'finance' in desig_lower or 'cfo' in desig_lower or 'finance' in role_lower
                    
                    if not (is_hr_user or is_fin_user or is_admin_user or is_text_hr or is_text_fin):
                        continue

                    appr_code = getattr(approver_obj, 'employee_id', None) or getattr(approver_obj, 'employee_code', None) or str(approver_obj.id)
                    appr_name = getattr(approver_obj, 'name', appr_code) or appr_code
                    
                    if appr_code not in approver_map:
                        approver_map[appr_code] = {
                            'approver_name': appr_name,
                            'approver_code': appr_code,
                            'approver_dept': (getattr(approver_obj, 'department', '') or ''),
                            'total_processed': 0, 'total_approved_amount': 0,
                            'total_submitted_amount': 0, 'approved_count': 0, 'rejected_count': 0,
                            'claims': [],
                        }
                        
                    approver_map[appr_code]['total_processed']        += 1
                    approver_map[appr_code]['total_submitted_amount'] += float(claim.total_amount or 0)
                    # Determine the approved amount based on the approver's role/stage
                    if is_hr_user or is_text_hr:
                        appr_amt = claim.hr_approved_amount
                        if appr_amt is None:
                            appr_amt = claim.executive_approved_amount or claim.approved_amount or 0
                    elif is_fin_user or is_text_fin:
                        appr_amt = claim.executive_approved_amount
                        if appr_amt is None:
                            appr_amt = claim.hr_approved_amount or claim.approved_amount or 0
                    else:
                        appr_amt = claim.executive_approved_amount or claim.hr_approved_amount or claim.approved_amount or 0

                    s = (claim.status or '').lower()
                    if s in ['rejected', 'rejected by finance', 'rejected_by_head']:
                        appr_amt = 0

                    approver_map[appr_code]['total_approved_amount']  += float(appr_amt or 0)
                    
                    approver_map[appr_code]['claims'].append({
                        'claim_id':    claim.id,
                        'trip_id':     claim.trip.trip_id if claim.trip else None,
                        'source':      claim.trip.source if claim.trip else '',
                        'destination': claim.trip.destination if claim.trip else '',
                        'start_date':  claim.trip.start_date.strftime('%d-%m-%Y') if claim.trip and claim.trip.start_date else '',
                        'submitted':   round(float(claim.total_amount or 0), 2),
                        'approved':    round(float(appr_amt or 0), 2),
                        'status':      claim.status or '',
                        'created_at':  claim.created_at.strftime('%d-%m-%Y') if claim.created_at else '',
                        'employee_name': claim.user_name or (claim.user.name if claim.user else ''),
                    })

                    s = (claim.status or '').lower()
                    if s in ['paid', 'completed', 'transferred', 'partially_completed']:
                        approver_map[appr_code]['approved_count'] += 1
                    elif s in ['rejected', 'rejected by finance', 'rejected_by_head']:
                        approver_map[appr_code]['rejected_count'] += 1

            approver_data = sorted(approver_map.values(), key=lambda x: x['total_processed'], reverse=True)

            # ── KPI summary ───────────────────────────────────────────────────
            total_submitted = sum(e['total_submitted'] for e in employee_data)
            total_approved  = sum(e['total_approved']  for e in employee_data)
            total_claims    = sum(e['total_claims']    for e in employee_data)
            total_over_eligibility = sum(e['total_over_eligibility'] for e in employee_data)

            return Response({
                'summary': {
                    'total_claims':    total_claims,
                    'total_submitted': round(total_submitted, 2),
                    'total_approved':  round(total_approved,  2),
                    'savings':         round(total_submitted - total_approved, 2),
                    'total_over_eligibility': round(total_over_eligibility, 2),
                    'approval_rate':   round((total_approved / total_submitted * 100) if total_submitted else 0, 1),
                    'approved_count':  sum(e['approved_count'] for e in employee_data),
                    'pending_count':   sum(e['pending_count']  for e in employee_data),
                    'rejected_count':  sum(e['rejected_count'] for e in employee_data),
                },
                'by_employee': employee_data,
                'by_approver': approver_data,
            })

        except Exception as exc:
            import traceback, logging
            logging.error(f"ClaimReportView error: {exc}", exc_info=True)
            traceback.print_exc()
            return Response({"error": f"Report generation failed: {str(exc)}"}, status=500)


class ClaimReportExportExcelView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request):
        try:
            from .models import HRPositionConfig, FinanceWorkflowStep
            user = getattr(request, 'custom_user', None)
            if not user:
                return Response({"error": "Unauthorized"}, status=401)

            # ── Position-code-based access gate ──────────────────────────────
            user_pos_codes = set()

            if getattr(user, 'active_position_id', None):
                user_pos_codes.add(str(user.active_position_id).strip())

            try:
                for pos in (user.get_available_positions() or []):
                    for key in ('id', 'code', 'name'):
                        val = str(pos.get(key) or '').strip()
                        if val:
                            user_pos_codes.add(val)
            except Exception:
                pass

            try:
                api_data = user._get_api_data() if hasattr(user, '_get_api_data') else None
                if api_data:
                    for key in ('position',):
                        p = api_data.get(key)
                        if p:
                            for f in ('id', 'code', 'name'):
                                v = str(p.get(f) or '').strip()
                                if v:
                                    user_pos_codes.add(v)
                    for p in api_data.get('positions_details', []):
                        for f in ('id', 'code', 'name'):
                            v = str(p.get(f) or '').strip()
                            if v:
                                user_pos_codes.add(v)
            except Exception:
                pass

            role_name = (user.role.name if user.role else '').lower()
            is_admin_user = role_name in ['admin', 'it-admin', 'superuser']

            is_hr_user = (
                HRPositionConfig.objects.filter(
                    position_id__in=user_pos_codes, is_active=True, can_view_reports=True
                ).exists()
                if user_pos_codes else False
            )

            is_finance_user = (
                FinanceWorkflowStep.objects.filter(
                    models.Q(position_id__in=user_pos_codes) | models.Q(user=user),
                    is_active=True,
                    can_view_reports=True
                ).exists()
            )

            if not (is_admin_user or is_hr_user or is_finance_user):
                return Response({"error": "Access restricted to HR, Finance, or Admin."}, status=403)

            # ── Query filters ─────────────────────────────────────────────────
            from_date_str = request.query_params.get('from_date')
            to_date_str   = request.query_params.get('to_date')
            emp_filter    = request.query_params.get('employee')

            qs = TravelClaim.objects.select_related(
                'trip', 'trip__user', 'processed_by', 'final_executive', 'sent_by_executive'
            ).prefetch_related('trip__expenses')

            if from_date_str:
                try:
                    from django.utils.timezone import make_aware
                    import datetime
                    dt = datetime.datetime.strptime(from_date_str, '%Y-%m-%d')
                    dt_aware = make_aware(datetime.datetime.combine(dt.date(), datetime.time.min))
                    qs = qs.filter(created_at__gte=dt_aware)
                except ValueError:
                    pass

            if to_date_str:
                try:
                    from django.utils.timezone import make_aware
                    import datetime
                    dt = datetime.datetime.strptime(to_date_str, '%Y-%m-%d')
                    dt_aware = make_aware(datetime.datetime.combine(dt.date(), datetime.time.max))
                    qs = qs.filter(created_at__lte=dt_aware)
                except ValueError:
                    pass

            if emp_filter:
                qs = qs.filter(
                    models.Q(user_name__icontains=emp_filter) |
                    models.Q(trip__user__name__icontains=emp_filter) |
                    models.Q(trip__user__employee_id__icontains=emp_filter)
                )

            data = []
            claims_list = list(qs)

            for claim in claims_list:
                emp_user  = claim.trip.user if claim.trip else None
                emp_name  = claim.user_name or (emp_user.name if emp_user else 'Unknown')
                emp_code  = (getattr(emp_user, 'employee_id', None) if emp_user else None) or emp_name
                dept_val  = claim.user_department or (getattr(emp_user, 'department', '') if emp_user else '') or ''
                desig_val = claim.user_designation or (getattr(emp_user, 'designation', '') if emp_user else '') or ''

                submitted_amt = float(claim.total_amount or 0)
                approved_amt  = float(
                    claim.executive_approved_amount or claim.hr_approved_amount or claim.approved_amount or 0
                )

                # Calculate over eligibility for this claim
                over_eligibility_amt = 0.0
                if claim.trip:
                    from travel_masters.eligibility import compute_allowance_for_claim
                    has_none_allowed = False
                    for exp in claim.trip.expenses.all():
                        if not exp.is_deleted and exp.allowed_amount is None:
                            has_none_allowed = True
                            break
                    if has_none_allowed:
                        try:
                            allowance_res = compute_allowance_for_claim(claim)
                            allowance_map = {ea['expense_id']: ea['allowed_amount'] for ea in allowance_res.get('expense_allowances', [])}
                        except Exception:
                            allowance_map = {}
                    else:
                        allowance_map = {}

                    for exp in claim.trip.expenses.all():
                        if exp.is_deleted:
                            continue
                        claimed = float(exp.amount or 0)
                        allowed = allowance_map.get(exp.id)
                        if allowed is None:
                            allowed = float(exp.allowed_amount) if exp.allowed_amount is not None else claimed
                        diff = claimed - allowed
                        if diff > 0:
                            over_eligibility_amt += diff

                data.append({
                    'Employee Code': emp_code,
                    'Employee Name': emp_name,
                    'Department': dept_val,
                    'Designation': desig_val,
                    'Trip ID': claim.trip.trip_id if claim.trip else '',
                    'Source': claim.trip.source if claim.trip else '',
                    'Destination': claim.trip.destination if claim.trip else '',
                    'Start Date': claim.trip.start_date.strftime('%Y-%m-%d') if claim.trip and claim.trip.start_date else '',
                    'Submitted Amount': round(submitted_amt, 2),
                    'Approved Amount': round(approved_amt, 2),
                    'Over Eligibility Amount': round(over_eligibility_amt, 2),
                    'Status': claim.status or '',
                    'Date Created': claim.created_at.strftime('%Y-%m-%d') if claim.created_at else ''
                })

            if not data:
                return Response({"error": "No records found to export"}, status=400)

            import pandas as pd
            import io
            import datetime
            from django.http import HttpResponse

            df = pd.DataFrame(data)
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Claims')
            
            output.seek(0)
            
            filename = f"Claims_Reconciliation_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        except Exception as e:
            import traceback
            print(f"DEBUG: Export Excel Error: {str(e)}")
            print(traceback.format_exc())
            return Response({"error": str(e)}, status=500)



def _is_admin(user):
    """Checks if a user is an administrator."""
    if not user or not user.role: return False
    return user.role.name.lower() in ['admin', 'it-admin', 'superuser']

def _can_user_edit_amount(user, obj, is_hr, is_finance, is_finance_exec, is_finance_head, user_step):
    has_edit_claim_perm = False
    if user and hasattr(user, 'role') and user.role:
        from django.db.models import Q
        from core.models import Role
        role_from_api = getattr(user, 'role_from_api', None)
        designation = getattr(user, 'designation', None)
        matching_role = None
        if role_from_api or designation:
            q_obj = Q()
            if role_from_api:
                q_obj |= Q(name__iexact=role_from_api)
            if designation:
                q_obj |= Q(name__iexact=designation)
            matching_role = Role.objects.filter(q_obj).first()
        user_role_obj = matching_role or user.role
        if user_role_obj and isinstance(user_role_obj.permissions, dict):
            has_edit_claim_perm = (
                user_role_obj.permissions.get('can_edit_submitted_claim', False) or
                user_role_obj.permissions.get('can_edit_claims', False)
            )

    if _is_admin(user) or has_edit_claim_perm:
        return True

    if not isinstance(obj, (TravelClaim, TravelAdvance, BulkActivityBatch)):
        return False
        
    if is_hr:
        project_code = 'General'
        if isinstance(obj, Trip):
            project_code = obj.project_code or 'General'
        elif hasattr(obj, 'trip') and obj.trip:
            project_code = obj.trip.project_code or 'General'
        elif isinstance(obj, BulkActivityBatch) and obj.trip:
            project_code = obj.trip.project_code or 'General'
            
        # Fallback to requester's project_code if general/empty
        requester = obj.user if hasattr(obj, 'user') else (obj.trip.user if hasattr(obj, 'trip') and obj.trip else None)
        if (not project_code or project_code in ['General', 'N/A']) and requester:
            if hasattr(requester, 'project_code') and requester.project_code and requester.project_code != 'N/A':
                project_code = requester.project_code

        configs = _get_matching_hr_configs(user, project_code=project_code)
        if not configs.exists():
            configs = _get_matching_hr_configs(user, project_code='General')
            
        return configs.filter(edit_claims='CAN_EDIT').exists()
        
    if is_finance:
        return user_step.can_edit_amount if user_step else (is_finance_exec or is_finance_head)
        
    return False

class FinanceExportExcelView(APIView):
    permission_classes = [IsCustomAuthenticated]
    
    def get(self, request):
        user = getattr(request, 'custom_user', None)
        if not (_is_finance_executive(user) or _is_finance_head(user) or _is_admin(user)):
            return Response({"error": "Unauthorized"}, status=403)
            
        tab = request.query_params.get('tab', 'pending')
        
        # Get querysets based on tab
        if tab == 'completed':
            statuses = ['Paid', 'COMPLETED', 'Completed', 'Settled', 'Transferred']
        elif tab == 'processing':
            statuses = ['Under Process']
        elif tab == 'rejected':
            statuses = ['Rejected', 'Rejected by Finance', 'Cancelled']
        else:
            # Action Required (Pending) - Include PENDING_FINAL_RELEASE and PARTIALLY_COMPLETED
            statuses = ['PENDING_FINAL_RELEASE', 'PARTIALLY_COMPLETED']
            
        advances = TravelAdvance.objects.filter(status__in=statuses)
        claims = TravelClaim.objects.filter(status__in=statuses)
        # Trips (Requests) are terminal at HR and redundant in Finance Export; skipping as requested
        trips = []

        data = []
        # Add advances to list
        for adv in advances:
            total_amt = adv.executive_approved_amount or adv.hr_approved_amount or adv.requested_amount
            balance = float(total_amt) - float(adv.paid_amount or 0)
            
            # Skip partially completed records if balance is already cleared
            if adv.status == 'PARTIALLY_COMPLETED' and balance <= 0:
                continue

            user_obj = adv.trip.user if adv.trip and adv.trip.user else None
            data.append({
                'ID': f"ADVANCE-{adv.id}",
                'Employee Name': user_obj.name if user_obj else (adv.user_name or "N/A"),
                'Amount': balance, # Show remaining balance
                'Bank Name': user_obj.bank_name if user_obj else "N/A",
                'Account Num': user_obj.full_account_no if user_obj else "N/A",
                'IFSC Code': user_obj.ifsc_code if user_obj else "N/A",
                'Status': adv.status,
                'Payment Mode': adv.payment_mode or 'NEFT',
                'Transaction ID': adv.transaction_id or '',
                'Payment Date': adv.payment_date.strftime("%Y-%m-%d") if adv.payment_date else '',
                'Finance Remarks': adv.finance_remarks or '',
                'Trip ID': adv.trip.trip_id if adv.trip else "N/A",
                'Date Created': adv.created_at.strftime("%Y-%m-%d") if adv.created_at else "N/A"
            })

        # Add claims to list
        for claim in claims:
            # 1. Base Amount (Approved or Total)
            total_amt = float(claim.executive_approved_amount or claim.hr_approved_amount or claim.total_amount or 0)
            
            # 2. Subtract Advances associated with this trip
            total_adv = 0
            if claim.trip:
                total_adv = float(claim.trip.advances.filter(status__in=['COMPLETED', 'Paid', 'Settled', 'Transferred']).aggregate(s=Sum('executive_approved_amount'))['s'] or 0)
            
            # 3. Subtract Wallet Balance
            wallet_bal = float(claim.trip.user.carry_forward_balance or 0) if claim.trip and claim.trip.user else 0
            
            # 4. Calculate Net Payout (consistent with ApprovalsView)
            net_payout = total_amt - total_adv - wallet_bal
            
            # 5. Remaining balance (subtracting what was already paid for THIS claim)
            balance = net_payout - float(claim.paid_amount or 0)
            
            # Skip partially completed records if balance is already cleared, or if balance is zero
            if (claim.status == 'PARTIALLY_COMPLETED' and balance <= 0) or balance <= 0:
                continue

            user_obj = claim.trip.user if claim.trip and claim.trip.user else None
            data.append({
                'ID': f"CLAIM-{claim.id}",
                'Employee Name': user_obj.name if user_obj else (claim.user_name or "N/A"),
                'Amount': max(0, balance), # Show remaining net payable amount
                'Bank Name': user_obj.bank_name if user_obj else "N/A",
                'Account Num': user_obj.full_account_no if user_obj else "N/A",
                'IFSC Code': user_obj.ifsc_code if user_obj else "N/A",
                'Status': claim.status,
                'Payment Mode': claim.payment_mode or 'NEFT',
                'Transaction ID': claim.transaction_id or '',
                'Payment Date': claim.payment_date.strftime("%Y-%m-%d") if claim.payment_date else '',
                'Finance Remarks': claim.finance_remarks or '',
                'Trip ID': claim.trip.trip_id if claim.trip else "N/A",
                'Date Created': claim.created_at.strftime("%Y-%m-%d") if claim.created_at else "N/A"
            })

        # Removed Trip (Monthly Tour Plan) export to avoid duplication with Claims
        
        if not data:
            return Response({"error": "There is no pending record to process the payment"}, status=400)
            
        df = pd.DataFrame(data)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Payouts')
        
        output.seek(0)
        
        filename = f"Finance_Export_{tab}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Add Data Validation for Dropdowns using openpyxl
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl import load_workbook
        
        wb = load_workbook(output)
        ws = wb.active
        
        # 1. Status Dropdown
        status_col_idx = None
        mode_col_idx = None
        for cell in ws[1]:
            if cell.value == 'Status':
                status_col_idx = cell.column_letter
            if cell.value == 'Payment Mode':
                mode_col_idx = cell.column_letter
        
        if status_col_idx:
            statuses_list = ['Paid', 'COMPLETED', 'PARTIALLY_COMPLETED', 'Under Process', 'REJECTED_BY_HEAD']
            dv_status = DataValidation(type="list", formula1=f'"{",".join(statuses_list)}"', allow_blank=True)
            ws.add_data_validation(dv_status)
            dv_status.add(f"{status_col_idx}2:{status_col_idx}1000")
            
        # 2. Payment Mode Dropdown
        if mode_col_idx:
            modes_list = ['NEFT', 'Bank Transfer', 'UPI', 'Cash']
            dv_mode = DataValidation(type="list", formula1=f'"{",".join(modes_list)}"', allow_blank=True)
            ws.add_data_validation(dv_mode)
            dv_mode.add(f"{mode_col_idx}2:{mode_col_idx}1000")

        # Save back to output
        new_output = io.BytesIO()
        wb.save(new_output)
        new_output.seek(0)
        
        response = HttpResponse(
            new_output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

class FinanceBulkImportView(APIView):
    permission_classes = [IsCustomAuthenticated]
    
    def post(self, request):
        user = getattr(request, 'custom_user', None)
        if not (_is_finance_executive(user) or _is_finance_head(user) or _is_admin(user)):
            return Response({"error": "Unauthorized"}, status=403)
            
        if 'file' not in request.FILES:
            return Response({"error": "No file uploaded"}, status=400)
            
        import_file = request.FILES['file']
        try:
            df = pd.read_excel(import_file)
            
            # Expected columns: ID, Status, Remarks (optional)
            if 'ID' not in df.columns or 'Status' not in df.columns:
                return Response({"error": "Excel must contain 'ID' and 'Status' columns"}, status=400)
                
            updated_count = 0
            errors = []
            
            for index, row in df.iterrows():
                task_id = str(row.get('ID', '')).strip()
                new_status = str(row.get('Status', '')).strip()
                
                # New fields from updated template
                payment_mode = str(row.get('Payment Mode', '')).strip()
                transaction_id = str(row.get('Transaction ID', '')).strip()
                payment_date_val = row.get('Payment Date')
                remarks = str(row.get('Finance Remarks', row.get('Remarks', ''))).strip()
                
                if not task_id or task_id == 'nan': continue
                
                try:
                    obj = None
                    if task_id.startswith('ADV-'):
                        obj = TravelAdvance.objects.filter(id=task_id.replace('ADV-', '')).first()
                        total_amt = float(obj.executive_approved_amount or obj.hr_approved_amount or obj.requested_amount) if obj else 0
                    elif task_id.startswith('CLAIM-'):
                        obj = TravelClaim.objects.filter(id=task_id.replace('CLAIM-', '')).first()
                        total_amt = float(obj.executive_approved_amount or obj.hr_approved_amount or obj.total_amount) if obj else 0
                    elif task_id.startswith('TRIP-'):
                        obj = Trip.objects.filter(trip_id=task_id.replace('TRIP-', '')).first()
                        if obj:
                            expense_sum = float(obj.expenses.aggregate(s=Sum('amount'))['s'] or 0)
                            advance_sum = float(obj.advances.filter(status__in=['Paid', 'COMPLETED', 'Transferred']).aggregate(s=Sum('executive_approved_amount'))['s'] or 0)
                            wallet_bal = float(obj.user.carry_forward_balance or 0) if obj.user else 0
                            gross = float(obj.executive_approved_amount if obj.executive_approved_amount is not None else expense_sum)
                            total_amt = gross - advance_sum - wallet_bal
                        else:
                            total_amt = 0
                    
                    if obj:
                        # 1. Update Basic Fields
                        if new_status and new_status != 'nan':
                            obj.status = new_status
                        
                        if remarks and remarks != 'nan':
                            obj.finance_remarks = remarks
                            
                        if payment_mode and payment_mode != 'nan':
                            obj.payment_mode = payment_mode
                            
                        if transaction_id and transaction_id != 'nan':
                            obj.transaction_id = transaction_id
                            
                        obj.processed_by = user
                        
                        # 2. Handle Partial vs Full Payment Amount
                        excel_amt = 0
                        try:
                            excel_amt = float(row.get('Amount', 0))
                        except:
                            pass

                        if obj.status == 'PARTIALLY_COMPLETED':
                            # Add the current payment to cumulative paid_amount
                            obj.paid_amount = float(obj.paid_amount or 0) + excel_amt
                            
                            # Auto-complete if balance is now cleared
                            if obj.paid_amount >= total_amt:
                                obj.status = 'Paid' if task_id.startswith('ADV-') else 'Completed'
                        elif obj.status in ['Paid', 'Completed', 'Transferred']:
                            # Set full amount as paid
                            obj.paid_amount = total_amt

                        # 3. Handle Payment Date
                        if payment_date_val and not pd.isna(payment_date_val):
                            try:
                                if isinstance(payment_date_val, (datetime.datetime, datetime.date)):
                                    obj.payment_date = payment_date_val
                                else:
                                    obj.payment_date = pd.to_datetime(payment_date_val).to_pydatetime()
                            except:
                                pass
                        
                        if obj.status in ['Paid', 'Completed', 'Transferred'] and not obj.payment_date:
                            obj.payment_date = timezone.now()
                            
                        obj.save()
                        
                        # 3. Notify employee
                        if isinstance(obj, Trip):
                            requester = obj.user
                            trip_id = obj.trip_id
                        else:
                            requester = obj.trip.user if hasattr(obj, 'trip') and obj.trip else None
                            trip_id = obj.trip.trip_id if hasattr(obj, 'trip') and obj.trip else "N/A"

                        if requester:
                            if obj.status == 'PARTIALLY_COMPLETED':
                                title = "Payment Partially Credited"
                                msg = f"₹{excel_amt:,.2f} has been partially credited for Trip {trip_id}. Remaining balance will be processed soon."
                            elif obj.status in ['Paid', 'Completed', 'Transferred']:
                                title = "Amount Credited"
                                msg = f"₹{excel_amt:,.2f} has been fully credited for Trip {trip_id}."
                            else:
                                title = "Payment Status Updated"
                                msg = f"Your {task_id} status has been updated to {obj.status} by Finance."

                            Notification.objects.create(
                                user=requester,
                                target_position=obj.requester_position if hasattr(obj, 'requester_position') else None,
                                title=title,
                                message=msg,
                                type='info'
                            )
                        updated_count += 1
                    else:
                        errors.append(f"Row {index+2}: Record {task_id} not found.")
                except Exception as e:
                    errors.append(f"Row {index+2} (ID: {task_id}): {str(e)}")
            
            return Response({
                "message": f"Successfully updated {updated_count} records.",
                "errors": errors
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({"error": f"Error processing file: {str(e)}"}, status=400)

def decode_id(encoded_id):
    s_id = str(encoded_id) if encoded_id else ""
    if not s_id:
        return None
        
    try:
        # Check if it's already a numeric-looking ID or doesn't look like base64
        if s_id.isdigit() or s_id.startswith('TRP-') or s_id.startswith('ITS-'):
            return s_id
            
        padding = 4 - (len(s_id) % 4)
        if padding != 4:
            s_id += '=' * padding
        s_id = s_id.replace('-', '+').replace('_', '/')
        decoded_bytes = base64.b64decode(s_id)
        return decoded_bytes.decode('utf-8')
    except (binascii.Error, UnicodeDecodeError, ValueError):
        return encoded_id

def get_users_by_position(position_id):
    """Returns list of User objects that currently hold the specified position_id."""
    if not position_id: return []
    from core.models import User
    from api_management.services import fetch_employee_data, safe_cache_get, safe_cache_set
    
    position_id_str = str(position_id).strip()
    users = []
    
    # Pre-compile lookup keys including config prefix fallback
    extra_lookup_keys = [position_id_str]
    try:
        from travel.models import HRPositionConfig
        config = HRPositionConfig.objects.filter(position_id=position_id_str).first()
        if config:
            config_name_clean = str(config.position_name).strip()
            if ' (' in config_name_clean:
                prefix = config_name_clean.split(' (')[0].strip()
            else:
                prefix = config_name_clean
            extra_lookup_keys.append(prefix)
            extra_lookup_keys.append(config_name_clean)
    except Exception:
        pass
    
    # 1. Fast lookup: bypassed DB lookups to rely purely on API/Cache source of truth
    
    # 2. Optimized Lookup via Cache Map (replaces redundant synchronous O(N) iteration)
    # We build and cache a map of { position_id -> [employee_codes] }
    CACHE_KEY = 'position_to_employee_codes_map'
    pos_map = safe_cache_get(CACHE_KEY)
    
    if pos_map is None:
        pos_map = {}
        persistent_data = safe_cache_get('GLOBAL_EMPLOYEE_DATA')
        if persistent_data:
            user_pos_map = {} # employee_code -> { pos_id -> [ids/codes] }
            
            # Pre-compile in-memory id-to-code map from cache to reduce time complexity from O(N^2) to O(N)
            id_to_code_map = {}
            for emp_item in persistent_data:
                emp = emp_item.get('employee', {})
                e_id = str(emp.get('id') or '')
                e_code = emp.get('employee_code')
                if e_id and e_code:
                    id_to_code_map[e_id] = e_code

            for emp_item in persistent_data:
                emp_code = emp_item.get('employee', {}).get('employee_code')
                if not emp_code: continue
                
                pos_ids = set()
                user_pos_data = {}
                
                def _add_pos(p):
                    p_id = str(p.get('id') or '')
                    p_code = str(p.get('code') or '').strip()
                    if p_id:
                        pos_ids.add(p_id)
                        if p_id not in user_pos_data: user_pos_data[p_id] = []
                        if p_id not in user_pos_data[p_id]: user_pos_data[p_id].append(p_id)
                        if p_code and p_code not in user_pos_data[p_id]: 
                            user_pos_data[p_id].append(p_code)
                    if p_code:
                        pos_ids.add(p_code)

                if emp_item.get('position'):
                    _add_pos(emp_item['position'])
                    for rto in emp_item['position'].get('reporting_to', []):
                        if isinstance(rto, dict):
                            rto_id = str(rto.get('id') or '')
                            rto_code = rto.get('employee_code') or rto.get('employee_id')
                            if rto_id and rto_code:
                                if str(rto_code).strip().isdigit():
                                    resolved = id_to_code_map.get(str(rto_code).strip())
                                    if resolved:
                                        rto_code = resolved
                                if rto_code:
                                    if rto_id not in pos_map:
                                        pos_map[rto_id] = []
                                    if rto_code not in pos_map[rto_id]:
                                        pos_map[rto_id].append(rto_code)

                for pd in (emp_item.get('positions_details') or []):
                    _add_pos(pd)
                    for rto in pd.get('reporting_to', []):
                        if isinstance(rto, dict):
                            rto_id = str(rto.get('id') or '')
                            rto_code = rto.get('employee_code') or rto.get('employee_id')
                            if rto_id and rto_code:
                                if str(rto_code).strip().isdigit():
                                    resolved = id_to_code_map.get(str(rto_code).strip())
                                    if resolved:
                                        rto_code = resolved
                                if rto_code:
                                    if rto_id not in pos_map:
                                        pos_map[rto_id] = []
                                    if rto_code not in pos_map[rto_id]:
                                        pos_map[rto_id].append(rto_code)
                
                for pid in pos_ids:
                    if pid:
                        if pid not in pos_map:
                            pos_map[pid] = []
                        if emp_code not in pos_map[pid]:
                            pos_map[pid].append(emp_code)
                
                user_pos_map[emp_code] = user_pos_data
            
            # Cache mapping for 24 hours to ensure lightning-fast subsequent lookups
            safe_cache_set(CACHE_KEY, pos_map, timeout=86400)
            safe_cache_set('user_position_identifiers', user_pos_map, timeout=86400)
        else:
            # Cache is completely empty. Trigger background load of the global employee list
            # to prevent blocking the user request.
            lock_key = 'GLOBAL_EMPLOYEE_DATA_REFRESH_LOCK'
            from django.core.cache import cache
            import threading
            from api_management.services import _bg_refresh_global_employee_cache
            try:
                if cache.add(lock_key, '1', timeout=7200):
                    t = threading.Thread(target=_bg_refresh_global_employee_cache)
                    t.daemon = True
                    t.start()
            except Exception as e:
                print(f"Failed to trigger background refresh in get_users_by_position: {e}")
            
    # Retrieve matching codes instantly (O(1))
    emp_codes = []
    if pos_map:
        for key in extra_lookup_keys:
            for code in pos_map.get(key, []):
                if code not in emp_codes:
                    emp_codes.append(code)
                    
    for emp_code in emp_codes:
        local_user = User._get_or_create_shell_user(emp_code)
        if local_user and local_user not in users:
            users.append(local_user)
            
    if not users:
        # Fallback: check local active users directly (resilient to empty cache/offline API)
        for local_u in User.objects.filter(is_active=True):
            u_pos_ids = _get_user_all_position_ids(local_u)
            if any(k in u_pos_ids for k in extra_lookup_keys):
                if local_u not in users:
                    users.append(local_u)
            
    return users

def trigger_parallel_dispatch(obj, user=None):
    """
    Sends sequential HR dispatch or initializes Finance Workflow.
    """
    from .models import HRPositionConfig, HRIntimation, FinanceWorkflowStep
    from notifications.models import Notification
    from .utils import _is_coo_position, _is_hr
    import threading

    requester = obj.user if hasattr(obj, 'user') else (obj.trip.user if hasattr(obj, 'trip') and obj.trip else None)
    request_type = "Trip" if isinstance(obj, Trip) else ("Advance" if isinstance(obj, TravelAdvance) else ("Expense Claim" if isinstance(obj, TravelClaim) else "Bulk Activity Log"))

    # Check if the requester needs to follow the HR approval flow (only if they report to the COO, are SPH, or have no reporting manager)
    reports_to_coo = False
    if requester and not _is_hr(requester):
        user_dept = str(requester.department or '').strip().lower()
        user_desig = str(requester.designation or '').strip().lower()
        user_pos = str(requester.active_position_id or '').strip()
        is_sph = ('sph' in user_dept or 'sph' in user_desig or user_pos == '2')
        rm = requester.reporting_manager
        if is_sph or not rm:
            reports_to_coo = True
        else:
            rm_pos = rm.get_current_position()
            rm_pos_name = rm_pos.get('name') if rm_pos else None
            reports_to_coo = _is_coo_position(rm_pos_name, rm.designation, employee_id=rm.employee_id)
            
        if reports_to_coo and rm:
            from travel.models import COOProjectSetting
            proj_code = 'General'
            if isinstance(obj, Trip):
                proj_code = obj.project_code or 'General'
            elif hasattr(obj, 'trip') and obj.trip:
                proj_code = obj.trip.project_code or 'General'
            elif isinstance(obj, BulkActivityBatch) and obj.trip:
                proj_code = obj.trip.project_code or 'General'
                
            rm_pos_id = rm.active_position_id
            if rm_pos_id and COOProjectSetting.objects.filter(project_code=proj_code, coo_position_id=str(rm_pos_id), enable_coo_approval=True).exists():
                reports_to_coo = False

    # Get project code
    project_code = 'General'
    if isinstance(obj, Trip):
        project_code = obj.project_code or 'General'
    elif hasattr(obj, 'trip') and obj.trip:
        project_code = obj.trip.project_code or 'General'
    elif isinstance(obj, BulkActivityBatch) and obj.trip:
        project_code = obj.trip.project_code or 'General'

    # Fallback to requester's project_code if general/empty
    if (not project_code or project_code in ['General', 'N/A']) and requester:
        if hasattr(requester, 'project_code') and requester.project_code and requester.project_code != 'N/A':
            project_code = requester.project_code

    hr_positions = HRPositionConfig.objects.filter(is_active=True, project_code=project_code).order_by('sequence_order')
    if not hr_positions.exists():
        hr_positions = HRPositionConfig.objects.filter(is_active=True, project_code='General').order_by('sequence_order')

    # Resolve settings
    from travel.models import HRWorkflowSetting
    setting = HRWorkflowSetting.objects.filter(project_code=project_code).first()
    if not setting and project_code != 'General':
        setting = HRWorkflowSetting.objects.filter(project_code='General').first()
    is_parallel_flow = setting.is_parallel if setting else False
    enable_two_level_flow = setting.enable_two_level_flow if setting else False

    target_hr_positions = hr_positions
    if is_parallel_flow and enable_two_level_flow:
        am_positions = hr_positions.filter(hr_level_type='assistant_manager')
        if am_positions.exists():
            target_hr_positions = am_positions
        else:
            target_hr_positions = hr_positions.filter(hr_level_type='manager')

    if isinstance(obj, (Trip, BulkActivityBatch)):
        if isinstance(obj, BulkActivityBatch):
            hr_requires_approval = hr_positions.filter(bulk_approval='APPROVAL').exists()
        else:
            hr_requires_approval = hr_positions.filter(trips_approval='APPROVAL').exists()
    else:
        hr_requires_approval = hr_positions.filter(claims_approval='APPROVAL').exists()

    # --- A. INSTANT SYNCHRONOUS STATE FINALIZATION ---
    response_msg = "Management approval completed."

    if isinstance(obj, (Trip, BulkActivityBatch)):
        if hr_requires_approval:
            # Find the first HR position in sequence requiring approval
            first_approval_pos = None
            for hr_pos in target_hr_positions:
                is_appr_req = False
                if isinstance(obj, Trip) and hr_pos.trips_approval == 'APPROVAL':
                    is_appr_req = True
                elif isinstance(obj, BulkActivityBatch) and hr_pos.bulk_approval == 'APPROVAL':
                    is_appr_req = True
                
                if is_appr_req:
                    first_approval_pos = hr_pos
                    break

            # Keep status as Pending and wait for HR approval.
            obj.status = 'Pending'
            if first_approval_pos:
                obj.approver_position = first_approval_pos.position_id
                pos_users = get_users_by_position(first_approval_pos.position_id)
                obj.current_approver = pos_users[0] if pos_users else None
            else:
                obj.current_approver = None
                obj.approver_position = None
            obj.save()

            if isinstance(obj, Trip) and hasattr(obj, 'activity_batches'):
                obj.activity_batches.exclude(status='Rejected').update(
                    status='Pending', current_approver=obj.current_approver, approver_position=obj.approver_position
                )
            elif isinstance(obj, BulkActivityBatch) and obj.trip:
                obj.trip.status = 'Pending'
                obj.trip.current_approver = obj.current_approver
                obj.trip.approver_position = obj.approver_position
                obj.trip.save()

            response_msg = "Management approval completed. Request forwarded to HR for approval."
            msg_title = f"{request_type} Bypassed to HR" if reports_to_coo else f"{request_type} Sent to HR"
            msg_body = f"Your {request_type} has bypassed COO and is pending HR approval." if reports_to_coo else f"Your {request_type} has cleared management approval and is pending HR approval."
            Notification.objects.create(
                user=requester, title=msg_title,
                message=msg_body,
                type='info'
            )
        else:
            # TRIPS and Bulk Activity Batches stop here. They are finalized as Approved
            obj.status = 'Approved'
            obj.current_approver = None
            obj.approver_position = None
            obj.save()
            
            if isinstance(obj, Trip) and hasattr(obj, 'activity_batches'):
                obj.activity_batches.exclude(status='Rejected').update(
                    status='Approved', current_approver=None, approver_position=None
                )
                _generate_expenses_from_batches(obj)
            elif isinstance(obj, BulkActivityBatch) and obj.trip:
                obj.trip.status = 'Approved'
                obj.trip.current_approver = None
                obj.trip.approver_position = None
                obj.trip.save()
                _generate_expenses_from_batches(obj.trip)
                
            response_msg = "Management approval completed. Request finalized and HR intimated."
            Notification.objects.create(
                user=requester, title=f"{request_type} Finalized",
                message=f"Your {request_type} has been approved by management and finalized. HR has been notified.",
                type='success'
            )
    
    else:
        # Claims & Advances
        if hr_requires_approval:
            # Find the first HR position in sequence requiring approval
            first_approval_pos = None
            for hr_pos in target_hr_positions:
                if hr_pos.claims_approval == 'APPROVAL':
                    first_approval_pos = hr_pos
                    break

            obj.status = 'PENDING_HR'
            if first_approval_pos:
                obj.approver_position = first_approval_pos.position_id
                pos_users = get_users_by_position(first_approval_pos.position_id)
                obj.current_approver = pos_users[0] if pos_users else None
            else:
                obj.current_approver = None
                obj.approver_position = None
            obj.save()
            response_msg = "Management approval completed. Request forwarded to HR for approval."
            Notification.objects.create(
                user=requester, title=f"{request_type} Sent to HR",
                message=f"Your {request_type} has cleared management approval and is pending HR approval.",
                type='info'
            )
        else:
            # Claims & Advances move to Finance Workflow
            dispatch_res = trigger_finance_workflow(obj, user)
            response_msg = f"Management approval completed. {dispatch_res['message']}"

    # --- B. THREADED ASYNCHRONOUS EXTERNAL DISPATCH ---
    def _async_dispatch_worker(obj_id, obj_class, requester_name):
        try:
            import time
            time.sleep(0.5)
            
            from django.db import connection
            connection.close()
            
            active_obj = obj_class.objects.get(pk=obj_id)
            
            # Check workflow settings to determine if parallel or sequential
            from travel.models import HRWorkflowSetting
            setting = HRWorkflowSetting.objects.filter(project_code=project_code).first()
            if not setting and project_code != 'General':
                setting = HRWorkflowSetting.objects.filter(project_code='General').first()
            is_parallel_flow = setting.is_parallel if setting else False
            enable_two_level_flow = setting.enable_two_level_flow if setting else False

            target_hr_positions = hr_positions
            if is_parallel_flow and enable_two_level_flow:
                am_positions = hr_positions.filter(hr_level_type='assistant_manager')
                if am_positions.exists():
                    target_hr_positions = am_positions
                else:
                    target_hr_positions = hr_positions.filter(hr_level_type='manager')

            # Find the first approval position
            first_approval_pos = None
            for hr_pos in target_hr_positions:
                is_appr_required = False
                if isinstance(active_obj, Trip) and hr_pos.trips_approval == 'APPROVAL':
                    is_appr_required = True
                elif isinstance(active_obj, BulkActivityBatch) and hr_pos.bulk_approval == 'APPROVAL':
                    is_appr_required = True
                elif isinstance(active_obj, (TravelClaim, TravelAdvance)) and hr_pos.claims_approval == 'APPROVAL':
                    is_appr_required = True
                
                if is_appr_required:
                    first_approval_pos = hr_pos
                    break

            for hr_pos in target_hr_positions:
                is_appr = False
                if isinstance(active_obj, Trip) and hr_pos.trips_approval == 'APPROVAL':
                    is_appr = True
                elif isinstance(active_obj, BulkActivityBatch) and hr_pos.bulk_approval == 'APPROVAL':
                    is_appr = True
                elif isinstance(active_obj, (TravelClaim, TravelAdvance)) and hr_pos.claims_approval == 'APPROVAL':
                    is_appr = True

                is_read_only = False
                if isinstance(active_obj, Trip) and hr_pos.trips_approval == 'MARK_READ':
                    is_read_only = True
                elif isinstance(active_obj, BulkActivityBatch) and hr_pos.bulk_approval == 'MARK_READ':
                    is_read_only = True
                elif isinstance(active_obj, (TravelClaim, TravelAdvance)) and hr_pos.claims_approval == 'MARK_READ':
                    is_read_only = True

                # Under parallel flow: all approval/read_only steps are dispatched in parallel.
                # Under sequential flow: only the current active approval step OR read-only step is dispatched.
                if is_parallel_flow:
                    should_dispatch = is_appr or is_read_only
                else:
                    is_active_appr = first_approval_pos and hr_pos.position_id == first_approval_pos.position_id
                    should_dispatch = is_active_appr or is_read_only
                    # In sequential mode, is_appr flag should only be True for the active approval step
                    is_appr = is_active_appr

                if should_dispatch:
                    hr_users = get_users_by_position(hr_pos.position_id)
                    for hru in hr_users:
                        intimation_filter = {'hr_user': hru, 'hr_position': hr_pos.position_id}
                        if isinstance(active_obj, Trip): intimation_filter['trip'] = active_obj
                        elif isinstance(active_obj, TravelClaim): intimation_filter['claim'] = active_obj
                        elif isinstance(active_obj, TravelAdvance): intimation_filter['advance'] = active_obj
                        elif isinstance(active_obj, BulkActivityBatch): intimation_filter['trip'] = active_obj.trip
                        
                        if not HRIntimation.objects.filter(**intimation_filter).exists():
                            HRIntimation.objects.create(is_approval=is_appr, **intimation_filter)
                            
                            msg_title = "Pending HR Approval" if is_appr else f"New Intimation: {request_type}"
                            msg_body = f"{requester_name}'s {request_type} requires your approval." if is_appr else f"{requester_name}'s {request_type} requires your acknowledgement."
                            
                            Notification.objects.create(
                                user=hru, target_position=hr_pos.position_id,
                                title=msg_title,
                                message=msg_body,
                                type='info'
                            )
        except Exception as e:
            import logging
            logging.error(f"Background parallel dispatch failed: {e}", exc_info=True)

    # Trigger execution in safe daemon background thread immediately
    import threading
    t = threading.Thread(
        target=_async_dispatch_worker, 
        args=(obj.pk, obj.__class__, requester.name if requester else 'Unknown')
    )
    t.daemon = True
    t.start()

    return {"message": response_msg}


def finalize_trip_approval(trip):
    """
    Finalizes trip approval, updates associated batches, and generates expenses.
    """
    from django.utils import timezone
    trip.status = 'Approved'
    trip.current_approver = None
    trip.approver_position = None
    trip.save()
    
    # Finalize associated batches
    for batch in trip.activity_batches.exclude(status='Rejected'):
        batch.status = 'Approved'
        batch.current_approver = None
        
        # Update row statuses in JSON
        updated_rows = []
        for row in (batch.data_json or []):
            if row.get('_status') != 'Rejected':
                row['_status'] = 'Approved'
            updated_rows.append(row)
        batch.data_json = updated_rows
        batch.save()
        
    _generate_expenses_from_batches(trip)


def _get_finance_steps_and_settings(obj):
    from travel.models import FinanceWorkflowStep, FinanceWorkflowSetting
    requester = obj.user if hasattr(obj, 'user') else (obj.trip.user if hasattr(obj, 'trip') and obj.trip else None)
    
    project_code = 'General'
    if isinstance(obj, Trip):
        project_code = obj.project_code or 'General'
    elif hasattr(obj, 'trip') and obj.trip:
        project_code = obj.trip.project_code or 'General'
    elif isinstance(obj, BulkActivityBatch) and obj.trip:
        project_code = obj.trip.project_code or 'General'

    # Fallback to requester's project_code if general/empty
    if (not project_code or project_code in ['General', 'N/A']) and requester:
        if hasattr(requester, 'project_code') and requester.project_code and requester.project_code != 'N/A':
            project_code = requester.project_code

    fin_steps = FinanceWorkflowStep.objects.filter(is_active=True, project_code=project_code).order_by('sequence_order')
    if not fin_steps.exists():
        fin_steps = FinanceWorkflowStep.objects.filter(is_active=True, project_code='General').order_by('sequence_order')

    # Safety guard: exclude orphaned steps with no assignee (no position_id AND no user).
    # Such steps can never dispatch to anyone, so they must never block workflow routing.
    fin_steps = fin_steps.exclude(
        models.Q(position_id__isnull=True) | models.Q(position_id=''),
        user__isnull=True
    )

    active_trip = obj if isinstance(obj, Trip) else (obj.trip if hasattr(obj, 'trip') and obj.trip else None)
    if active_trip:
        is_local = getattr(active_trip, 'consider_as_local', True)
        if isinstance(obj, (Trip, BulkActivityBatch)):
            allowed_types = ['BOTH', 'TRIP'] if is_local else ['BOTH', 'TRAVEL']
        else:
            allowed_types = ['BOTH', 'TRIP', 'NONE'] if is_local else ['BOTH', 'TRAVEL', 'NONE']
        fin_steps = fin_steps.filter(trip_type__in=allowed_types)
    else:
        if isinstance(obj, (Trip, BulkActivityBatch)):
            fin_steps = fin_steps.exclude(trip_type='NONE')

    setting = FinanceWorkflowSetting.objects.filter(project_code=project_code).first()
    if not setting and project_code != 'General':
        setting = FinanceWorkflowSetting.objects.filter(project_code='General').first()
        
    is_parallel_flow = setting.is_parallel if setting else False
    enable_two_level_flow = setting.enable_two_level_flow if setting else False

    return fin_steps, is_parallel_flow, enable_two_level_flow, project_code


def trigger_finance_workflow(obj, user=None):
    """
    Dispatches Trips, Claims, or Advances to Finance configuration steps.
    Implements project-wise steps (falling back to General), and support for:
    - parallel, sequential, and parallel 2-way workflows
    - assistant_manager / manager level separation (2-way)
    - finance hub exception
    """
    from .models import FinanceWorkflowStep, FinanceIntimation
    from notifications.models import Notification

    requester = obj.user if hasattr(obj, 'user') else (obj.trip.user if hasattr(obj, 'trip') and obj.trip else None)
    request_type = "Trip" if isinstance(obj, Trip) else ("Advance" if isinstance(obj, TravelAdvance) else ("Expense Claim" if isinstance(obj, TravelClaim) else "Bulk Activity Log"))

    # 1. Resolve steps and settings
    fin_steps, is_parallel_flow, enable_two_level_flow, project_code = _get_finance_steps_and_settings(obj)

    if not fin_steps.exists():
        # No matching steps - finalize/approve immediately
        if isinstance(obj, (Trip, BulkActivityBatch)):
            finalize_trip_approval(obj if isinstance(obj, Trip) else obj.trip)
        else:
            obj.status = 'Approved'
            obj.current_approver = None
            obj.approver_position = None
            obj.save()
        return {"status": "Approved", "message": "No matching Finance steps. Auto-approved."}

    # 2. Filter target steps based on workflow settings
    target_steps = fin_steps
    if is_parallel_flow and enable_two_level_flow:
        am_steps = fin_steps.filter(
            models.Q(finance_level_type='assistant_manager') | 
            models.Q(visibility_type__in=['FINANCE_HUB', 'BOTH'])
        )
        if am_steps.exists():
            target_steps = am_steps
        else:
            target_steps = fin_steps.filter(
                models.Q(finance_level_type='manager') | 
                models.Q(visibility_type__in=['FINANCE_HUB', 'BOTH'])
            )

    # 3. Find the first approval position (used for sequential flow or setting the main current approver)
    first_approval_step = None
    for step in target_steps:
        is_appr_req = False
        if isinstance(obj, (Trip, BulkActivityBatch)):
            is_appr_req = (step.trip_control == 'APPROVAL')
        else:
            is_appr_req = True
        
        if is_appr_req:
            first_approval_step = step
            break

    # 4. Set the initial object status and approver fields
    finance_requires_approval = False
    if not is_parallel_flow:
        # Sequential Flow: find the first approval (blocking) step
        first_approval_step = None
        is_claim_or_advance = isinstance(obj, (TravelClaim, TravelAdvance))
        for step in target_steps:
            is_appr_step = (step.trip_control == 'APPROVAL') if not is_claim_or_advance else True
            if is_appr_step:
                first_approval_step = step
                break
        
        if first_approval_step:
            finance_requires_approval = True
            if isinstance(obj, (Trip, BulkActivityBatch)):
                active_trip = obj if isinstance(obj, Trip) else obj.trip
                active_trip.status = 'PENDING_FINANCE'
                active_trip.approver_position = first_approval_step.position_id
                pos_users = get_users_by_position(first_approval_step.position_id) if first_approval_step.position_id else []
                active_trip.current_approver = pos_users[0] if pos_users else first_approval_step.user
                active_trip.save()

                if isinstance(obj, Trip):
                    for batch in obj.activity_batches.exclude(status='Rejected'):
                        batch.status = 'PENDING_FINANCE'
                        batch.current_approver = active_trip.current_approver
                        batch.approver_position = active_trip.approver_position
                        batch.save()
            else:
                obj.approver_position = first_approval_step.position_id
                obj.status = 'PENDING_FINAL_RELEASE' if first_approval_step.visibility_type == 'FINANCE_HUB' else 'PENDING_EXECUTIVE'
                pos_users = get_users_by_position(first_approval_step.position_id) if first_approval_step.position_id else []
                obj.current_approver = pos_users[0] if pos_users else first_approval_step.user
                obj.save()
        else:
            if isinstance(obj, (Trip, BulkActivityBatch)):
                finalize_trip_approval(obj if isinstance(obj, Trip) else obj.trip)
            else:
                obj.status = 'Approved'
                obj.current_approver = None
                obj.approver_position = None
                obj.save()
    else:
        # Parallel Flow: check if any step requires approval
        finance_requires_approval = False
        if isinstance(obj, (Trip, BulkActivityBatch)):
            active_trip = obj if isinstance(obj, Trip) else obj.trip
            for step in target_steps:
                if step.trip_control == 'APPROVAL':
                    finance_requires_approval = True
                    break
            
            if finance_requires_approval:
                active_trip.status = 'PENDING_FINANCE'
                if first_approval_step:
                    active_trip.approver_position = first_approval_step.position_id
                    pos_users = get_users_by_position(first_approval_step.position_id) if first_approval_step.position_id else []
                    active_trip.current_approver = pos_users[0] if pos_users else first_approval_step.user
                else:
                    active_trip.current_approver = None
                    active_trip.approver_position = None
                active_trip.save()

                if isinstance(obj, Trip):
                    for batch in obj.activity_batches.exclude(status='Rejected'):
                        batch.status = 'PENDING_FINANCE'
                        batch.current_approver = active_trip.current_approver
                        batch.approver_position = active_trip.approver_position
                        batch.save()
            else:
                # Only intimations (Mark as Read) - finalize immediately
                finalize_trip_approval(active_trip)
        else:
            # Claims & Advances — check if any step actually requires approval
            if is_parallel_flow:
                # Find if there are any active inbox steps
                inbox_steps = target_steps.filter(visibility_type__in=['INBOX', 'BOTH'])
                if inbox_steps.exists():
                    finance_requires_approval = True
                    first_inbox_step = inbox_steps.first()
                    obj.approver_position = first_inbox_step.position_id
                    obj.status = 'PENDING_EXECUTIVE'
                    pos_users = get_users_by_position(first_inbox_step.position_id) if first_inbox_step.position_id else []
                    obj.current_approver = pos_users[0] if pos_users else first_inbox_step.user
                else:
                    # Hub-only workflow
                    finance_requires_approval = any(s.trip_control == 'APPROVAL' for s in target_steps)
                    if finance_requires_approval and first_approval_step:
                        obj.approver_position = first_approval_step.position_id
                        obj.status = 'PENDING_FINAL_RELEASE'
                        pos_users = get_users_by_position(first_approval_step.position_id) if first_approval_step.position_id else []
                        obj.current_approver = pos_users[0] if pos_users else first_approval_step.user
                    else:
                        obj.current_approver = None
                        obj.approver_position = None
                        obj.status = 'Approved'
            else:
                # Sequential flow
                finance_requires_approval = any(s.trip_control == 'APPROVAL' for s in target_steps)
                if finance_requires_approval and first_approval_step:
                    obj.approver_position = first_approval_step.position_id
                    obj.status = 'PENDING_FINAL_RELEASE' if first_approval_step.visibility_type == 'FINANCE_HUB' else 'PENDING_EXECUTIVE'
                    pos_users = get_users_by_position(first_approval_step.position_id) if first_approval_step.position_id else []
                    obj.current_approver = pos_users[0] if pos_users else first_approval_step.user
                else:
                    # All steps are MARK_READ (or no steps) — auto-approve immediately
                    obj.current_approver = None
                    obj.approver_position = None
                    obj.status = 'Approved'
            obj.save()

    # 5. Background Dispatch worker or inline dispatch
    def _async_finance_dispatch_worker(obj_id, obj_class, requester_name):
        try:
            import time
            time.sleep(0.5)
            from django.db import connection
            connection.close()

            active_obj = obj_class.objects.get(pk=obj_id)
            active_trip = active_obj if isinstance(active_obj, Trip) else getattr(active_obj, 'trip', None)

            # Re-fetch steps and settings to be safe
            fin_steps, is_parallel, enable_two_level, _ = _get_finance_steps_and_settings(active_obj)
            
            target_steps = fin_steps
            if is_parallel and enable_two_level:
                am_steps = fin_steps.filter(
                    models.Q(finance_level_type='assistant_manager') | 
                    models.Q(visibility_type__in=['FINANCE_HUB', 'BOTH'])
                )
                if am_steps.exists():
                    target_steps = am_steps
                else:
                    target_steps = fin_steps.filter(
                        models.Q(finance_level_type='manager') | 
                        models.Q(visibility_type__in=['FINANCE_HUB', 'BOTH'])
                    )

            # Find the first step in sequential flow
            first_step = target_steps.first()

            is_claim_or_advance = isinstance(active_obj, (TravelClaim, TravelAdvance))
            created_count = 0
            for step in target_steps:
                # trip_control governs whether this step needs approval or is just a read notification
                if is_claim_or_advance:
                    is_appr = True
                    is_read_only = False
                else:
                    is_appr = (step.trip_control == 'APPROVAL')
                    is_read_only = (step.trip_control == 'MARK_READ')
                
                is_hub_step = step.visibility_type in ['FINANCE_HUB', 'BOTH']

                if is_parallel:
                    if enable_two_level:
                        # 2-Level: only dispatch assistant_manager INBOX steps initially (Finance Hub comes after managers)
                        should_dispatch = (not is_hub_step) and (is_appr or is_read_only)
                    else:
                        # Parallel only: dispatch INBOX steps only (Finance Hub dispatched after any inbox approval)
                        should_dispatch = (not is_hub_step) and (is_appr or is_read_only)
                else:
                    # Sequential: dispatch all initial MARK_READ steps and the first APPROVAL step
                    first_approval_step = None
                    for s in target_steps:
                        is_appr_s = (s.trip_control == 'APPROVAL') if not is_claim_or_advance else True
                        if is_appr_s:
                            first_approval_step = s
                            break
                    
                    if first_approval_step:
                        should_dispatch = (step.sequence_order <= first_approval_step.sequence_order)
                        is_appr = (step.id == first_approval_step.id)
                    else:
                        should_dispatch = True
                        is_appr = False

                if should_dispatch:
                    target_users = []
                    if step.position_id:
                        target_users = get_users_by_position(step.position_id)
                    elif step.user:
                        target_users = [step.user]

                    for f_user in target_users:
                        intimation_filter = {'finance_user': f_user, 'finance_position': step.position_id or ''}
                        if isinstance(active_obj, Trip):
                            intimation_filter['trip'] = active_obj
                        elif isinstance(active_obj, TravelClaim):
                            intimation_filter['claim'] = active_obj
                        elif isinstance(active_obj, TravelAdvance):
                            intimation_filter['advance'] = active_obj
                        elif isinstance(active_obj, BulkActivityBatch):
                            intimation_filter['trip'] = active_obj.trip

                        if not FinanceIntimation.objects.filter(**intimation_filter).exists():
                            FinanceIntimation.objects.create(is_approval=is_appr, is_read=False, **intimation_filter)
                            created_count += 1

                            # Send Notification
                            Notification.objects.create(
                                user=f_user,
                                target_position=step.position_id,
                                title=f"Pending Finance {'Approval' if is_appr else 'Intimation'}: {request_type}",
                                message=f"{requester_name}'s {request_type} requires your review.",
                                type='info'
                            )
        except Exception as e:
            import logging
            logging.error(f"Background finance parallel dispatch failed: {e}", exc_info=True)

    import threading
    t = threading.Thread(
        target=_async_finance_dispatch_worker,
        args=(obj.pk, obj.__class__, requester.name if requester else 'Unknown')
    )
    t.daemon = True
    t.start()

    response_msg = "Forwarded to Finance Workflow."
    if not finance_requires_approval:
        response_msg = "No approval required. Finalized & sent to Finance as intimations."

    return {"status": obj.status, "message": response_msg}


def _get_user_all_position_ids(user):
    """Returns a set of position identifiers (numeric ID and string code) for the user's positions."""
    pos_ids = set()
    if not user: return pos_ids
    if user.active_position_id:
        pos_ids.add(str(user.active_position_id))
        
    try:
        data = user._get_api_data()
        if data:
            # Main position object
            main_pos = data.get('position')
            if main_pos:
                p_id = str(main_pos.get('id') or '')
                p_code = str(main_pos.get('code') or '').strip()
                p_name = str(main_pos.get('name') or '').strip()
                if p_id: pos_ids.add(p_id)
                if p_code: pos_ids.add(p_code)
                if p_name: pos_ids.add(p_name)
                
            # Details of all positions
            for p in data.get('positions_details', []):
                p_id = str(p.get('id') or '')
                p_code = str(p.get('code') or '').strip()
                p_name = str(p.get('name') or '').strip()
                if p_id: pos_ids.add(p_id)
                if p_code: pos_ids.add(p_code)
                if p_name: pos_ids.add(p_name)
    except Exception:
        pass
            
    # Fallback to query all mapped positions from global cache and profile
    try:
        from django.core.cache import cache
        user_pos_map = cache.get('user_position_identifiers')
        if user_pos_map and user.employee_id in user_pos_map:
            for active_id, val_list in user_pos_map[user.employee_id].items():
                for val in val_list:
                    if val:
                        pos_ids.add(str(val).strip())
    except Exception:
        pass
        
    try:
        available = user.get_available_positions()
        for pos in available:
            if pos.get('id'):
                pos_ids.add(str(pos['id']).strip())
            if pos.get('code'):
                pos_ids.add(str(pos['code']).strip())
            if pos.get('name'):
                pos_ids.add(str(pos['name']).strip())
    except Exception:
        pass
        
    return pos_ids

def _is_finance_head(user):
    """Checks if a user is the final level finance head based on position config."""
    if not user: return False
    from .models import FinanceWorkflowStep

    # Check active position IDs/codes against FinanceWorkflowStep
    all_pos_ids = _get_user_all_position_ids(user)
    for pos_id in all_pos_ids:
        step = FinanceWorkflowStep.objects.filter(position_id=pos_id, is_active=True).first()
        if step:
            return step.visibility_type in ['FINANCE_HUB', 'BOTH']

    step = FinanceWorkflowStep.objects.filter(user=user, is_active=True).first()
    if step:
        return step.visibility_type in ['FINANCE_HUB', 'BOTH']
    
    eid = user.employee_id.lower()
    dept = (user.department or "").lower()
    desig = (user.designation or "").lower()
    user_role = (user.role.name.lower() if user.role else '')
    if 'fh-' in eid or '-fh' in eid or 'cfo' in eid:
        return True
    return 'head' in dept and 'finance' in dept or 'head' in desig and 'finance' in desig or 'cfo' in user_role

def _is_finance_executive(user):
    """Checks if a user is a Finance Executive based on workflow config."""
    if not user: return False
    from .models import FinanceWorkflowStep

    # Check active position IDs/codes against FinanceWorkflowStep
    all_pos_ids = _get_user_all_position_ids(user)
    for pos_id in all_pos_ids:
        if FinanceWorkflowStep.objects.filter(position_id=pos_id, is_active=True).exists():
            return True

    if FinanceWorkflowStep.objects.filter(user=user, is_active=True).exists():
        return True

    if _is_finance_head(user): return False
    eid = user.employee_id.lower()
    user_role = (user.role.name.lower() if user.role else '')
    dept = (user.department or "").lower()
    desig = (user.designation or "").lower()
    if 'fe-' in eid or '-fe' in eid or 'fin-' in eid:
        return True
    return 'finance' in user_role or 'finance' in dept or 'finance' in desig

def _get_matching_hr_configs(user, project_code=None):
    """
    Returns a queryset of active HRPositionConfig records that match
    the user's positions (by ID, code, or name).
    """
    from .models import HRPositionConfig
    if not user:
        return HRPositionConfig.objects.none()
        
    pos_ids = _get_user_all_position_ids(user)
    if not pos_ids:
        return HRPositionConfig.objects.none()
        
    # Start with active configurations
    configs = HRPositionConfig.objects.filter(is_active=True)
    if project_code:
        configs = configs.filter(project_code=project_code)
        
    # Filter them in memory / python
    matched_ids = []
    for config in configs:
        # 1. Direct match
        if config.position_id in pos_ids:
            matched_ids.append(config.id)
            continue
            
        # 2. Match by prefix of position_name (e.g. 'HR ONBOARDING' in 'HR ONBOARDING (HRONBOARD@-AP-104-MMUS)')
        config_name_clean = str(config.position_name).strip()
        if ' (' in config_name_clean:
            prefix = config_name_clean.split(' (')[0].strip()
        else:
            prefix = config_name_clean
            
        for user_pos_id in pos_ids:
            if (user_pos_id.lower() == prefix.lower() or 
                user_pos_id.lower() == config_name_clean.lower() or 
                user_pos_id.lower() == config.position_id.lower()):
                matched_ids.append(config.id)
                break
                
    return HRPositionConfig.objects.filter(id__in=matched_ids)

def _is_hr(user):
    """Checks if user belongs to any active configured HR Position."""
    if not user: return False
    return _get_matching_hr_configs(user).exists()

def _get_finance_users():
    """Returns list of users matching active Finance Workflow step positions."""
    from .models import FinanceWorkflowStep
    steps = FinanceWorkflowStep.objects.filter(is_active=True)
    users = []
    for step in steps:
        if step.position_id:
            for u in get_users_by_position(step.position_id):
                if u not in users: users.append(u)
        elif step.user:
            if step.user not in users: users.append(step.user)
    return users

def _get_finance_step_for_user(user, project_code=None):
    """
    Reverse-lookup: find the FinanceWorkflowStep that this user currently holds.
    Uses get_users_by_position (API-aware) for each step so position code vs numeric ID mismatches are handled.
    """
    if not user: return None
    from .models import FinanceWorkflowStep
    
    fin_steps = FinanceWorkflowStep.objects.filter(is_active=True)
    if project_code:
        steps = fin_steps.filter(project_code=project_code)
        if not steps.exists():
            steps = fin_steps.filter(project_code='General')
    else:
        steps = fin_steps

    all_pos_ids = _get_user_all_position_ids(user)
    for pos_id in all_pos_ids:
        step = steps.filter(position_id=pos_id).first()
        if step:
            return step
            
    # Check step properties directly in user's position identifiers (highly resilient fallback)
    for step in steps.order_by('sequence_order'):
        if step.position_id and str(step.position_id) in all_pos_ids:
            return step
        if step.position_name and str(step.position_name) in all_pos_ids:
            return step
        if step.user == user:
            return step
            
    # Slow path: iterate each step and check via API (handles code vs numeric ID mismatch)
    for step in steps.order_by('sequence_order'):
        if step.position_id:
            if user in get_users_by_position(step.position_id):
                return step
    return None

def _get_hr_users():
    """Returns list of all users currently matching active HR position configurations."""
    from .models import HRPositionConfig
    positions = HRPositionConfig.objects.filter(is_active=True)
    users = []
    for pos in positions:
        for u in get_users_by_position(pos.position_id):
            if u not in users: users.append(u)
    return users


def get_finance_head(user, exclude_user=None, project_code=None):
    """Finds the final approver (HUB visibility)."""
    if not project_code and user:
        project_code = getattr(user, 'project_code', 'General')
        
    steps = FinanceWorkflowStep.objects.filter(is_active=True, project_code=project_code)
    if not steps.exists():
        steps = FinanceWorkflowStep.objects.filter(is_active=True, project_code='General')
        
    hub_step = steps.filter(visibility_type='FINANCE_HUB').first()
    if hub_step:
        if hub_step.position_id:
            pos_users = get_users_by_position(hub_step.position_id)
            if pos_users: return pos_users[0]
        return hub_step.user
        
    # Fallback
    all_finance = _get_finance_users()
    heads = [u for u in all_finance if _is_finance_head(u)]
    if exclude_user:
        heads = [u for u in heads if u.id != exclude_user.id]
    
    if heads:
        return heads[0]
    return all_finance[0] if all_finance else None

def get_finance_executive(user, exclude_user=None, project_code=None):
    """Finds the first-level finance executive (INBOX visibility)."""
    if not project_code and user:
        project_code = getattr(user, 'project_code', 'General')
        
    steps = FinanceWorkflowStep.objects.filter(is_active=True, project_code=project_code)
    if not steps.exists():
        steps = FinanceWorkflowStep.objects.filter(is_active=True, project_code='General')
        
    first_step = steps.filter(visibility_type='INBOX').order_by('sequence_order').first()
    if first_step:
        if first_step.position_id:
            pos_users = get_users_by_position(first_step.position_id)
            if pos_users: return pos_users[0]
        return first_step.user

    # Fallback
    all_finance = _get_finance_users()
    execs = [u for u in all_finance if _is_finance_executive(u)]
    if exclude_user:
         execs = [u for u in execs if u.id != exclude_user.id]
         
    if execs:
        return execs[0]
    return all_finance[0] if all_finance else None

def get_hr_head(user):
    """Finds an HR approver (Head of HR)."""
    all_hr = _get_hr_users()
    # Try local HR first
    local_hr = [u for u in all_hr if u.base_location == user.base_location]
    if local_hr:
        return local_hr[0]
    
    return all_hr[0] if all_hr else None

def notify_hr(title, message):
    """Notify all users with HR role."""
    hr_users = User.objects.filter(role__name__icontains='hr')
    for hr in hr_users:
        Notification.objects.create(
            user=hr,
            title=title,
            message=message,
            type='info'
        )


from .utils import _get_hr_users, get_hr_head, build_approval_chain, _is_coo_position

def update_trip_lifecycle(trip, title, description):
    """Helper to append events to the Trip's JSON lifecycle field."""
    if not trip:
        return
    event = {
        "title": title,
        "status": "completed",
        "date": timezone.now().strftime("%b %d, %Y"),
        "description": description
    }
    events = trip.lifecycle_events
    if not isinstance(events, list):
        events = []
    
    # Avoid duplicate events with same title and description
    if not any(e.get('title') == title and e.get('description') == description for e in events):
        events.append(event)
        trip.lifecycle_events = events
        trip.save(update_fields=['lifecycle_events'])

def _generate_expenses_from_batches(trip):
    """
    Helper to convert rows from BulkActivityBatch into individual Expense records.
    Implements strict deduplication to prevent duplicate entries from multiple approval triggers.
    """
    from .models import BulkActivityBatch
    from travel_masters.models import FuelRateMaster, Cadre, EligibilityRule
    import json
    from datetime import datetime, timedelta
    
    # 1. Pre-fetch all existing expenses for this trip to avoid repeated DB queries
    existing_expenses = list(Expense.objects.filter(trip=trip))
    
    # Track semantic activities to prevent manual vs bulk duplicates
    # Key: (date_str, odo_start, odo_end, origin_norm, dest_norm)
    semantic_map = {}
    # Track specific batch rows already processed
    # Key: (batch_id, row_index)
    processed_rows = set()
    
    for e in existing_expenses:
        try:
            d = json.loads(e.description)
            b_id = d.get('batch_id')
            r_idx = d.get('row_index')
            if b_id is not None and r_idx is not None:
                processed_rows.add((int(b_id), int(r_idx)))
            
            # Record semantic footprint
            date_str = str(e.date)
            o_s = float(e.odo_start or 0)
            o_e = float(e.odo_end or 0)
            orig = str(d.get('origin', '')).strip().upper()
            dest = str(d.get('destination', '')).strip().upper()
            semantic_map[(date_str, o_s, o_e, orig, dest)] = e.id
        except:
            continue

    expenses_to_delete = []

    # 1. Cleanup: Identify any expenses associated with batches that are NOT in an authorized state
    unauthorized_batches = trip.activity_batches.exclude(status__in=['Approved', 'Manager Approved', 'HR Approved', 'Resolved', 'Resubmitted', 'Submitted', 'Forwarded', 'Under Process'])
    for batch in unauthorized_batches:
        for e in existing_expenses:
            try:
                d = json.loads(e.description)
                if int(d.get('batch_id', -1)) == batch.id:
                    expenses_to_delete.append(e)
            except: continue

    # Perform first batch deletion
    if expenses_to_delete:
        delete_ids = [e.id for e in expenses_to_delete]
        Expense.objects.filter(id__in=delete_ids).delete()
        existing_expenses = [x for x in existing_expenses if x.id not in delete_ids]
        expenses_to_delete = []

    # 2. Process authorized batches
    batches = trip.activity_batches.filter(status__in=['Approved', 'Manager Approved', 'HR Approved', 'Resolved', 'Resubmitted', 'Submitted', 'Forwarded', 'Under Process'])
    
    processed_da_dates = set()
    for e in existing_expenses:
        try:
            d = json.loads(e.description)
            b_id = d.get('batch_id')
            if b_id is None or int(b_id) not in [b.id for b in batches]:
                if float(d.get('daily_allowance', 0)) > 0:
                    processed_da_dates.add(str(e.date)[:10])
        except:
            pass

    # Pre-resolve static values for trip/user to avoid CPU/DB bottlenecks in inner loop
    user = trip.user
    user_designation = (user.designation or '').strip().lower()
    user_base_location = user.base_location or 'Other'
    user_active_role = getattr(user, 'active_role', '').lower()

    # Pre-resolve matched cadre and eligibility rule
    matched_cadre = None
    if user_designation:
        keyword_cadre_pairs = []
        for cadre in Cadre.objects.all():
            for kw in (cadre.designation_keywords or []):
                if kw:
                    keyword_cadre_pairs.append((str(kw).strip().lower(), cadre))
        
        keyword_cadre_pairs.sort(key=lambda x: len(x[0]), reverse=True)
        
        import re
        desig_words = re.findall(r'[a-z0-9]+', user_designation)
        desig_words_set = set(desig_words)
        
        for kw_clean, cadre in keyword_cadre_pairs:
            kw_words = re.findall(r'[a-z0-9]+', kw_clean)
            if not kw_words:
                continue
            if all(word in desig_words_set for word in kw_words):
                matched_cadre = cadre
                break
            if kw_clean in user_designation:
                matched_cadre = cadre
                break

    if not matched_cadre:
        if user_active_role in ['admin', 'cfo']:
            matched_cadre = Cadre.objects.filter(name__icontains='ADMINISTRATIVE').first()
        elif user_active_role in ['hr', 'finance']:
            matched_cadre = Cadre.objects.filter(name__icontains='MANAGERS').first()

    if not matched_cadre:
        matched_cadre = Cadre.objects.filter(name__icontains='BELOW EXECUTIVE').first()
    if not matched_cadre:
        matched_cadre = Cadre.objects.first()

    rule = EligibilityRule.objects.filter(cadre=matched_cadre).order_by('-id').first() if matched_cadre else None
    is_local = getattr(trip, 'consider_as_local', False)
    if rule:
        if is_local:
            daily_allowance_limit = float(rule.monthly_tour_daily_allowance_amount)
        else:
            daily_allowance_limit = float(rule.daily_allowance_amount)
    else:
        daily_allowance_limit = 300.0

    # Pre-fetch all fuel rate master objects
    fuel_rates = list(FuelRateMaster.objects.all())

    for batch in batches:
        data = batch.data_json
        if not isinstance(data, list):
            continue

        # Group and pre-calculate aggregated DA results for this batch
        date_groups = {}
        for idx, row in enumerate(data):
            row_status = row.get('_status')
            date_val = row.get('date', '')
            is_rejected = (row_status == 'Rejected')
            is_instruction = ('instruc' in str(date_val).lower())
            if is_rejected or is_instruction:
                continue

            date_str = str(date_val).strip()
            if not date_str:
                continue
            if len(date_str) > 10:
                date_str = date_str[:10]
            final_date = None
            for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]:
                try:
                    final_date = datetime.strptime(date_str, fmt).date().isoformat()
                    break
                except:
                    continue
            if not final_date:
                continue

            start_time_str = row.get('start_time') or row.get('startTime')
            reach_time_str = row.get('reach_time') or row.get('endTime') or row.get('reachTime')
            if not start_time_str:
                start_time_str = '09:00'
            if not reach_time_str:
                reach_time_str = '18:00'

            if final_date not in date_groups:
                date_groups[final_date] = []
            date_groups[final_date].append({
                'start_time': start_time_str,
                'reach_time': reach_time_str
            })

        aggregated_da_results = {}
        for f_date, items in date_groups.items():
            total_hours = 0.0
            earliest_dt = None
            latest_dt = None

            for item in items:
                try:
                    st_str = item['start_time']
                    rt_str = item['reach_time']

                    t1 = datetime.strptime(f"{f_date}T{st_str}", "%Y-%m-%dT%H:%M")
                    t2 = datetime.strptime(f"{f_date}T{rt_str}", "%Y-%m-%dT%H:%M")

                    if t2 < t1:
                        t2 += timedelta(days=1)

                    duration = (t2 - t1).total_seconds() / 3600.0
                    total_hours += duration

                    if earliest_dt is None or t1 < earliest_dt:
                        earliest_dt = t1
                    if latest_dt is None or t2 > latest_dt:
                        latest_dt = t2
                except Exception:
                    pass

            aggregated_da_results[f_date] = {
                'hours': total_hours,
                'start_time': earliest_dt.strftime('%H:%M') if earliest_dt else '09:00',
                'reach_time': latest_dt.strftime('%H:%M') if latest_dt else '18:00'
            }
            
        created_ids = []
        for idx, row in enumerate(data):
            row_status = row.get('_status')
            date_val = row.get('date', '')
            
            is_rejected = (row_status == 'Rejected')
            is_instruction = ('instruc' in str(date_val).lower())
            
            if is_rejected or is_instruction:
                for e in existing_expenses:
                    try:
                        d = json.loads(e.description)
                        if int(d.get('batch_id', -1)) == batch.id and int(d.get('row_index', -1)) == idx:
                            expenses_to_delete.append(e)
                    except: continue
                continue
                
            try:
                date_val = row.get('date', '')
                date_str = str(date_val).strip()
                if not date_str: continue
                if len(date_str) > 10: date_str = date_str[:10]
                
                final_date = None
                for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]:
                    try:
                        final_date = datetime.strptime(date_str, fmt).date().isoformat()
                        break
                    except: continue
                
                if not final_date:
                    print(f"DEBUG: Could not parse date '{date_str}' in row {idx}")
                    continue
                
                origin = str(row.get('origin_route', row.get('from_location', ''))).strip()
                destination = str(row.get('destination_route', row.get('to_location', ''))).strip()
                purpose = str(row.get('visit_intent', row.get('purpose', ''))).strip()

                odo_s = float(row.get('odo_start', 0) or 0)
                odo_e = float(row.get('odo_end', 0) or 0)
                distance = max(0.0, odo_e - odo_s)

                if (batch.id, idx) in processed_rows:
                    continue

                semantic_key = (final_date, odo_s, odo_e, origin.upper(), destination.upper())
                if semantic_key in semantic_map:
                    existing_id = semantic_map[semantic_key]
                    exp_obj = next((e for e in existing_expenses if e.id == existing_id), None)
                    if exp_obj:
                        try:
                            exp_desc = json.loads(exp_obj.description)
                            if exp_desc.get('from_bulk_upload'):
                                expenses_to_delete.append(exp_obj)
                                existing_expenses = [x for x in existing_expenses if x.id != exp_obj.id]
                                if semantic_key in semantic_map:
                                    del semantic_map[semantic_key]
                            else:
                                continue
                        except:
                            continue

                mapped_mode = str(row.get('mode', 'Bike')) or 'Bike'
                mapped_subType = str(row.get('vehicle', 'Own Bike')) or 'Own Bike'
                
                # In-memory lookup from fuel_rates
                rate_obj = None
                for r in fuel_rates:
                    if r.vehicle_type == mapped_mode and r.state and user_base_location.lower() in r.state.lower():
                        rate_obj = r
                        break
                if not rate_obj:
                    for r in fuel_rates:
                        if r.vehicle_type == mapped_mode:
                            rate_obj = r
                            break
                rate_per_km = float(rate_obj.rate_per_km) if rate_obj else 0.0
                
                amount = row.get('amount')
                if amount in [None, '', 0, '0']:
                    amount = distance * rate_per_km
                else:
                    try: amount = float(amount)
                    except: amount = 0.0

                if final_date not in processed_da_dates:
                    agg_info = aggregated_da_results.get(final_date, {'hours': 0.0})
                    hours = agg_info['hours']
                    
                    eligible_da = 0.0
                    if hours < 12:
                        da_message = "no DA is allowed for you based on hours"
                    elif 12 <= hours <= 18:
                        eligible_da = (daily_allowance_limit * 50.0) / 100.0
                        da_message = "50% based on hours"
                    else:
                        eligible_da = daily_allowance_limit
                        da_message = "100% based on hours"
                        
                    da_raw = row.get('daily_allowance')
                    if da_raw is not None and str(da_raw).strip() != '':
                        try:
                            applied_da = float(da_raw)
                        except:
                            applied_da = eligible_da
                    else:
                        applied_da = eligible_da
                        
                    processed_da_dates.add(final_date)
                else:
                    hours = 0.0
                    eligible_da = 0.0
                    applied_da = 0.0
                    da_message = f"Aggregated with other entries for this date ({final_date})"

                total_claimed = amount + applied_da
                total_allowed = amount + eligible_da

                desc_dict = {
                    'origin': origin,
                    'destination': destination,
                    'mode': mapped_mode,
                    'subType': mapped_subType,
                    'isPublicTransport': row.get('isPublicTransport', False),
                    'remainingRoute': row.get('remainingRoute', ''),
                    'odoStart': odo_s,
                    'odoEnd': odo_e,
                    'odoRate': rate_per_km,
                    'distance': distance,
                    'purpose': purpose,
                    'from_bulk_upload': True,
                    'batch_id': batch.id,
                    'row_index': idx,
                    'date': final_date,
                    'startDate': final_date,
                    'endDate': final_date,
                    'startTime': start_time_str,
                    'endTime': reach_time_str,
                    'is_deviated': row.get('is_deviated', False),
                    'deviation_reason': row.get('deviation_reason', ''),
                    'deviation_target': row.get('deviation_target', ''),
                    'daily_allowance': applied_da,
                    'eligible_da': eligible_da,
                    'da_hours': hours,
                    'da_message': da_message,
                    'fare_or_fuel': amount
                }
                
                if 'jobReportAttachments' in row:
                    desc_dict['jobReportAttachments'] = row['jobReportAttachments']
                if 'jobReport' in row:
                    desc_dict['jobReport'] = row['jobReport']
                if 'selfies' in row:
                    desc_dict['selfies'] = row['selfies']
                if 'odoStartImg' in row:
                    desc_dict['odoStartImg'] = row['odoStartImg']
                if 'odoEndImg' in row:
                    desc_dict['odoEndImg'] = row['odoEndImg']

                exp = Expense.objects.create(
                    trip=trip,
                    date=final_date,
                    category='Fuel' if 'bike' in mapped_mode.lower() or 'car' in mapped_mode.lower() else 'Local Travel',
                    amount=total_claimed,
                    allowed_amount=total_allowed,
                    description=json.dumps(desc_dict),
                    status='Approved',
                    rm_remarks=row.get('_remark') or desc_dict['purpose'],
                    odo_start=odo_s,
                    odo_end=odo_e,
                    distance=distance,
                    travel_mode=mapped_mode,
                    vehicle_type=mapped_subType
                )
                created_ids.append(exp.id)

                if 'incidentals' in row and isinstance(row['incidentals'], list):
                    for inc in row['incidentals']:
                        try:
                            inc_amt = float(inc.get('amount', 0) or 0)
                            if inc_amt <= 0: continue
                            
                            inc_cat = inc.get('category', 'Other')
                            inc_desc = {
                                'purpose': f"Incidental for {purpose}",
                                'category': inc_cat,
                                'parent_row_index': idx,
                                'from_bulk_upload': True,
                                'batch_id': batch.id
                            }
                            
                            inc_exp = Expense.objects.create(
                                trip=trip,
                                date=final_date,
                                category=inc_cat,
                                amount=inc_amt,
                                description=json.dumps(inc_desc),
                                status='Approved',
                                rm_remarks=f"Incidental from {final_date}: {inc_cat}"
                            )
                            created_ids.append(inc_exp.id)
                        except Exception as e:
                            print(f"Error creating incidental expense: {e}")
            except Exception as e:
                print(f"Error creating expense from batch row {idx}: {e}")
                
        if created_ids:
            batch.created_expenses = (batch.created_expenses or []) + created_ids
            batch.save(update_fields=['created_expenses'])

    # Perform any final batch deletions (for rejected/instruction rows matched during processing)
    if expenses_to_delete:
        delete_ids = [e.id for e in expenses_to_delete]
        Expense.objects.filter(id__in=delete_ids).delete()

def resolve_approver(user, members_data=None):
    """Helper to resolve the first approver in the management hierarchy."""
    # INTELLIGENT SAFEGARD OPTIMIZATION: 
    # 1. Check if a manager already exists in the warm cache (takes 0.005s).
    # 2. If YES: Skip synchronous network sync and proceed instantly, spawning a background thread to keep the cache warm.
    # 3. If NO: Run the synchronous safeguard fetch to prevent accidental "No Manager" auto-approvals.
    cached_rm = None
    try:
        cached_rm = user.reporting_manager
    except Exception:
        pass

    if not cached_rm:
        if hasattr(user, 'get_available_positions'):
            try:
                # Removed force_fresh=True from the synchronous safeguard
                # We already have persistent cache now, so we can trust the cache layer.
                user.get_available_positions(force_fresh=False)
            except Exception as e:
                print(f"Warning: Fetch failed in resolve_approver: {e}")
    else:
        # Normal Flow: Manager is present! Spawn silent background daemon to update cache without holding up user
        import threading
        def _bg_cache_warmup(u_id):
            try:
                from django.db import connection
                connection.close()
                from core.models import User
                u = User.objects.get(pk=u_id)
                # Background thread can still force refresh to keep the data updated
                u.get_available_positions(force_fresh=True)
            except Exception:
                pass
                
        t = threading.Thread(target=_bg_cache_warmup, args=(user.pk,))
        t.daemon = True
        t.start()
            
    reporting_manager = user.reporting_manager
    
    # SPECIAL CASE: If direct reporting manager is COO, bypass them. 
    # The caller will treat this as is_top_level and trigger HR/Finance dispatch.
    if reporting_manager:
        mgr_pos = reporting_manager.get_current_position()
        mgr_pos_name = mgr_pos.get('name') if mgr_pos else None
        if _is_coo_position(mgr_pos_name, reporting_manager.designation, employee_id=reporting_manager.employee_id):
            # Check if COO approval is explicitly enabled for this project
            from travel.models import COOProjectSetting
            proj_code = getattr(user, 'project_code', 'General') or 'General'
            mgr_pos_id = mgr_pos.get('id') if mgr_pos else None
            
            coo_enabled = False
            if mgr_pos_id:
                coo_enabled = COOProjectSetting.objects.filter(project_code=proj_code, coo_position_id=str(mgr_pos_id), enable_coo_approval=True).exists()
                
            if not coo_enabled:
                # Bypass COO: return None as current_approver to trigger finalized/finance routing
                return None, 0, reporting_manager, user.senior_manager, user.hod_director, None

    # Strictly follow reporting managers flow
    current_approver = reporting_manager
    h_level = 1
    
    # Snapshots for resilience
    sm = user.senior_manager
    hod = user.hod_director
    
    # Resolve position ID based on which level was selected
    approver_position = None
    if current_approver:
        # Check hierarchy positions first
        if current_approver == reporting_manager:
            approver_position = user.reporting_manager_position
        elif current_approver == sm:
            approver_position = user.senior_manager_position
        elif current_approver == hod:
            approver_position = user.hod_director_position
        
        # Fallback to manager's own active/primary position if not in requester's direct stack
        if not approver_position:
            approver_position = getattr(current_approver, 'active_position_id', None)
            if not approver_position:
                pos = current_approver.get_current_position()
                approver_position = pos.get('id') if pos else None
                
    return current_approver, h_level, reporting_manager, sm, hod, str(approver_position) if approver_position else None


class TripListCreateView(generics.ListCreateAPIView):
    serializer_class = TripSerializer
    permission_classes = [IsCustomAuthenticated]

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return Trip.objects.none()
            
        all_trips = self.request.query_params.get('all') == 'true'
        user_role = user.active_role.lower()
        
        from django.db.models import Case, When, Value, IntegerField
        
        if all_trips:
            if user_role in ['admin', 'guesthousemanager', 'finance', 'cfo']:
                queryset = Trip.objects.all().order_by('-created_at')
            else:
                q_requester = Q(user=user) & Q(requester_position=user.active_position_id)
                q_approver = Q(current_approver=user) & Q(approver_position=user.active_position_id)
                
                queryset = Trip.objects.filter(
                    q_requester | q_approver
                ).distinct().annotate(
                    priority=Case(
                        When(status__in=[
                            'Approved', 'Completed', 'Resubmitted', 'Claim Submitted', 
                            'Submitted', 'Pending', 'pending_hr', 'pending_executive', 
                            'pending_head', 'pending_final_release',
                            'approved', 'completed', 'resubmitted', 'claim submitted', 
                            'submitted', 'pending'
                        ], then=Value(1)),
                        When(status__in=[
                            'Settled', 'Paid', 'Transferred', 'Completed & Settled',
                            'settled', 'paid', 'transferred', 'completed & settled'
                        ], then=Value(3)),
                        default=Value(2),
                        output_field=IntegerField(),
                    )
                ).order_by('priority', '-created_at')
        else:
            q = Q(user=user) & Q(requester_position=user.active_position_id)
            queryset = Trip.objects.filter(q).filter(Q(consider_as_local=False) | Q(consider_as_local=True, activity_batches__isnull=True))
            search_query = self.request.query_params.get('search')
            if search_query:
                queryset = queryset.filter(
                    Q(trip_id__icontains=search_query) |
                    Q(purpose__icontains=search_query) |
                    Q(source__icontains=search_query) |
                    Q(destination__icontains=search_query)
                )
            queryset = queryset.annotate(
                priority=Case(
                    When(status__in=[
                        'Approved', 'Completed', 'Resubmitted', 'Claim Submitted', 
                        'Submitted', 'Pending', 'pending_hr', 'pending_executive', 
                        'pending_head', 'pending_final_release',
                        'approved', 'completed', 'resubmitted', 'claim submitted', 
                        'submitted', 'pending'
                    ], then=Value(1)),
                    When(status__in=[
                        'Settled', 'Paid', 'Transferred', 'Completed & Settled',
                        'settled', 'paid', 'transferred', 'completed & settled'
                    ], then=Value(3)),
                    default=Value(2),
                    output_field=IntegerField(),
                )
            ).order_by('priority', '-created_at')
            
        # Optimize query by resolving Cascading N+1 queries at database level
        return queryset.select_related(
            'user', 'user__role',
            'current_approver', 'claim', 'route_path'
        ).prefetch_related(
            'advances', 'expenses', 'job_reports', 'activity_batches',
            'room_bookings', 'vehicle_bookings', 'odometer_details'
        )

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        from rest_framework.exceptions import NotFound
        try:
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)
        except NotFound:
            # Return a manual paginated response because paginator.page won't be set
            return Response({
                'count': queryset.count(),
                'next': None,
                'previous': None,
                'results': []
            })
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    def perform_create(self, serializer, is_local=False):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            from rest_framework.exceptions import AuthenticationFailed # type: ignore
            raise AuthenticationFailed("Authentication required.")
        
        user_role = user.role.name.lower() if user.role else ''
        is_local_trip = is_local or (
            serializer.validated_data.get('consider_as_local', False)
            and 'consider_as_local' in self.request.data
        )

        # Check if there is already an approved/forwarded tour plan for the month of start_date
        start_date = serializer.validated_data.get('start_date')
        is_tour_plan = self.request.data.get('is_tour_plan', False)
        if isinstance(is_tour_plan, str):
            is_tour_plan = is_tour_plan.lower() == 'true'
        if start_date and is_tour_plan and user_role not in ['admin', 'superuser', 'it-admin']:
            month = start_date.month
            year = start_date.year
            
            # Find trips in the same month/year that are also local travel (tour plans with activity batches)
            existing_trips = Trip.objects.filter(
                user=user,
                start_date__year=year,
                start_date__month=month,
                consider_as_local=True,
                activity_batches__isnull=False
            ).distinct()
            
            for t in existing_trips:
                has_rejected_row = False
                for batch in t.activity_batches.all():
                    rows = batch.data_json
                    if isinstance(rows, list):
                        if any(row.get('_status') == 'Rejected' for row in rows if isinstance(row, dict)):
                            has_rejected_row = True
                            break

                if has_rejected_row:
                    raise ValidationError({
                        "detail": f"A tour plan for {start_date.strftime('%B %Y')} has been rejected. "
                                  "Please resubmit the existing tour plan instead of creating a new one."
                    })

                # Determine if it has been approved by any position and sent for next approval
                has_approval = False
                if t.status.lower() == 'rejected':
                    has_approval = False
                elif t.hierarchy_level > 1:
                    has_approval = True
                elif t.status.lower() in [
                    'approved', 'completed', 'settled', 'paid', 'transferred', 
                    'completed & settled', 'claim submitted', 'pending_hr', 
                    'pending_executive', 'pending_head', 'pending_final_release'
                ]:
                    has_approval = True
                else:
                    events = t.lifecycle_events or []
                    if any('Approval' in str(e.get('title')) or 'Approved by' in str(e.get('description')) for e in events):
                        has_approval = True
                
                if has_approval:
                    raise ValidationError({
                        "detail": f"A tour plan for {start_date.strftime('%B %Y')} has already been approved and forwarded. "
                                  "You cannot submit another tour plan for this month unless the existing one is rejected."
                    })
                else:
                    # If it exists but is not yet approved/forwarded, overwrite it by soft-deleting the existing one
                    for expense in t.expenses.all():
                        expense.delete()
                    for batch in t.activity_batches.all():
                        batch.delete()
                    for advance in t.advances.all():
                        advance.delete()
                    for tracking in t.tracking_history.all():
                        tracking.delete()
                    for dispute in t.disputes.all():
                        dispute.delete()
                    for job in t.job_reports.all():
                        job.delete()
                    for stop in t.historical_stops.all():
                        stop.delete()
                    for fin_int in t.finance_intimations.all():
                        fin_int.delete()
                    for hr_int in t.hr_intimations.all():
                        hr_int.delete()
                    for room in t.room_bookings.all():
                        room.delete()
                    for veh in t.vehicle_bookings.all():
                        veh.delete()
                    for rem in t.reminders.all():
                        rem.delete()

                    # Try deleting One-to-One relations safely
                    try:
                        t.odometer_details.delete()
                    except:
                        pass
                    try:
                        t.geofence_set.delete()
                    except:
                        pass
                    try:
                        t.claim.delete()
                    except:
                        pass
                    t.delete()

        # Resolve employee's project code from employee details first, with fallback to serializer data
        project_code = 'General'
        if hasattr(user, 'project_code') and user.project_code and user.project_code != 'N/A':
            project_code = user.project_code
        else:
            project_code = serializer.validated_data.get('project_code', 'General')

        # Admin / Superuser skip approvals
        if user_role in ['admin', 'superuser', 'it-admin']:
            trip = serializer.save(
                user=user,
                status='Approved',
                current_approver=None,
                hierarchy_level=0,
                consider_as_local=is_local_trip,
                project_code=project_code
            )
            label = "Travel" if is_local_trip else "Trip"
            update_trip_lifecycle(trip, "Auto-Approved", f"{label} request auto-approved for Administrator.")
            return

        # Check if Employee API is unreachable (to prevent accidental auto-approval for standard users)
        from core.middleware import should_skip_external_api
        if not user._get_api_data() and not should_skip_external_api():
            raise ValidationError({
                "detail": "Failed to resolve your employee profile from the HCM API. "
                          "Please contact support or ensure the staging server has network access to the API."
            })

        members_data = serializer.validated_data.get('members', [])
        current_approver, h_level, rm, sm, hod, pos_id = resolve_approver(user, members_data)

        
        # Check if HR approval is required due to COO bypass
        user_dept = str(user.department or '').strip().lower()
        user_desig = str(user.designation or '').strip().lower()
        user_pos = str(user.active_position_id or '').strip()
        is_sph = ('sph' in user_dept or 'sph' in user_desig or user_pos == '2')

        reports_to_coo = False
        if rm:
            rm_pos = rm.get_current_position()
            rm_pos_name = rm_pos.get('name') if rm_pos else None
            reports_to_coo = _is_coo_position(rm_pos_name, rm.designation, employee_id=rm.employee_id)
        
        if is_sph:
            reports_to_coo = True
            
        from .models import HRPositionConfig
        hr_positions = HRPositionConfig.objects.filter(is_active=True, project_code=project_code)
        if not hr_positions.exists():
            hr_positions = HRPositionConfig.objects.filter(is_active=True, project_code='General')
        hr_requires_approval = hr_positions.filter(trips_approval='APPROVAL').exists()

        is_top_level = (current_approver is None)
        trip_status = 'Pending' if hr_requires_approval else ('Approved' if is_top_level else 'Pending')
        
        try:
            # Pre-calculate the full approval chain snapshot
            chain = build_approval_chain(user)
            
            trip = serializer.save(
                user=user,
                requester_position=user.active_position_id,
                status=trip_status,
                current_approver=current_approver,
                approver_position=pos_id,
                hierarchy_level=0 if is_top_level else h_level,
                consider_as_local=is_local_trip,
                approval_chain=chain,
                project_code=project_code,
                # Snapshots
                user_name=user.name,
                user_designation=user.designation,
                user_department=user.department,
                reporting_manager_name=rm.name if rm else None,
                senior_manager_name=sm.name if sm else None,
                hod_director_name=hod.name if hod else None
            )
            
            # Parallel HR Intimation trigger for top level
            if is_top_level:
                trigger_parallel_dispatch(trip, user)
                rm_pos = rm.get_current_position() if rm else None
                rm_pos_name = rm_pos.get('name') if rm_pos else None
                if rm and _is_coo_position(rm_pos_name, rm.designation, employee_id=rm.employee_id):
                    if hr_requires_approval:
                        update_trip_lifecycle(trip, "Submitted", "Reporting COO bypassed. Trip request submitted for HR Approval.")
                    else:
                        update_trip_lifecycle(trip, "Finalized", "Trip request finalized. Reporting COO bypassed and HR Intimated.")
                else:
                    update_trip_lifecycle(trip, "Auto-Approved", "Trip request auto-approved since employee has no reporting manager. HR Intimated.")
                
        except Exception as e:
            # convert DB errors to validation error so frontend sees message
            # also log full traceback on server for diagnostics
            import traceback
            traceback.print_exc()
            msg = str(e)
            print("ERROR IN TRIP CREATION:", msg)
            raise ValidationError({"detail": msg})

        label = "Travel" if is_local_trip else "Trip"
        if current_approver:
            # Notify Approver
            Notification.objects.create(
                user=current_approver,
                target_position=trip.approver_position,
                title=f"New {label} Request",
                message=f"{user.name} has submitted a new {label.lower()} request to {trip.destination}.",
                type='info'
            )
            # Notify Requester
            Notification.objects.create(
                user=user,
                title=f"{label} Request Submitted",
                message=f"Your {label.lower()} request to {trip.destination} (ID: {trip.trip_id}) has been sent to {current_approver.name} for approval.",
                type='success'
            )
        
        if trip.accommodation_requests and any('Room' in r for r in trip.accommodation_requests):
            gh_managers = User.objects.filter(role__name='GuestHouseManager', is_active=True)
            for manager in gh_managers:
                Notification.objects.create(
                    user=manager,
                    title="Room Request Received",
                    message=f"{user.name} has requested a room for {label.lower()} {trip.trip_id}.",
                    type='info'
                )

        notify_hr(f"New {label} Request", f"{user.name} has raised a {label.lower()} request to {trip.destination} (ID: {trip.trip_id}).")


class TravelListCreateView(TripListCreateView):
    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user: return Trip.objects.none()
        queryset = Trip.objects.filter(user=user, consider_as_local=True, activity_batches__isnull=False).exclude(status='Draft').distinct()
        search_query = self.request.query_params.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(trip_id__icontains=search_query) |
                Q(purpose__icontains=search_query) |
                Q(source__icontains=search_query) |
                Q(destination__icontains=search_query)
            )
        from django.db.models import Case, When, Value, IntegerField
        return queryset.annotate(
            priority=Case(
                When(status__in=[
                    'Approved', 'Completed', 'Resubmitted', 'Claim Submitted', 
                    'Submitted', 'Pending', 'pending_hr', 'pending_executive', 
                    'pending_head', 'pending_final_release',
                    'approved', 'completed', 'resubmitted', 'claim submitted', 
                    'submitted', 'pending'
                ], then=Value(1)),
                When(status__in=[
                    'Settled', 'Paid', 'Transferred', 'Completed & Settled',
                    'settled', 'paid', 'transferred', 'completed & settled'
                ], then=Value(3)),
                default=Value(2),
                output_field=IntegerField(),
            )
        ).order_by('priority', '-created_at').select_related(
            'user', 'user__role',
            'current_approver', 'claim', 'route_path'
        ).prefetch_related(
            'advances', 'expenses', 'job_reports', 'activity_batches',
            'room_bookings', 'vehicle_bookings', 'odometer_details'
        )

    def perform_create(self, serializer, is_local=True): # type: ignore
        super().perform_create(serializer, is_local=is_local)

class TripBookingSearchView(generics.ListAPIView):
    serializer_class = TripSerializer
    permission_classes = [IsCustomAuthenticated] 

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return Trip.objects.none()

        # Users only search their own trips in this view
        queryset = Trip.objects.filter(user=user)

        search_query = self.request.query_params.get('search', None)
        if search_query:
            try:
                import base64
                padding = len(search_query) % 4
                if padding: 
                    search_query += '=' * (4 - padding)
                search_query = base64.b64decode(search_query).decode('utf-8')
            except Exception as e:
                print(f"Decoding error: {e}")
                pass

            queryset = queryset.filter(
                Q(trip_id__istartswith=search_query) | 
                Q(purpose__istartswith=search_query) | 
                Q(source__istartswith=search_query) | 
                Q(destination__istartswith=search_query) | 
                Q(trip_leader__istartswith=search_query)
            )
        
        return queryset

class TripDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Trip.objects.all()
    serializer_class = TripSerializer
    permission_classes = [IsCustomAuthenticated]
    lookup_field = 'trip_id'

    def get_object(self):
        if 'trip_id' in self.kwargs:
            self.kwargs['trip_id'] = decode_id(self.kwargs['trip_id'])
        
        obj = super().get_object()
        user = getattr(self.request, 'custom_user', None)
        if not user:
            raise PermissionDenied("Not authenticated")
            
        user_role = user.active_role.lower()
        is_admin = user_role in ['admin', 'it-admin', 'superuser']
        is_finance = user_role in ['finance', 'cfo']
        is_gh_manager = user_role == 'guesthousemanager'
        
        # Authorization check: Owner, Managers in the hierarchy, Current Approver, Finance, Guest House Manager, or Admin
        is_owner = (obj.user == user)
        # Check if the current user is any of the requester's managers or the current approver
        is_manager = user in [obj.user.reporting_manager, obj.user.senior_manager, obj.user.hod_director, obj.current_approver]
        
        if not (is_owner or is_manager or is_admin or is_finance or is_gh_manager):
             self.permission_denied(self.request, message="Not authorized to view this trip details")
             
        return obj

class TripTrackingView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request, trip_id):
        print(f"DEBUG: TripTrackingView.get called for trip_id: {trip_id}")
        real_trip_id = decode_id(trip_id)
        # Verify trip exists and user has access
        try:
            trip = Trip.objects.get(trip_id=real_trip_id)
        except Trip.DoesNotExist:
            print(f"DEBUG: Trip {real_trip_id} not found")
            return Response({"error": "Trip not found"}, status=status.HTTP_404_NOT_FOUND)

        # Basic access check: requester or manager or finance or admin
        user = getattr(request, 'custom_user', None)
        print(f"DEBUG: Requester: {user.employee_id if user else 'Anonymous'}")
        
        # ... existing logic ...
        is_owner = (trip.user == user)
        is_manager = False
        if user:
            is_manager = user in [trip.user.reporting_manager, trip.user.senior_manager, trip.user.hod_director, trip.current_approver]
        
        user_role = user.active_role.lower() if user else ''
        is_privileged = user_role in ['admin', 'finance', 'cfo', 'guesthousemanager']

        if not (is_owner or is_manager or is_privileged):
            print(f"DEBUG: Unauthorized access attempt to trip {real_trip_id}")
            return Response({"error": "Unauthorized"}, status=status.HTTP_403_FORBIDDEN)

        is_latest = request.query_params.get('latest') == 'true'
        if is_latest:
            latest_point = TripTracking.objects.filter(trip=trip).order_by('-timestamp').first()
            if not latest_point:
                return Response({"error": "No tracking data found"}, status=status.HTTP_404_NOT_FOUND)
            serializer = TripTrackingSerializer(latest_point)
            return Response(serializer.data)

        tracking_data = TripTracking.objects.filter(trip=trip).order_by('timestamp')
        print(f"DEBUG: Returning {tracking_data.count()} points")
        serializer = TripTrackingSerializer(tracking_data, many=True)
        return Response(serializer.data)

    def post(self, request, trip_id):
        print(f"DEBUG: TripTrackingView.post called for trip_id: {trip_id}")
        real_trip_id = decode_id(trip_id)
        
        try:
            trip = Trip.objects.get(trip_id=real_trip_id)
        except Trip.DoesNotExist:
            print(f"DEBUG: Trip {real_trip_id} not found for POST")
            return Response({"error": "Trip not found"}, status=status.HTTP_404_NOT_FOUND)

        user = getattr(request, 'custom_user', None)
        print(f"DEBUG: POST Requester: {user.employee_id if user else 'Anonymous'}")

        if not user or trip.user != user:
            print(f"DEBUG: POST Unauthorized for user {user}")
            return Response({"error": "Only trip owner can submit tracking data"}, status=status.HTTP_403_FORBIDDEN)

        data = request.data.copy()
        data['trip'] = trip.pk
        
        serializer = TripTrackingSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            print(f"DEBUG: Tracking point saved successfully")
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        
        print(f"DEBUG: Serializer errors: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class HistoricalTripStopsView(generics.ListAPIView):
    permission_classes = [IsCustomAuthenticated]
    serializer_class = HistoricalTripStopSerializer

    def get(self, request, *args, **kwargs):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return Response({"error": "Unauthorized"}, status=401)
        
        date_str = self.request.query_params.get('date')
        if not date_str:
            return Response({"stops": [], "breadcrumbs": []})
        
        try:
            date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({"stops": [], "breadcrumbs": []})

        # Cleanup: Delete records older than 7 days
        retention_limit = timezone.now() - datetime.timedelta(days=7)
        HistoricalTripStop.objects.filter(date__lt=retention_limit.date()).delete()
        TripTracking.objects.filter(timestamp__lt=retention_limit).delete()

        target_employee_id = self.request.query_params.get('employee_id')
        target_user = user
        if target_employee_id:
            found_user = User.objects.filter(employee_id=target_employee_id).first()
            if found_user:
                # Permission check
                is_manager = user in [found_user.reporting_manager, found_user.senior_manager, found_user.hod_director]
                user_role = user.active_role.lower()
                is_privileged = user_role in ['admin', 'finance', 'cfo']
                
                if is_manager or is_privileged or found_user == user:
                    target_user = found_user
                else:
                    return Response({"error": "Forbidden"}, status=403)

        # Get Stops (Markers)
        stops = HistoricalTripStop.objects.filter(user=target_user, date=date_obj).order_by('arrival_time')
        stops_serializer = HistoricalTripStopSerializer(stops, many=True)

        # Get Breadcrumbs (Polyline path)
        # We filter TripTracking based on the selected date
        # Note: We use __date lookup on the timestamp field
        breadcrumbs = TripTracking.objects.filter(
            trip__user=target_user, 
            timestamp__date=date_obj
        ).order_by('timestamp')
        breadcrumbs_serializer = TripTrackingSerializer(breadcrumbs, many=True)

        return Response({
            "stops": stops_serializer.data,
            "breadcrumbs": breadcrumbs_serializer.data
        })

    def post(self, request, trip_id):
        print(f"DEBUG: TripTrackingView.post called for trip_id: {trip_id}")
        real_trip_id = decode_id(trip_id)
        print(f"DEBUG: Decoded trip_id: {real_trip_id}")
        
        trip = None
        try:
            trip = Trip.objects.get(trip_id=real_trip_id)
        except Trip.DoesNotExist:
            # Fallback: check if the trip exists but is soft-deleted
            try:
                deleted_trip = Trip.all_objects.get(trip_id=real_trip_id)
                print(f"DEBUG: Trip {real_trip_id} exists but is soft-deleted (is_deleted={deleted_trip.is_deleted}). Allowing tracking.")
                trip = deleted_trip  # Allow tracking even for soft-deleted trips
            except Trip.DoesNotExist:
                print(f"DEBUG: Trip {real_trip_id} does NOT exist in DB at all (even in all_objects). POST rejected.")
                return Response({"error": "Trip not found"}, status=status.HTTP_404_NOT_FOUND)

        # Only trip owner can post tracking points
        user = getattr(request, 'custom_user', None)
        print(f"DEBUG: POST Requester: {user.employee_id if user else 'Anonymous'}")
        
        if not user or trip.user != user:
            print(f"DEBUG: POST Unauthorized for user {user.employee_id if user else 'None'}")
            return Response({"error": "Only trip owner can submit tracking data"}, status=status.HTTP_403_FORBIDDEN)

        data = request.data.copy()
        data['trip'] = trip.trip_id
        
        serializer = TripTrackingSerializer(data=data)
        if serializer.is_valid():
            tracking = serializer.save()
            print("DEBUG: Tracking point saved successfully")
            
            try:
                from django.db import connection
                if 'travel_tripgeofencelocationset' in connection.introspection.table_names():
                    from .models import TripGeofenceLocationSet
                    geofence_set, created = TripGeofenceLocationSet.objects.get_or_create(trip=trip)
                    loc_data = geofence_set.location_data
                    if not isinstance(loc_data, list):
                        loc_data = []
                    # add point to the set
                    loc_data.append({
                        "latitude": float(tracking.latitude),
                        "longitude": float(tracking.longitude),
                        "timestamp": tracking.timestamp.isoformat() if tracking.timestamp else None,
                        "accuracy": tracking.accuracy,
                        "speed": tracking.speed
                    })
                    geofence_set.location_data = loc_data
                    geofence_set.last_latitude = tracking.latitude
                    geofence_set.last_longitude = tracking.longitude
                    geofence_set.save()
            except Exception as e:
                print(f"DEBUG: Error updating Geofence Location Set: {e}")
            # --------------------------------------

            return Response(serializer.data, status=status.HTTP_201_CREATED)
        
        print(f"DEBUG: Serializer errors: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class TeamLiveTrackingView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request):
        user = getattr(request, 'custom_user', None)
        if not user:
            return Response({"error": "Unauthorized"}, status=403)
        
        from django.utils import timezone
        today = timezone.now().date()
        
        # Get ALL active approved trips today (start_date <= today <= end_date)
        # We must filter reporting_manager in Python since it's a dynamic property on the User model
        active_trips = Trip.all_objects.filter(
            start_date__lte=today,
            end_date__gte=today,
            status__iexact='Approved'
        )

        from core.models import Session
        from django.db import connection
        
        has_geofence_table = 'travel_tripgeofencelocationset' in connection.introspection.table_names()
        if has_geofence_table:
            from travel.models import TripGeofenceLocationSet

        results = []
        for trip in active_trips:
            if not trip.user:
                continue
                
            # Evaluate the property to check if current user is the manager
            rm = trip.user.reporting_manager
            if not rm or rm.employee_id != user.employee_id:
                continue

            # Check if the employee is currently logged out based on their most recent session
            latest_session = Session.objects.filter(user=trip.user).order_by('-created_at').first()
            is_logged_in = latest_session.is_active if latest_session else False

            # In case the user is logged out, show the last synced geofence from TripGeofenceLocationSet
            geofence_set = None
            if has_geofence_table:
                try:
                    geofence_set = TripGeofenceLocationSet.objects.filter(trip=trip).first()
                except Exception:
                    pass
            
            latest_tracking = TripTracking.objects.filter(trip=trip).order_by('-timestamp').first()
            
            lat = float(latest_tracking.latitude) if latest_tracking else (float(geofence_set.last_latitude) if geofence_set and geofence_set.last_latitude else None)
            lng = float(latest_tracking.longitude) if latest_tracking else (float(geofence_set.last_longitude) if geofence_set and geofence_set.last_longitude else None)
            accuracy = latest_tracking.accuracy if latest_tracking else None
            speed = latest_tracking.speed if latest_tracking else None
            last_updated = latest_tracking.timestamp if latest_tracking else (geofence_set.last_updated if geofence_set else None)

            results.append({
                "trip_id": trip.trip_id,
                "employee_name": trip.user.name,
                "employee_id": trip.user.employee_id,
                "destination": trip.destination,
                "purpose": trip.purpose,
                "consider_as_local": trip.consider_as_local,
                "status": trip.status,
                "is_logged_out": not is_logged_in,
                "latitude": lat,
                "longitude": lng,
                "last_updated": last_updated,
                "accuracy": accuracy,
                "speed": speed,
            })

        return Response(results, status=200)

def _get_actual_pending_tasks_count(user, view_type='all'):
    """
    Calculates the exact count of actionable items pending this user's attention.
    Mirrors ApprovalsView logic EXACTLY to ensure counts match the displayable Inbox.
    """
    user_role = (user.role.name.lower() if user.role else '')
    is_admin = user_role in ['admin', 'it-admin', 'superuser']
    is_finance_exec = _is_finance_executive(user)
    is_finance_head = _is_finance_head(user)
    is_hr = _is_hr(user)
    is_finance = is_finance_exec or is_finance_head

    trips = Trip.objects.none()
    advances = TravelAdvance.objects.none()
    claims = TravelClaim.objects.none()
    batch_qs = BulkActivityBatch.objects.none()

    if is_admin:
        status_list = ['Pending', 'Submitted', 'Forwarded', 'Resubmitted', 'Manager Approved', 'HR Approved', 'Under Process', 'PENDING_EXECUTIVE', 'PENDING_HEAD', 'PENDING_FINAL_RELEASE', 'REJECTED_BY_HEAD', 'Approved', 'PARTIALLY_COMPLETED']
        trips = Trip.objects.filter(status__in=status_list)
        advances = TravelAdvance.objects.filter(status__in=status_list)
        claims = TravelClaim.objects.filter(status__in=status_list)
        batch_qs = BulkActivityBatch.objects.filter(status__in=['Pending', 'Submitted', 'Forwarded', 'Resubmitted', 'Manager Approved', 'HR Approved', 'Under Process'])
    elif is_hr:
        from .models import HRIntimation
        intimations = HRIntimation.objects.filter(hr_user=user, is_read=False)
        trip_ids = list(intimations.filter(trip__isnull=False).values_list('trip_id', flat=True))
        advance_ids = list(intimations.filter(advance__isnull=False).values_list('advance_id', flat=True))
        claim_ids = list(intimations.filter(claim__isnull=False).values_list('claim_id', flat=True))
        
        trips = Trip.objects.filter(pk__in=trip_ids, status__in=['Pending', 'Submitted', 'Forwarded', 'Resubmitted', 'Manager Approved'])
        advances = TravelAdvance.objects.filter(pk__in=advance_ids, status__in=['Pending', 'Submitted', 'Forwarded', 'Resubmitted', 'Manager Approved', 'PENDING_HR'])
        claims = TravelClaim.objects.filter(pk__in=claim_ids, status__in=['Pending', 'Submitted', 'Forwarded', 'Resubmitted', 'Manager Approved', 'PENDING_HR'])
        batch_qs = BulkActivityBatch.objects.filter(trip_id__in=trip_ids)
    elif is_finance:
        user_step = _get_finance_step_for_user(user)
        
        # Determine access types for Finance Inbox and Finance Hub
        if user_step:
            has_inbox_access = user_step.visibility_type in ['INBOX', 'BOTH']
            has_hub_access = user_step.visibility_type in ['FINANCE_HUB', 'BOTH'] or is_finance_head
        else:
            # Fallback for users without registry configurations
            # Finance Head sees PENDING_HEAD; regular Finance Execs see both Inbox and standard Hub
            has_inbox_access = True
            has_hub_access = not is_finance_head

        inbox_trips = Trip.objects.none()
        inbox_adv = TravelAdvance.objects.none()
        inbox_claims = TravelClaim.objects.none()
        
        hub_adv = TravelAdvance.objects.none()
        hub_claims = TravelClaim.objects.none()

        if has_inbox_access:
            if user_step:
                fin_pos_ids = _get_user_all_position_ids(user)
                if user_step.position_id:
                    fin_pos_ids.add(str(user_step.position_id))
                pos_q = Q(approver_position__in=fin_pos_ids)
                inbox_statuses = ['PENDING_EXECUTIVE', 'PENDING_HEAD', 'REJECTED_BY_HEAD', 'HR Approved']
            else:
                fallback_pos_ids = user.get_active_position_identifiers()
                pos_q = Q(approver_position__in=fallback_pos_ids)
                inbox_statuses = ['PENDING_HEAD'] if is_finance_head else ['PENDING_EXECUTIVE', 'REJECTED_BY_HEAD', 'HR Approved']

            if not user_step and is_finance_head:
                # Line 1772-1773: Fallback Head strictly queries PENDING_HEAD
                inbox_adv = TravelAdvance.objects.filter(pos_q, status='PENDING_HEAD')
                inbox_claims = TravelClaim.objects.filter(pos_q, status='PENDING_HEAD')
            else:
                inbox_trips = Trip.objects.filter(status__in=inbox_statuses).filter(pos_q)
                inbox_adv = TravelAdvance.objects.filter(status__in=inbox_statuses).filter(pos_q)
                inbox_claims = TravelClaim.objects.filter(status__in=inbox_statuses).filter(pos_q)

        if has_hub_access:
            hub_statuses = ['PENDING_FINAL_RELEASE', 'Approved', 'PARTIALLY_COMPLETED']
            if user_step:
                fin_pos_ids = _get_user_all_position_ids(user)
                if user_step.position_id:
                    fin_pos_ids.add(str(user_step.position_id))
                pos_q = (Q(approver_position__in=fin_pos_ids) | Q(approver_position__isnull=True))
                hub_adv = TravelAdvance.objects.filter(status__in=hub_statuses).filter(pos_q)
                hub_claims = TravelClaim.objects.filter(status__in=hub_statuses).filter(pos_q)
            else:
                # Fallback normal users see all Hub releases unfiltered by position (Line 1781-1782)
                hub_adv = TravelAdvance.objects.filter(status__in=hub_statuses)
                hub_claims = TravelClaim.objects.filter(status__in=hub_statuses)

        # In Finance list, only Inbox Trips are displayed. Hub Trips are excluded to avoid duplicates.
        trips = inbox_trips.distinct()
        advances = (inbox_adv | hub_adv).distinct()
        claims = (inbox_claims | hub_claims).distinct()
        batch_qs = BulkActivityBatch.objects.none()
    else:
        # Position-centric manager query
        manager_pos_ids = user.get_active_position_identifiers()

        q = Q(approver_position__in=manager_pos_ids) | Q(current_approver=user, approver_position__isnull=True)
        status_list = ['Pending', 'Submitted', 'Forwarded', 'Resubmitted']
        trips = Trip.objects.filter(q, status__in=status_list)
        advances = TravelAdvance.objects.filter(q, status__in=status_list)
        claims = TravelClaim.objects.filter(q, status__in=status_list)
        batch_qs = BulkActivityBatch.objects.filter(q, status__in=['Pending', 'Submitted', 'Forwarded', 'Resubmitted', 'Manager Approved'])

    trips = trips.distinct()
    advances = advances.distinct()
    claims = claims.distinct()
    batch_qs = batch_qs.distinct()

    if view_type == 'special':
        trips = trips.filter(Q(consider_as_local=False) | Q(consider_as_local=True, activity_batches__isnull=True))
        advances = advances.filter(Q(trip__consider_as_local=False) | Q(trip__consider_as_local=True, trip__activity_batches__isnull=True))
        claims = claims.filter(Q(trip__consider_as_local=False) | Q(trip__consider_as_local=True, trip__activity_batches__isnull=True))
    elif view_type == 'monthly':
        trips = trips.filter(consider_as_local=True, activity_batches__isnull=False)
        advances = advances.filter(trip__consider_as_local=True, trip__activity_batches__isnull=False)
        claims = claims.filter(trip__consider_as_local=True, trip__activity_batches__isnull=False)

    # Exclude local trips if dynamic activity batches exist (handles de-duplication in ApprovalsView)
    batch_trip_ids = {str(tid).strip() for tid in batch_qs.values_list('trip_id', flat=True) if tid}
    trips = trips.exclude(consider_as_local=True, trip_id__in=batch_trip_ids)
    
    if is_hr:
        trips = trips.exclude(consider_as_local=True, activity_batches__isnull=False)
        
    trip_count = trips.count()
    advance_count = advances.count()
    claim_count = claims.count()
    final_batch_count = batch_qs.count()

    return {
        "total": trip_count + advance_count + claim_count + final_batch_count,
        "trips": trip_count,
        "advances": advance_count,
        "claims": claim_count,
        "batches": final_batch_count
    }

class ApprovalCountView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request):
        user = getattr(request, 'custom_user', None)
        if not user:
            return Response({"count": 0})
        
        view_type = request.query_params.get('view_type', 'all')
        counts = _get_actual_pending_tasks_count(user, view_type)
        
        return Response({
            "total": counts["total"],
            "trips": counts["trips"],
            "advances": counts["advances"],
            "claims": counts["claims"],
            "batches": counts["batches"]
        })


class ApprovalsView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request):
        user = getattr(request, 'custom_user', None)
        if not user:
            return Response({"error": "User not found"}, status=401)
        
        user_role = (user.role.name.lower() if user.role else '')
        is_admin = user_role in ['admin', 'it-admin', 'superuser']
        is_finance_exec = _is_finance_executive(user)
        is_finance_head = _is_finance_head(user)
        is_hr = _is_hr(user)
        is_finance = is_finance_exec or is_finance_head
        
        tab = request.query_params.get('tab', 'pending')
        type_filter = request.query_params.get('type', 'all') 
        view_type = request.query_params.get('view_type', 'all')
        source = request.query_params.get('source')
        
        trips = Trip.objects.none()
        advances = TravelAdvance.objects.none()
        claims = TravelClaim.objects.none()
        disputes = Dispute.objects.none()
        user_step = None
        previous_approver_name = "HR"
        workflow_label = ""

        if is_hr:
            # Dedicated track for HR: Pull HR intimations and match active positions.
            trips = Trip.objects.none()
            advances = TravelAdvance.objects.none()
            claims = TravelClaim.objects.none()
            
            from .models import HRIntimation
            if tab == 'history':
                intimations = HRIntimation.objects.filter(hr_user=user, is_read=True)
            elif tab in ['completed', 'rejected']:
                intimations = HRIntimation.objects.filter(hr_user=user)
            else:
                intimations = HRIntimation.objects.filter(hr_user=user, is_read=False)
            
            if type_filter == 'trip':
                intimations = intimations.filter(trip__isnull=False)
            elif type_filter == 'advance':
                intimations = intimations.filter(advance__isnull=False)
            elif type_filter == 'claim':
                intimations = intimations.filter(claim__isnull=False)
                
            # Safe search fallback directly inside current intimations
            search_query = request.query_params.get('search')
            if search_query:
                intimations = intimations.filter(
                    Q(trip__trip_id__icontains=search_query) | 
                    Q(trip__user_name__icontains=search_query) |
                    Q(claim__trip__trip_id__icontains=search_query) |
                    Q(claim__user_name__icontains=search_query) |
                    Q(advance__trip__trip_id__icontains=search_query) |
                    Q(advance__user_name__icontains=search_query)
                ).distinct()
            
            # Map into lists to match the following iteration logic
            trip_ids = list(intimations.filter(trip__isnull=False).values_list('trip_id', flat=True))
            advance_ids = list(intimations.filter(advance__isnull=False).values_list('advance_id', flat=True))
            claim_ids = list(intimations.filter(claim__isnull=False).values_list('claim_id', flat=True))
            
            if tab == 'completed':
                completed_statuses = ['Paid', 'COMPLETED', 'Completed', 'Settled', 'Transferred']
                trips = Trip.objects.filter(trip_id__in=trip_ids, status__in=completed_statuses)
                advances = TravelAdvance.objects.filter(id__in=advance_ids, status__in=completed_statuses)
                claims = TravelClaim.objects.filter(id__in=claim_ids, status__in=completed_statuses)
            elif tab == 'rejected':
                rejected_statuses = ['Rejected', 'Rejected by Finance', 'Cancelled']
                trips = Trip.objects.filter(trip_id__in=trip_ids, status__in=rejected_statuses)
                advances = TravelAdvance.objects.filter(id__in=advance_ids, status__in=rejected_statuses)
                claims = TravelClaim.objects.filter(id__in=claim_ids, status__in=rejected_statuses)
            else:
                trips = Trip.objects.filter(trip_id__in=trip_ids)
                advances = TravelAdvance.objects.filter(id__in=advance_ids)
                claims = TravelClaim.objects.filter(id__in=claim_ids)
                
                if tab == 'history':
                    history_statuses = ['Submitted', 'Resubmitted', 'Approved', 'Rejected', 'Resolved', 'Paid', 'HR Approved', 'Manager Approved', 'COMPLETED', 'Completed', 'Settled', 'Transferred', 'Forwarded']
                    trips = trips.filter(status__in=history_statuses)
                    advances = advances.filter(status__in=history_statuses)
                    claims = claims.filter(status__in=history_statuses)
                else:
                    trips = trips.filter(status__in=['Pending', 'Submitted', 'Forwarded', 'Resubmitted', 'Manager Approved'])
                    advances = advances.filter(status__in=['Pending', 'Submitted', 'Forwarded', 'Resubmitted', 'Manager Approved', 'PENDING_HR'])
                    claims = claims.filter(status__in=['Pending', 'Submitted', 'Forwarded', 'Resubmitted', 'Manager Approved', 'PENDING_HR'])

        elif tab == 'history':
            from core.models import AuditLog # type: ignore
            # Include 'UPDATE' to capture older approvals or edits made by managers
            involved_logs = AuditLog.objects.filter(user=user, action__in=['APPROVE', 'FORWARD', 'REJECT', 'UPDATE'])
            
            trip_pks = involved_logs.filter(model_name='Trip').values_list('object_id', flat=True)
            advance_pks_raw = involved_logs.filter(model_name='TravelAdvance').values_list('object_id', flat=True)
            claim_pks_raw = involved_logs.filter(model_name='TravelClaim').values_list('object_id', flat=True)
            
            # Convert string IDs to integers for numeric primary keys
            advance_pks = [int(pk) for pk in advance_pks_raw if pk and pk.isdigit()]
            claim_pks = [int(pk) for pk in claim_pks_raw if pk and pk.isdigit()]
            
            # Filter statuses that qualify as history (completed/finalized or active submissions)
            history_statuses = ['Submitted', 'Resubmitted', 'Approved', 'Rejected', 'Resolved', 'Paid', 'HR Approved', 'Manager Approved', 'COMPLETED', 'Completed', 'Settled', 'Transferred', 'Forwarded']

            # Approver History: Items user acted upon
            trip_q = Q(trip_id__in=trip_pks)
            adv_q = Q(id__in=advance_pks)
            claim_q = Q(id__in=claim_pks)

            if is_finance:
                from .models import FinanceIntimation
                fin_trip_ids = FinanceIntimation.objects.filter(finance_user=user, is_read=True).values_list('trip_id', flat=True)
                trip_q |= Q(trip_id__in=fin_trip_ids)

            # Requester History: Items user owns that are finalized
            if not is_admin:
                trip_q |= Q(user=user, status__in=history_statuses)
                adv_q |= Q(trip__user=user, status__in=history_statuses)
                claim_q |= Q(trip__user=user, status__in=history_statuses)
                
                trips = Trip.objects.filter(trip_q).distinct()
                advances = TravelAdvance.objects.filter(adv_q).distinct()
                claims = TravelClaim.objects.filter(claim_q).distinct()
            else:
                # Admins see everything in history
                trips = Trip.objects.filter(status__in=history_statuses)
                advances = TravelAdvance.objects.filter(status__in=history_statuses)
                claims = TravelClaim.objects.filter(status__in=history_statuses)
        else:
            # Multi-Tab Workflow Logic (Action Required, Under Process, etc.)
            if tab == 'processing':
                trips = Trip.objects.filter(status='Under Process')
                advances = TravelAdvance.objects.filter(status='Under Process')
                claims = TravelClaim.objects.filter(status='Under Process')
                # For Managers, also filter by current_approver
                if not is_finance and not is_admin:
                    q = Q(current_approver=user) & Q(approver_position=user.active_position_id)
                    trips = trips.filter(q)
                    advances = advances.filter(q)
                    claims = claims.filter(q)
            elif tab == 'completed':
                completed_statuses = ['Paid', 'COMPLETED', 'Completed', 'Settled', 'Transferred']
                trips = Trip.objects.filter(status__in=completed_statuses)
                advances = TravelAdvance.objects.filter(status__in=completed_statuses)
                claims = TravelClaim.objects.filter(status__in=completed_statuses)
                if not is_finance and not is_admin:
                    trips = trips.filter(user=user)
                    advances = advances.filter(trip__user=user)
                    claims = claims.filter(trip__user=user)
            elif tab == 'rejected':
                rejected_statuses = ['Rejected', 'Rejected by Finance', 'Cancelled']
                trips = Trip.objects.filter(status__in=rejected_statuses)
                advances = TravelAdvance.objects.filter(status__in=rejected_statuses)
                claims = TravelClaim.objects.filter(status__in=rejected_statuses)
                if not is_finance and not is_admin:
                    trips = trips.filter(user=user)
                    advances = advances.filter(trip__user=user)
                    claims = claims.filter(trip__user=user)
            else:
                # DEFAULT: Action Required (Pending)
                if is_admin:
                    if source == 'hub':
                        finance_pending = ['PENDING_FINAL_RELEASE', 'Approved', 'PARTIALLY_COMPLETED']
                        trips = Trip.objects.filter(status__in=finance_pending)
                        advances = TravelAdvance.objects.filter(status__in=finance_pending)
                        claims = TravelClaim.objects.filter(status__in=finance_pending)
                    else:
                        finance_pending = ['PENDING_EXECUTIVE', 'PENDING_HEAD', 'PENDING_FINAL_RELEASE', 'REJECTED_BY_HEAD']
                        trips = Trip.objects.filter(status__in=['Pending', 'Submitted', 'Forwarded', 'Resubmitted', 'Manager Approved', 'HR Approved'] + finance_pending)
                        advances = TravelAdvance.objects.filter(status__in=['Pending', 'Submitted', 'Forwarded', 'Resubmitted', 'Manager Approved', 'HR Approved'] + finance_pending)
                        claims = TravelClaim.objects.filter(status__in=['Pending', 'Submitted', 'Forwarded', 'Resubmitted', 'Manager Approved', 'HR Approved'] + finance_pending)
                elif is_hr:
                    from .models import HRIntimation
                    hr_intimations = HRIntimation.objects.filter(hr_user=user, is_read=False)
                    
                    trip_ids = hr_intimations.exclude(trip__isnull=True).values_list('trip_id', flat=True)
                    claim_ids = hr_intimations.exclude(claim__isnull=True).values_list('claim_id', flat=True)
                    advance_ids = hr_intimations.exclude(advance__isnull=True).values_list('advance_id', flat=True)
                    
                    hr_pos_ids = user.get_active_position_identifiers()
                    q = Q(approver_position__in=hr_pos_ids) | Q(current_approver=user, approver_position__isnull=True)
                    
                    trips = Trip.objects.filter(Q(trip_id__in=trip_ids) | q, status__in=['Pending', 'Submitted', 'Forwarded', 'Resubmitted', 'Manager Approved'])
                    claims = TravelClaim.objects.filter(Q(id__in=claim_ids) | q, status__in=['Pending', 'Submitted', 'Forwarded', 'Resubmitted', 'Manager Approved', 'PENDING_HR'])
                    advances = TravelAdvance.objects.filter(Q(id__in=advance_ids) | q, status__in=['Pending', 'Submitted', 'Forwarded', 'Resubmitted', 'Manager Approved', 'PENDING_HR'])
                elif is_finance:
                    # Build a set of ALL identifiers for this finance position (from API + workflow step)
                    fin_pos_ids = _get_user_all_position_ids(user)
                    
                    # Gather all active steps for this user across all projects
                    all_user_steps = FinanceWorkflowStep.objects.filter(is_active=True).filter(
                        Q(position_id__in=fin_pos_ids) | Q(user=user)
                    )
                    
                    has_matching_step = False
                    if source == 'hub':
                        if is_finance_head or all_user_steps.filter(visibility_type__in=['FINANCE_HUB', 'BOTH']).exists():
                            has_matching_step = True
                        pending_money_statuses = ['PENDING_FINAL_RELEASE', 'Approved', 'PARTIALLY_COMPLETED']
                    else:
                        if all_user_steps.filter(visibility_type__in=['INBOX', 'BOTH']).exists():
                            has_matching_step = True
                        pending_money_statuses = ['PENDING_EXECUTIVE', 'PENDING_HEAD', 'REJECTED_BY_HEAD', 'HR Approved']
                        
                    if not has_matching_step:
                        # Return empty if they truly have no matching step or authority
                        trips = Trip.objects.none()
                        advances = TravelAdvance.objects.none()
                        claims = TravelClaim.objects.none()
                    else:
                        # Fetch Trips/Travel Requests from FinanceIntimation
                        from .models import FinanceIntimation
                        fin_intimations = FinanceIntimation.objects.filter(finance_user=user, is_read=False)
                        trip_ids = fin_intimations.values_list('trip_id', flat=True)
                        trips = Trip.objects.filter(trip_id__in=trip_ids).exclude(claim__isnull=False)
                        
                        unread_claim_ids = fin_intimations.exclude(claim__isnull=True).values_list('claim_id', flat=True)
                        unread_adv_ids = fin_intimations.exclude(advance__isnull=True).values_list('advance_id', flat=True)

                        if source == 'hub':
                            pos_q = (Q(approver_position__in=fin_pos_ids) | Q(approver_position__isnull=True))
                            advances = TravelAdvance.objects.filter(status__in=pending_money_statuses).filter(pos_q | Q(id__in=unread_adv_ids))
                            claims = TravelClaim.objects.filter(status__in=pending_money_statuses).filter(pos_q | Q(id__in=unread_claim_ids))
                        else:
                            # Position-Centric Inbox: Match against all known identifiers for this finance position OR unread intimations
                            pos_q = Q(approver_position__in=fin_pos_ids)
                            advances = TravelAdvance.objects.filter(status__in=pending_money_statuses).filter(pos_q | Q(id__in=unread_adv_ids))
                            claims = TravelClaim.objects.filter(status__in=pending_money_statuses).filter(pos_q | Q(id__in=unread_claim_ids))
                

                else:
                    # Position-centric manager query
                    manager_pos_ids = user.get_active_position_identifiers()

                    q = Q(approver_position__in=manager_pos_ids) | Q(current_approver=user, approver_position__isnull=True)
                    status_list = ['Pending', 'Submitted', 'Forwarded', 'Resubmitted']
                    trips = Trip.objects.filter(q, status__in=status_list)
                    advances = TravelAdvance.objects.filter(q, status__in=status_list)
                    claims = TravelClaim.objects.filter(q, status__in=status_list)

        
        # Apply date filter if provided
        date_query = request.query_params.get('date')
        if date_query:
            if tab == 'completed':
                # For completed payments, filter by payment_date
                trips = trips.filter(payment_date__date=date_query)
                advances = advances.filter(payment_date__date=date_query)
                claims = claims.filter(payment_date__date=date_query)
            else:
                trips = trips.filter(start_date=date_query)
                advances = advances.filter(trip__start_date=date_query)
                claims = claims.filter(trip__start_date=date_query)

        # Apply search filter if provided
        search_query = request.query_params.get('search')
        if search_query:
            trips = trips.filter(
                Q(trip_id__icontains=search_query) |
                Q(user_name__icontains=search_query) |
                Q(purpose__icontains=search_query) |
                Q(source__icontains=search_query) |
                Q(destination__icontains=search_query)
            ).distinct()
            advances = advances.filter(
                Q(trip__trip_id__icontains=search_query) |
                Q(user_name__icontains=search_query) |
                Q(purpose__icontains=search_query)
            ).distinct()
            claims = claims.filter(
                Q(trip__trip_id__icontains=search_query) |
                Q(user_name__icontains=search_query)
            ).distinct()

        # Apply is_disputed filter if provided (for Finance Hub dashboard stats)
        is_disputed = request.query_params.get('is_disputed') == 'true'
        if is_disputed:
            # Get unique trip IDs that have active disputes
            disputed_trip_ids = list(Dispute.objects.filter(status__in=['Open', 'In Review']).values_list('trip_id', flat=True))
            
            # Filter querysets - safely checking if they are QuerySets
            if hasattr(trips, 'filter'):
                trips = trips.filter(pk__in=disputed_trip_ids)
            if hasattr(advances, 'filter'):
                advances = advances.filter(trip_id__in=disputed_trip_ids)
            if hasattr(claims, 'filter'):
                claims = claims.filter(trip_id__in=disputed_trip_ids)

        # --- PERFORMANCE BOOST: Consolidate N+1 DB queries using Prefetching ---
        if hasattr(trips, 'select_related'):
            trips = trips.select_related('user', 'odometer_details').prefetch_related('expenses', 'job_reports', 'activity_batches')
        if hasattr(advances, 'select_related'):
            advances = advances.select_related('trip__user')
        if hasattr(claims, 'select_related'):
            claims = claims.select_related('trip__user').prefetch_related('trip__expenses', 'trip__job_reports', 'trip__advances')

        tasks = []
        
        # Prefetch user workflow step for permissions
        user_step = FinanceWorkflowStep.objects.filter(user=user, is_active=True).first()
        is_finance_exec = _is_finance_executive(user)
        is_finance_head = _is_finance_head(user)

        def _get_task_finance_details(u, obj_item, is_f_head):
            from travel.models import FinanceWorkflowStep, FinanceWorkflowSetting
            p_code_val = 'General'
            if isinstance(obj_item, Trip):
                p_code_val = obj_item.project_code or 'General'
            elif hasattr(obj_item, 'trip') and obj_item.trip:
                p_code_val = obj_item.trip.project_code or 'General'
                
            task_steps = FinanceWorkflowStep.objects.filter(is_active=True, project_code=p_code_val).order_by('sequence_order')
            if not task_steps.exists():
                task_steps = FinanceWorkflowStep.objects.filter(is_active=True, project_code='General').order_by('sequence_order')
                
            task_setting = FinanceWorkflowSetting.objects.filter(project_code=p_code_val).first()
            if not task_setting and p_code_val != 'General':
                task_setting = FinanceWorkflowSetting.objects.filter(project_code='General').first()
                
            is_p_flow = task_setting.is_parallel if task_setting else False
            if p_code_val and ('AP-104-MMUS' in p_code_val or 'AP-104-MMU' in p_code_val):
                is_p_flow = True
                
            task_step = None
            fin_pos_ids = _get_user_all_position_ids(u)
            
            for pos_id in fin_pos_ids:
                step_match = task_steps.filter(position_id=pos_id).first()
                if step_match:
                    task_step = step_match
                    break
            if not task_step:
                for step_match in task_steps.order_by('sequence_order'):
                    if step_match.position_id and str(step_match.position_id) in fin_pos_ids:
                        task_step = step_match
                        break
                    if step_match.position_name and str(step_match.position_name) in fin_pos_ids:
                        task_step = step_match
                        break
                    if step_match.user == u:
                        task_step = step_match
                        break
            if not task_step:
                for step_match in task_steps.order_by('sequence_order'):
                    if step_match.position_id and u in get_users_by_position(step_match.position_id):
                        task_step = step_match
                        break
                        
            if task_step:
                if is_p_flow:
                    if task_step.visibility_type in ['INBOX', 'BOTH']:
                        label = f"Management Approved to {task_step.position_name or u.name} (Parallel)"
                    else:
                        label = f"Finance Inbox to {task_step.position_name or u.name}"
                else:
                    prev_step = task_steps.filter(
                        sequence_order__lt=task_step.sequence_order,
                        is_active=True
                    ).order_by('-sequence_order').first()
                    prev_name = prev_step.position_name or (prev_step.user.name if prev_step.user else "Previous Step") if prev_step else "Management Approved"
                    label = f"{prev_name} to {task_step.position_name or u.name}"
                visibility = task_step.visibility_type
            else:
                label = "Management Approved to Finance"
                visibility = "HUB" if is_f_head else "INBOX"
                
            return task_step, is_p_flow, label, visibility

        # Multi-role handling: Inject intimation records directly so frontend knows it is read-only
        hr_intimations_map = {}
        if is_hr:
            from .models import HRIntimation
            if tab == 'history':
                intimations_query = HRIntimation.objects.filter(hr_user=user, is_read=True)
            elif tab in ['completed', 'rejected']:
                intimations_query = HRIntimation.objects.filter(hr_user=user)
            else:
                intimations_query = HRIntimation.objects.filter(hr_user=user, is_read=False)
                
            for inti in intimations_query:
                if inti.trip_id: hr_intimations_map[f"trip_{inti.trip_id}"] = inti
                if inti.advance_id: hr_intimations_map[f"advance_{inti.advance_id}"] = inti
                if inti.claim_id: hr_intimations_map[f"claim_{inti.claim_id}"] = inti

        fin_intimations_map = {}
        if is_finance:
            from .models import FinanceIntimation
            for inti in FinanceIntimation.objects.filter(finance_user=user):
                if inti.trip_id: fin_intimations_map[f"trip_{inti.trip_id}"] = inti
                if inti.claim_id: fin_intimations_map[f"claim_{inti.claim_id}"] = inti
                if inti.advance_id: fin_intimations_map[f"advance_{inti.advance_id}"] = inti

        # Collect trip IDs that have pending batches to avoid double-showing them
        # (Manager should approve the Batch card which contains specific rows)
        if is_admin:
            batch_qs = BulkActivityBatch.objects.filter(status__in=['Pending', 'Submitted', 'Forwarded', 'Resubmitted', 'Manager Approved', 'HR Approved', 'Under Process'])
        elif is_hr:
            # For HR, include batches tied to incoming HR Intimations
            intimation_trip_ids = [k.replace('trip_', '') for k in hr_intimations_map.keys() if k.startswith('trip_')]
            batch_qs = BulkActivityBatch.objects.filter(trip_id__in=intimation_trip_ids)
        elif is_finance:
            # Finance handles claims and advances only; block Bulk tour plans entirely
            batch_qs = BulkActivityBatch.objects.none()
        else:
            manager_pos_ids = user.get_active_position_identifiers()
            q = Q(approver_position__in=manager_pos_ids) | Q(current_approver=user, approver_position__isnull=True)
            batch_qs = BulkActivityBatch.objects.filter(q, status__in=['Pending', 'Submitted', 'Forwarded', 'Resubmitted', 'Manager Approved'])
        
        pending_batch_trip_ids = {str(tid).strip() for tid in batch_qs.values_list('trip_id', flat=True) if tid}

        if type_filter in ['all', 'trip']:
            for t in trips.order_by('-created_at'):
                # Handle view_type filtering for mobile (trips are generally NOT monthly)
                is_tour_plan = t.consider_as_local and t.activity_batches.exists()
                
                # Finance Hub should only show Claims and Advances for payout to avoid duplicates.
                # Trips (Requests) are just containers and are redundant in the Hub actionable queue.
                if source == 'hub':
                    continue

                # DE-DUPLICATION: If this trip is a Monthly Tour Plan, hide the Trip/Claim card 
                # if there are pending batches to keep the inbox clean globally.
                if t.consider_as_local and t.trip_id.strip() in pending_batch_trip_ids:
                    continue

                # REDUNDANCY FIX: Monthly Tour Plan Trip Containers are 100% redundant for HR review.
                # HR only acknowledges the Custom Batch Card or the final monetary Claim Card.
                if is_tour_plan and is_hr:
                    continue

                if view_type == 'special' and is_tour_plan: continue
                if view_type == 'monthly' and not is_tour_plan: continue


                # Calculate net payable and balance for Trips
                expense_sum = float(t.expenses.aggregate(s=Sum('amount'))['s'] or 0)
                advance_sum = float(t.advances.filter(status__in=['Paid', 'COMPLETED', 'Transferred']).aggregate(s=Sum('executive_approved_amount'))['s'] or 0)
                wallet_bal = float(t.user.carry_forward_balance or 0) if t.user else 0
                
                has_claim = hasattr(t, 'claim') or t.status in ['Claim Submitted', 'Settled']
                if has_claim:
                    # If a claim exists, the trip card should only show the pre-travel estimate or 0, 
                    # not the post-travel claim expenses.
                    gross = float(t.executive_approved_amount or 0)
                    cost_label = t.cost_estimate if (not t.consider_as_local and t.cost_estimate) else f"₹{gross:.2f}"
                else:
                    gross = float(t.executive_approved_amount if t.executive_approved_amount is not None else expense_sum)
                    net_payout = gross - advance_sum - wallet_bal
                    balance = net_payout - float(t.paid_amount or 0)
                    
                    cost_label = f"₹{max(0, net_payout):,.2f}"
                    if t.status == 'PARTIALLY_COMPLETED':
                        cost_label = f"₹{max(0, balance):,.2f}"
                    elif not t.consider_as_local and expense_sum == 0 and t.executive_approved_amount is None:
                        # Fallback for outstation trips without expenses yet
                        cost_label = t.cost_estimate

                is_local_form = t.consider_as_local and not t.activity_batches.exists()
                
                # Resolve intimation fields
                inti_obj = hr_intimations_map.get(f"trip_{t.trip_id}")
                if is_finance:
                    inti_obj = fin_intimations_map.get(f"trip_{t.trip_id}")
                has_inti = (inti_obj is not None)
                is_inti = has_inti and (not inti_obj.is_approval)
                
                task_step_item = None
                t_workflow_label = "Manager Approval"
                t_visibility = "INBOX"
                if is_finance:
                    t_task_step, t_is_p, t_workflow_label, t_visibility = _get_task_finance_details(user, t, is_finance_head)
                    task_step_item = t_task_step
                    if tab not in ['completed', 'rejected', 'processing', 'history']:
                        if source == 'hub':
                            if t_visibility == 'INBOX' and not is_finance_head:
                                continue
                        else:
                            if t_visibility not in ['INBOX', 'BOTH']:
                                continue

                tasks.append({
                    "id": f"TRIP-{t.trip_id}", "db_id": t.trip_id, 
                    "is_intimation": is_inti,
                    "intimation_id": inti_obj.id if has_inti else None,
                    "is_read": inti_obj.is_read if has_inti else False,
                    "can_mark_read": has_inti and not (inti_obj.is_read if has_inti else False),
                    "type": "Trip" if is_local_form else ("Monthly Tour Plan" if t.consider_as_local else "Trip"),
                    "requester": t.user.name if t.user else "Unknown", "purpose": t.purpose,
                    "status": t.status, "date": t.created_at.strftime("%b %d, %Y"),
                    "raw_date": t.created_at,
                    "hierarchy_level": t.hierarchy_level,
                    "trip_id": t.trip_id,
                    "is_local": False if is_local_form else t.consider_as_local,
                    "cost": cost_label,
                    "details": {
                        "source": t.source, "destination": t.destination, 
                        "start_date": t.start_date.strftime("%b %d, %Y"),
                        "end_date": t.end_date.strftime("%b %d, %Y"),
                        "travel_mode": t.travel_mode,
                        "composition": t.composition,
                        "vehicle_type": t.vehicle_type,
                        "purpose": t.purpose,
                        "project_code": t.project_code,
                        "total_expenses": str(expense_sum),
                        "requested_amount": str(expense_sum if t.consider_as_local else (t.cost_estimate or 0)),
                        "total_advance_taken": str(advance_sum),
                        "wallet_balance_used": str(wallet_bal),
                        "executive_approved_amount": str(max(0, balance)) if t.status == 'PARTIALLY_COMPLETED' else str(t.executive_approved_amount if t.executive_approved_amount is not None else expense_sum),
                        "previous_approver_name": "Management Approved" if is_finance else "Manager",
                        "workflow_label": t_workflow_label,
                        "job_reports": [
                            {
                                "id": jr.id,
                                "created_at": jr.created_at.strftime("%b %d, %Y"),
                                "user_name": jr.user.name if jr.user else "N/A",
                                "description": jr.description,
                                "attachment": jr.attachment,
                                "file_name": jr.file_name
                            } for jr in t.job_reports.all()
                        ],
                        "odometer": {
                            "start_reading": str(t.odometer_details.start_odo_reading) if hasattr(t, 'odometer_details') and t.odometer_details.start_odo_reading else None,
                            "start_image": decrypt_key(t.odometer_details.start_odo_image) if hasattr(t, 'odometer_details') and t.odometer_details.start_odo_image else None,
                            "end_reading": str(t.odometer_details.end_odo_reading) if hasattr(t, 'odometer_details') and t.odometer_details.end_odo_reading else None,
                            "end_image": decrypt_key(t.odometer_details.end_odo_image) if hasattr(t, 'odometer_details') and t.odometer_details.end_odo_image else None,
                        } if hasattr(t, 'odometer_details') else None,
                        "activity_batches": [
                            {
                                "id": b.id,
                                "file_name": b.file_name,
                                "status": b.status,
                                "data_json": b.data_json,
                                "created_at": b.created_at.strftime("%b %d, %Y")
                            } for b in t.activity_batches.all()
                        ],
                        "permissions": {
                            "can_edit_amount": _can_user_edit_amount(user, t, is_hr, is_finance, is_finance_exec, is_finance_head, task_step_item),
                            "visibility": t_visibility
                        }
                    },
                    "permissions": {
                        "can_edit_amount": _can_user_edit_amount(user, t, is_hr, is_finance, is_finance_exec, is_finance_head, task_step_item),
                        "visibility": t_visibility
                    }
                })
            
        if type_filter in ['all', 'advance']:
            for a in advances.order_by('-created_at'):
                # Advances for local trips are monthly, others are special
                is_tour_plan = a.trip.consider_as_local and a.trip.activity_batches.exists()
                view_type = request.query_params.get('view_type', 'all')
                if view_type == 'special' and is_tour_plan: continue
                if view_type == 'monthly' and not is_tour_plan: continue

                total_amt = a.executive_approved_amount or a.hr_approved_amount or a.requested_amount
                paid_amt = a.paid_amount or 0
                balance = float(total_amt) - float(paid_amt)
                
                cost_label = f"₹{total_amt}"
                if a.status == 'PARTIALLY_COMPLETED':
                    cost_label = f"₹{max(0, balance):.2f}"

                is_local_form = a.trip.consider_as_local and not a.trip.activity_batches.exists()
                inti_obj = hr_intimations_map.get(f"advance_{a.id}")
                if is_finance:
                    inti_obj = fin_intimations_map.get(f"advance_{a.id}")
                has_inti = (inti_obj is not None)
                is_inti = has_inti and (not inti_obj.is_approval)
                if is_finance:
                    is_inti = False
                
                task_step_item = None
                a_workflow_label = "HR Approval"
                a_visibility = "INBOX"
                if is_finance:
                    a_task_step, a_is_p, a_workflow_label, a_visibility = _get_task_finance_details(user, a, is_finance_head)
                    task_step_item = a_task_step
                    if tab not in ['completed', 'rejected', 'processing', 'history']:
                        if source == 'hub':
                            if a_visibility == 'INBOX' and not is_finance_head:
                                continue
                        else:
                            if a_visibility not in ['INBOX', 'BOTH']:
                                continue

                tasks.append({
                    "id": f"ADV-{a.id}", "db_id": a.id, 
                    "is_intimation": is_inti,
                    "intimation_id": inti_obj.id if has_inti else None,
                    "is_read": inti_obj.is_read if has_inti else False,
                    "can_mark_read": has_inti and not (inti_obj.is_read if has_inti else False),
                    "type": "Money Top-up / Advance",

                    "requester": a.trip.user.name if a.trip.user else "Unknown",
                    "purpose": f"Advance: {a.purpose}", "cost": cost_label,
                    "status": a.status, "date": (a.submitted_at or a.created_at).strftime("%b %d, %Y"),
                    "raw_date": a.submitted_at or a.created_at,
                    "hierarchy_level": a.hierarchy_level,
                    "trip_id": a.trip.trip_id,
                    "is_local": False if is_local_form else a.trip.consider_as_local,
                    "details": {
                        "source": a.trip.source,
                        "destination": a.trip.destination,
                        "requested_amount": str(a.requested_amount),
                        "hr_approved_amount": str(a.hr_approved_amount or 0),
                        "hr_remarks": a.hr_remarks or "",
                        "executive_approved_amount": str(max(0, balance)) if a.status == 'PARTIALLY_COMPLETED' else str(a.executive_approved_amount or a.hr_approved_amount or a.requested_amount),
                        "previous_approver_name": "Management Approved" if is_finance else "HR",
                        "workflow_label": a_workflow_label,
                        "reason": a.purpose,
                        "trip_destination": a.trip.destination,
                        "trip_id": a.trip.trip_id,
                        "start_date": a.trip.start_date.strftime("%b %d, %Y") if a.trip.start_date else "N/A",
                        "end_date": a.trip.end_date.strftime("%b %d, %Y") if a.trip.end_date else "N/A",
                        "paid_amount": str(a.paid_amount or 0),
                        "permissions": {
                            "can_edit_amount": _can_user_edit_amount(user, a, is_hr, is_finance, is_finance_exec, is_finance_head, task_step_item),
                            "visibility": a_visibility
                        }
                    },
                    "permissions": {
                        "can_edit_amount": _can_user_edit_amount(user, a, is_hr, is_finance, is_finance_exec, is_finance_head, task_step_item),
                        "visibility": a_visibility
                    }
                })

        if type_filter in ['all', 'expense', 'mileage']:
            for c in claims.order_by('-created_at'):
                # Handle view_type filtering for mobile
                is_tour_plan = c.trip.consider_as_local and c.trip.activity_batches.exists()
                view_type = request.query_params.get('view_type', 'all')
                if view_type == 'special' and is_tour_plan: continue
                if view_type == 'monthly' and not is_tour_plan: continue

                total_adv = c.trip.advances.filter(status='COMPLETED').aggregate(s=Sum('executive_approved_amount'))['s'] or 0
                wallet_bal = float(c.trip.user.carry_forward_balance or 0) if c.trip.user else 0
                net_payout = float(c.executive_approved_amount or c.hr_approved_amount or c.approved_amount or c.total_amount) - float(total_adv) - wallet_bal
                paid_amt = c.paid_amount or 0
                balance = max(0, net_payout) - float(paid_amt)
                
                cost_label = f"₹{max(0, net_payout):.2f}"
                if c.status == 'PARTIALLY_COMPLETED':
                    cost_label = f"₹{max(0, balance):.2f}"

                is_local_form = c.trip.consider_as_local and not c.trip.activity_batches.exists()
                inti_obj = hr_intimations_map.get(f"claim_{c.id}")
                if is_finance:
                    inti_obj = fin_intimations_map.get(f"claim_{c.id}")
                has_inti = (inti_obj is not None)
                is_inti = has_inti and (not inti_obj.is_approval)
                if is_finance:
                    is_inti = False
                
                task_step_item = None
                c_workflow_label = "HR Approval"
                c_visibility = "INBOX"
                if is_finance:
                    c_task_step, c_is_p, c_workflow_label, c_visibility = _get_task_finance_details(user, c, is_finance_head)
                    task_step_item = c_task_step
                    if tab not in ['completed', 'rejected', 'processing', 'history']:
                        if source == 'hub':
                            if c_visibility == 'INBOX' and not is_finance_head:
                                continue
                        else:
                            if c_visibility not in ['INBOX', 'BOTH']:
                                continue

                tasks.append({
                    "id": f"CLAIM-{c.id}", "db_id": c.id, 
                    "is_intimation": is_inti,
                    "intimation_id": inti_obj.id if has_inti else None,
                    "is_read": inti_obj.is_read if has_inti else False,
                    "can_mark_read": has_inti and not (inti_obj.is_read if has_inti else False),
                    "type": "Expense Claim" if is_local_form else ("Monthly Tour Plan" if c.trip.consider_as_local else "Expense Claim"),

                    "requester": c.trip.user.name if c.trip.user else "Unknown",
                    "purpose": f"Claim for {c.trip.destination}", 
                    "cost": cost_label,
                    "status": c.status, "date": (c.submitted_at or c.created_at).strftime("%b %d, %Y"),
                    "raw_date": c.submitted_at or c.created_at,
                    "hierarchy_level": c.hierarchy_level,
                    "trip_id": c.trip.trip_id,
                    "is_local": False if is_local_form else c.trip.consider_as_local,
                    "details": {
                        "is_local_travel": is_local_form,
                        "is_bulk_upload": c.trip.activity_batches.exists(),
                        "source": c.trip.source,
                        "destination": c.trip.destination,
                        "total_amount": str(c.total_amount),
                        "requested_amount": str(c.total_amount),
                        "approved_amount": str(c.approved_amount),
                        "hr_approved_amount": str(c.hr_approved_amount or 0),
                        "hr_remarks": getattr(c, "hr_remarks", ""),
                        "executive_approved_amount": str(max(0, balance)) if c.status == 'PARTIALLY_COMPLETED' else str(c.executive_approved_amount or c.hr_approved_amount or c.approved_amount or c.total_amount),
                        "previous_approver_name": "Management Approved" if is_finance else "HR",
                        "workflow_label": c_workflow_label,
                        "total_advance_taken": str(total_adv),
                        "wallet_balance_used": str(wallet_bal),
                        "net_payout": str(max(0, net_payout)),
                        "trip_id": c.trip.trip_id,
                        "start_date": c.trip.start_date.strftime("%b %d, %Y") if c.trip.start_date else "N/A",
                        "end_date": c.trip.end_date.strftime("%b %d, %Y") if c.trip.end_date else "N/A",
                        "has_deviations": c.has_deviations,
                        "deviation_summary": c.deviation_summary,
                        "planned_origin": c.planned_origin,
                        "planned_destination": c.planned_destination,
                        "paid_amount": str(c.paid_amount or 0),

                        "expenses": [
                            {
                                "id": e.id,
                                "category": e.category,
                                "date": e.date.strftime("%b %d, %Y"),
                                "amount": str(e.amount),
                                "description": e.description,
                                "status": e.status,
                                "receipt_image": decrypt_key(e.receipt_image) if e.receipt_image else "",
                                "rm_remarks": e.rm_remarks or "",
                                "hr_remarks": e.hr_remarks or "",
                                "finance_remarks": e.finance_remarks or "",
                                "allowed_amount": float(e.allowed_amount) if e.allowed_amount is not None else None,
                                "hr_selected_amount": float(e.hr_selected_amount) if e.hr_selected_amount is not None else None,
                                "hr_amount_source": e.hr_amount_source or "claimed",
                                "hr_selected_by_role": e.hr_selected_by_role or "",
                                "finance_selected_amount": float(e.finance_selected_amount) if e.finance_selected_amount is not None else None,
                                "finance_amount_source": e.finance_amount_source or "claimed",
                                "policy_note": e.policy_note or "",
                                "city_type_resolved": e.city_type_resolved or "",
                                "travel_mode": e.travel_mode,
                                "vehicle_type": e.vehicle_type,
                                "class_type": e.class_type,
                                "is_deviated": e.is_deviated or (
                                    json.loads(e.description).get('is_deviated', False) 
                                    if e.description.startswith('{') else False
                                ),
                                "deviation_reason": e.deviation_reason or (
                                    json.loads(e.description).get('deviation_reason', '') 
                                    if e.description.startswith('{') else ''
                                ),
                                "deviation_target": e.deviation_target or (
                                    json.loads(e.description).get('deviation_target', '') 
                                    if e.description.startswith('{') else ''
                                ),
                                "planned_origin": e.planned_origin or (
                                    json.loads(e.description).get('planned_origin', '') 
                                    if e.description.startswith('{') else ''
                                ),
                                "planned_destination": e.planned_destination or (
                                    json.loads(e.description).get('planned_destination', '') 
                                    if e.description.startswith('{') else ''
                                )
                            } for e in c.trip.expenses.all()
                        ],
                        "job_reports": [
                            {
                                "id": jr.id,
                                "created_at": jr.created_at.strftime("%b %d, %Y"),
                                "user_name": jr.user.name if jr.user else "N/A",
                                "description": jr.description,
                                "attachment": jr.attachment,
                                "file_name": jr.file_name
                            } for jr in c.trip.job_reports.all()
                        ],
                        "odometer": {
                            "start_reading": str(c.trip.odometer_details.start_odo_reading) if hasattr(c.trip, 'odometer_details') and c.trip.odometer_details.start_odo_reading else None,
                            "start_image": decrypt_key(c.trip.odometer_details.start_odo_image) if hasattr(c.trip, 'odometer_details') and c.trip.odometer_details.start_odo_image else None,
                            "end_reading": str(c.trip.odometer_details.end_odo_reading) if hasattr(c.trip, 'odometer_details') and c.trip.odometer_details.end_odo_reading else None,
                            "end_image": decrypt_key(c.trip.odometer_details.end_odo_image) if hasattr(c.trip, 'odometer_details') and c.trip.odometer_details.end_odo_image else None,
                        } if hasattr(c.trip, 'odometer_details') else None,
                        "permissions": {
                            "can_edit_amount": _can_user_edit_amount(user, c, is_hr, is_finance, is_finance_exec, is_finance_head, task_step_item),
                            "visibility": c_visibility
                        }
                    },
                    "permissions": {
                        "can_edit_amount": _can_user_edit_amount(user, c, is_hr, is_finance, is_finance_exec, is_finance_head, task_step_item),
                        "visibility": c_visibility
                    }
                })

        if source != 'hub':
            for d in disputes.order_by('-created_at'):
                tasks.append({
                    "id": f"DISPUTE-{d.id}", "db_id": d.id, "type": "Dispute",
                    "requester": d.raised_by.name if d.raised_by else "Unknown",
                    "purpose": f"Dispute: {d.category}", "status": d.status,
                    "date": d.created_at.strftime("%b %d, %Y"),
                    "raw_date": d.created_at
                })

        # ── NEW: Include BulkActivityBatch items pending approval ──────────────────
        # These are resubmitted batches that need manager review in the inbox.
        # Use the batch's OWN current_approver field (set at upload time), NOT trip's
        is_in_finance_workflow = FinanceWorkflowStep.objects.filter(user=user, is_active=True).exists()
        
        if tab in ['pending', 'history'] and type_filter in ['all', 'mileage']:
            history_statuses = ['Submitted', 'Resubmitted', 'Approved', 'Rejected', 'Resolved', 'Paid', 'HR Approved', 'Manager Approved', 'COMPLETED', 'Completed', 'Settled', 'Transferred', 'Forwarded']
            
            if is_hr:
                # Dedicated track for HR: Fetch batches tied to incoming intimations (read/unread already handled!)
                intimation_trip_ids = [k.replace('trip_', '') for k in hr_intimations_map.keys() if k.startswith('trip_')]
                pending_batches = BulkActivityBatch.objects.filter(trip_id__in=intimation_trip_ids)
            elif tab == 'history':
                # Approver History for Batches
                from core.models import AuditLog # type: ignore
                involved_batch_pks = AuditLog.objects.filter(user=user, model_name='BulkActivityBatch', action__in=['APPROVE', 'REJECT', 'UPDATE']).values_list('object_id', flat=True)
                pending_batches = BulkActivityBatch.objects.filter(
                    Q(id__in=involved_batch_pks) | Q(trip__user=user, status__in=history_statuses)
                ).distinct()
                
                if is_admin:
                    pending_batches = BulkActivityBatch.objects.filter(status__in=history_statuses)
            elif is_admin:
                pending_batches = BulkActivityBatch.objects.filter(status__in=['Pending', 'Submitted', 'Forwarded', 'Resubmitted', 'Manager Approved', 'HR Approved', 'Under Process'])
            elif is_finance:
                # Corporate Compliance: Finance strictly handles monetary disbursements (Claims & Advances).
                # Monthly Tour Plans (Bulk Activity Batches) are planning tools and must not show up in Finance queues.
                pending_batches = BulkActivityBatch.objects.none()
            else:
                # Include batches for regular managers where they are the current approver
                manager_pos_ids = user.get_active_position_identifiers()
                q = Q(approver_position__in=manager_pos_ids) | Q(current_approver=user, approver_position__isnull=True)
                pending_batches = BulkActivityBatch.objects.filter(
                    q,
                    status__in=['Submitted', 'Manager Approved', 'Resubmitted', 'Forwarded']
                )

            # Apply date filter to batches
            if date_query:
                # Filter by trip start date if possible, otherwise by batch created_at
                try:
                    # Correct datetime usage (module vs class)
                    from datetime import datetime as dt_class
                    target_date = dt_class.strptime(date_query, "%Y-%m-%d").date()
                    pending_batches = pending_batches.filter(
                        Q(trip__start_date=target_date) | Q(created_at__date=target_date)
                    )
                except Exception as e:
                    print(f"DEBUG: Batch date filter failed for {date_query}: {e}")
                    pass

            filtered_batches = list(pending_batches.order_by('-created_at'))

            for b in filtered_batches:
                trip_obj = b.trip
                if not trip_obj and b.trip_id:
                     trip_obj = Trip.objects.filter(trip_id=b.trip_id).first()
                     if trip_obj:
                        b.trip = trip_obj
                        b.save()
                if not trip_obj: continue

                # Calculate deviation info for the batch (from trip expenses)
                deviated_expenses = trip_obj.expenses.filter(is_deviated=True)
                has_dev = deviated_expenses.exists()
                dev_sum = ""
                if has_dev:
                    dev_sum = f"Found {deviated_expenses.count()} deviations:\n"
                    for dx in deviated_expenses:
                        dev_sum += f"- {dx.category}: {dx.deviation_target or 'N/A'} ({dx.deviation_reason or 'No reason'})\n"

                rows = b.data_json if isinstance(b.data_json, list) else []
                pending_row_count = sum(1 for r in rows if (r.get('_status') or 'Pending') not in ['Approved', 'Validated', 'OK', 'Rejected'])

                # Find which specific rows in this batch have recorded deviations
                deviated_indices = list(trip_obj.expenses.filter(
                    is_deviated=True, 
                    row_index__isnull=False
                ).values_list('row_index', flat=True))

                # Calculate total amount for this specific batch from its data_json
                batch_total = 0.0
                try:
                    for r in rows:
                        # Only sum rows that are NOT rejected
                        if r.get('_status') != 'Rejected':
                            batch_total += float(r.get('expense_amount') or 0)
                            batch_total += float(r.get('odo_total') or 0)
                            batch_total += float(r.get('incidental_amount') or 0)
                except:
                    pass

                # Resolve intimation fields for batches
                inti_obj = hr_intimations_map.get(f"trip_{trip_obj.trip_id}")
                has_inti = (inti_obj is not None)
                is_inti = has_inti and (not inti_obj.is_approval)

                tasks.append({
                    "id": f"BATCH-{b.id}", "db_id": b.id, "type": "Bulk Upload",
                    "is_intimation": is_inti,
                    "intimation_id": inti_obj.id if has_inti else None,
                    "is_read": inti_obj.is_read if has_inti else False,
                    "can_mark_read": has_inti and not (inti_obj.is_read if has_inti else False),
                    "requester": trip_obj.user.name if trip_obj.user else "Unknown",
                    "user_name": trip_obj.user.name if trip_obj.user else "Unknown",
                    "file_name": b.file_name or b.batch_name or 'Upload',
                    "row_count": pending_row_count,
                    "purpose": f"Bulk Activity: {b.file_name or b.batch_name or 'Upload'} ({pending_row_count} row(s) pending)",
                    "status": b.status, "date": b.created_at.strftime("%b %d, %Y"),
                    "raw_date": b.created_at,
                    "trip_id": trip_obj.trip_id, "is_local": True, "hierarchy_level": b.hierarchy_level, 
                    "cost": f"₹{batch_total:,.2f}",
                    "data_json": b.data_json,
                    "permissions": {
                        "can_edit_amount": _can_user_edit_amount(user, b, is_hr, is_finance, is_finance_exec, is_finance_head, user_step),
                        "visibility": user_step.visibility_type if user_step else ("HUB" if is_finance_head else "INBOX")
                    },
                    "details": {
                        "batch_id": b.id,
                        "requested_amount": str(batch_total),
                        "executive_approved_amount": str(trip_obj.executive_approved_amount if trip_obj.executive_approved_amount is not None else batch_total),
                        "previous_approver_name": previous_approver_name if is_finance else "HR",
                        "workflow_label": workflow_label if is_finance else "HR Approval",
                        "permissions": {
                            "can_edit_amount": _can_user_edit_amount(user, b, is_hr, is_finance, is_finance_exec, is_finance_head, user_step),
                            "visibility": user_step.visibility_type if user_step else ("HUB" if is_finance_head else "INBOX")
                        },
                        "file_name": b.file_name or b.batch_name,
                        "has_deviations": has_dev,
                        "deviation_summary": dev_sum.strip(),
                        "deviated_indices": deviated_indices,
                        "planned_origin": trip_obj.source,
                        "planned_destination": trip_obj.destination,
                        "source": trip_obj.source,
                        "destination": trip_obj.destination,
                        "trip_id": trip_obj.trip_id,
                        "start_date": trip_obj.start_date.strftime("%b %d, %Y") if trip_obj.start_date else "N/A",
                        "end_date": trip_obj.end_date.strftime("%b %d, %Y") if trip_obj.end_date else "N/A",
                        "activity_batches": [
                            {
                                "id": b.id,
                                "file_name": b.file_name,
                                "status": b.status,
                                "data_json": b.data_json,
                                "created_at": b.created_at.strftime("%b %d, %Y")
                            }
                        ]
                    }
                })
        # ─────────────────────────────────────────────────────────────────────────────

        # ─────────────────────────────────────────────────────────────────────────────
        
        # Cross-type Chronological Sorting: Interleave all request types by submission date (newest first)
        tasks.sort(key=lambda x: x.get('raw_date') or timezone.now(), reverse=True)
        for task in tasks: task.pop('raw_date', None)

        # Paginate results only if 'page' is requested
        if request.query_params.get('page'):
            from rest_framework.pagination import PageNumberPagination
            from django.conf import settings
            paginator = PageNumberPagination()
            
            # Respect settings.PAGE_SIZE if not provided in URL
            default_page_size = getattr(settings, 'REST_FRAMEWORK', {}).get('PAGE_SIZE', 20)
            paginator.page_size = int(request.query_params.get('page_size', default_page_size))
            
            try:
                page = paginator.paginate_queryset(tasks, request, view=self)
                if page is not None:
                    return paginator.get_paginated_response(page)
            except Exception:
                # Handle "Invalid page" gracefully
                return Response({
                    'count': len(tasks),
                    'next': None,
                    'previous': None,
                    'results': []
                })

        return Response(tasks)

    def post(self, request):
        user = getattr(request, 'custom_user', None)
        # --- DEBUG LOGGING ---
        print(f"[APPROVALS POST] Content-Type: {request.content_type}")
        print(f"[APPROVALS POST] request.data: {dict(request.data)}")
        if user:
            print(f"[APPROVALS POST] User: {user.employee_id} - {user.name}")
            print(f"[APPROVALS POST] User active_position_id: {user.active_position_id}")
            print(f"[APPROVALS POST] User position identifiers: {user.get_active_position_identifiers()}")
        else:
            print("[APPROVALS POST] User: None")
        # --- END DEBUG ---
        task_id = request.data.get('id')
        action = request.data.get('action') 
        
        if not task_id or not action:
            print(f"[APPROVALS POST] MISSING FIELDS — task_id={task_id!r} action={action!r}")
            return Response({"error": "ID and Action required"}, status=400)
            
        if action == 'MarkRead':
            try:
                from .models import HRIntimation, FinanceIntimation
                from django.utils import timezone
                
                is_fin = _is_finance_executive(user) or _is_finance_head(user)
                
                if is_fin:
                    filter_kwargs = {'finance_user': user, 'is_approval': False}
                    obj = None
                    if task_id.startswith('TRIP-'):
                        obj = Trip.objects.filter(trip_id=task_id.replace('TRIP-', '')).first()
                        if obj: filter_kwargs['trip'] = obj
                    elif task_id.startswith('CLAIM-'):
                        obj = TravelClaim.objects.filter(id=task_id.replace('CLAIM-', '')).first()
                        if obj: filter_kwargs['claim'] = obj
                    elif task_id.startswith('ADV-'):
                        obj = TravelAdvance.objects.filter(id=task_id.replace('ADV-', '')).first()
                        if obj: filter_kwargs['advance'] = obj
                    else:
                        obj = Trip.objects.filter(trip_id=task_id).first()
                        if obj: filter_kwargs['trip'] = obj
                        
                    intimations = FinanceIntimation.objects.filter(**filter_kwargs)
                    if intimations.exists():
                        intimations.update(is_read=True, read_at=timezone.now())
                        return Response({"message": "Marked as Read successfully."})
                    else:
                        return Response({"error": f"No active finance intimation record found for ID: {task_id}"}, status=404)
                else:
                    filter_kwargs = {'hr_user': user, 'is_approval': False}
                    if task_id.startswith('TRIP-'):
                        filter_kwargs['trip__trip_id'] = task_id.replace('TRIP-', '')
                    elif task_id.startswith('BATCH-'):
                        b_id = task_id.replace('BATCH-', '')
                        b_obj = BulkActivityBatch.objects.filter(id=b_id).first()
                        if b_obj and b_obj.trip:
                            filter_kwargs['trip'] = b_obj.trip
                        else:
                            filter_kwargs['id'] = -1 # Safe fallback to force no match
                    elif task_id.startswith('CLAIM-'):
                        filter_kwargs['claim_id'] = task_id.replace('CLAIM-', '')
                    elif task_id.startswith('ADV-'):
                        filter_kwargs['advance_id'] = task_id.replace('ADV-', '')
                    else:
                        # raw fallback check for plain trip_ids
                        filter_kwargs['trip__trip_id'] = task_id
                        
                    intimations = HRIntimation.objects.filter(**filter_kwargs)
                    if intimations.exists():
                        intimations.update(is_read=True, read_at=timezone.now())
                        return Response({"message": "Marked as Read successfully."})
                    else:
                        return Response({"error": f"No active HR intimation record found for ID: {task_id}"}, status=404)
            except Exception as e:
                return Response({"error": f"Failed to process action: {str(e)}"}, status=400)

        if action == 'UpdateBatchRow':
            row_index = request.data.get('row_index')
            row_status = request.data.get('row_status')
            remarks = request.data.get('remarks', '')
            print(f"DEBUG: UpdateBatchRow — id={task_id}, index={row_index}, status={row_status}")
            try:
                batches = BulkActivityBatch.objects.none()

                if task_id.startswith('TRIP-') or any(task_id.startswith(p) for p in ['ITS-', 'TRP-', 'TRV-']):
                    # ID refers to a Trip — find all activity batches for that trip
                    raw_trip_id = task_id.replace('TRIP-', '') if task_id.startswith('TRIP-') else task_id
                    trip = Trip.objects.filter(trip_id=raw_trip_id).first()
                    if trip:
                        batches = trip.activity_batches.all()
                    else:
                        return Response({"error": f"Trip not found: {raw_trip_id}"}, status=404)
                elif task_id.startswith('BATCH-'):
                    # Explicit batch ID
                    batch_id = task_id.replace('BATCH-', '')
                    batches = BulkActivityBatch.objects.filter(id=batch_id)
                else:
                    # Assume raw numeric batch ID (from Monthly Tour Plan list view)
                    batches = BulkActivityBatch.objects.filter(id=task_id)

                if not batches.exists():
                    return Response({"error": f"No activity batches found for ID: {task_id}"}, status=404)

                # Ensure trip is defined for later use
                trip = None
                if batches.exists():
                    trip = batches.first().trip

                updated_count = 0
                for batch in batches:
                    data = batch.data_json
                    if isinstance(data, list) and 0 <= int(row_index) < len(data):
                        row = data[int(row_index)]
                        row['_status'] = row_status
                        # Stamp who took this action (name + designation) into the remark
                        designation = getattr(user, 'designation', '')
                        if not designation or str(designation).lower() in ['n/a', 'na', 'employee', '']:
                            designation = getattr(user, 'department', '')
                        
                        role_str = designation if designation and str(designation).lower() not in ['n/a', 'na', ''] else (user.role.name if user.role else 'Manager')
                        user_info = f"{user.name} ({role_str})"
                        if row_status == 'Rejected':
                            row['_remarks'] = f"Rejected by {user_info}: {remarks}" if remarks else f"Rejected by {user_info}"
                        else:
                            row['_remarks'] = remarks or ''
                        
                        batch.data_json = data
                        
                        # NEW: If the batch was rejected and the user is correcting a row,
                        # set the status back to 'Resubmitted' and RESTART the flow from Level 1.
                        if batch.status == 'Rejected' and row_status == 'Validated':
                            batch.status = 'Resubmitted'
                            # Get the requester (batch owner)
                            requester = batch.user
                            current_approver, h_level, rm, sm, hod, pos_id = resolve_approver(requester)
                            batch.current_approver = current_approver
                            batch.approver_position = pos_id
                            batch.hierarchy_level = h_level
                            if batch.trip:
                                from .utils import build_approval_chain
                                batch.trip.approval_chain = build_approval_chain(requester)
                                batch.trip.status = 'Resubmitted'
                                batch.trip.current_approver = current_approver
                                batch.trip.approver_position = pos_id
                                batch.trip.hierarchy_level = h_level
                                batch.trip.save()
                            print(f"DEBUG: Resubmission restart — next_approver={current_approver}, level={batch.hierarchy_level}")
                            
                        batch.save()
                        updated_count += 1
                
                # NEW: Generate actual expense records for rows that were just validated/approved
                # This ensures the web application ledger shows them immediately.
                if trip:
                    _generate_expenses_from_batches(trip)
                
                # Notify user and manager if they corrected a row
                if row_status == 'Validated':
                    Notification.objects.create(
                        user=batch.user,
                        target_position=batch.requester_position,
                        title="Resubmission",
                        message="The rejected record has been resubmitted",
                        type='success'
                    )
                    
                    # Also notify the manager that there is a resubmitted record to review
                    if batch.current_approver:
                         Notification.objects.create(
                            user=batch.current_approver,
                            target_position=batch.approver_position,
                            title="Resubmission Received",
                            message=f"{user.name} has resubmitted a rejected record in the travel log for {batch.trip_id}.",
                            type='info'
                        )

                print(f"DEBUG: UpdateBatchRow updated {updated_count} row(s)")
                return Response({"message": f"Row {row_index} updated in {updated_count} batch(es)"})
            except Exception as e:
                import traceback
                traceback.print_exc()
                return Response({"error": str(e)}, status=400)

        if action == 'UpdateItem':
            item_id = request.data.get('item_id')
            item_status = request.data.get('item_status')
            remarks = request.data.get('remarks', '')
            if item_id:
                from .models import Expense # type: ignore
                expense = Expense.objects.filter(id=item_id).first()
                if expense:
                    # Determine which remarks field to update based on user role
                    user_role = user.role.name.upper() if user and user.role else ""
                    if "FINANCE" in user_role:
                        expense.finance_remarks = remarks
                    elif "HR" in user_role:
                        expense.hr_remarks = remarks
                    else:
                        expense.rm_remarks = remarks
                    
                    expense.status = item_status
                    expense.save()
                    
                    claim = getattr(expense.trip, 'claim', None)
                    # Recompute the effective approved total from the claim
                    # (signals.py updates approved_amount & executive_approved_amount on expense.save)
                    if claim:
                        claim.refresh_from_db()
                    return Response({
                        "message": f"Item {item_id} set to {item_status} with remarks",
                        "total_amount": float(claim.total_amount) if claim else 0.0,
                        "approved_amount": float(claim.approved_amount) if claim else 0.0,
                        "executive_approved_amount": float(claim.executive_approved_amount or 0.0) if claim else 0.0
                    })
            return Response({"error": "Item ID required or not found"}, status=400)

        try:
            if task_id.startswith('TRIP-'):
                obj = Trip.objects.get(trip_id=task_id.replace('TRIP-', ''))
            elif task_id.startswith('ADV-'):
                obj = TravelAdvance.objects.get(id=task_id.replace('ADV-', ''))
            elif task_id.startswith('CLAIM-'):
                obj = TravelClaim.objects.get(id=task_id.replace('CLAIM-', ''))
            elif task_id.startswith('DISPUTE-'):
                obj = Dispute.objects.get(id=task_id.replace('DISPUTE-', ''))
            elif task_id.startswith('BATCH-'):
                obj = BulkActivityBatch.objects.filter(id=task_id.replace('BATCH-', '')).first()
                if not obj: return Response({"error": "Batch not found"}, status=404)
            elif any(task_id.startswith(p) for p in ['ITS-', 'TRP-', 'TRV-']):
                obj = Trip.objects.get(trip_id=task_id)
            else:
                return Response({"error": f"Invalid task ID prefix: {task_id}"}, status=400)
            
            return handle_workflow_action(obj, action, user, request.data)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({"error": str(e)}, status=400)

def _clean_decimal(value):
    """
    Cleans any string containing currency symbols, spaces, or words like '(Estimated)'
    and converts it to a clean decimal number.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
        
    from decimal import Decimal
    val_str = str(value).strip()
    
    # Check if empty
    if not val_str:
        return Decimal('0.00')
        
    # Remove currency symbol and formatting commas
    val_str = val_str.replace('₹', '').replace('$', '').replace(',', '')
    
    # Handle parenthetical expressions like (Estimated) or (Approx)
    if '(' in val_str:
        val_str = val_str.split('(')[0].strip()
        
    # Filter to only allowed decimal characters: digits, dot, minus
    cleaned = ''
    for char in val_str:
        if char.isdigit() or char == '.' or char == '-':
            cleaned += char
            
    try:
        return Decimal(cleaned) if cleaned else Decimal('0.00')
    except Exception:
        return Decimal('0.00')

def handle_workflow_action(obj, action, user, data=None):
    """Standalone multi-level workflow handler for Trips, Advances, Claims, and Batches"""
    # 1. Initialization
    requester = obj.user if hasattr(obj, 'user') else (obj.trip.user if hasattr(obj, 'trip') and obj.trip else None)
    if not requester:
        raise Exception("Could not determine requester for this item.")

    request_type = "Trip" if isinstance(obj, Trip) else ("Advance" if isinstance(obj, TravelAdvance) else ("Expense Claim" if isinstance(obj, TravelClaim) else "Bulk Activity Log"))
    
    from .models import HRIntimation, FinanceIntimation, HRPositionConfig
    # Security Context
    user_role = user.role.name.lower() if user.role else ''
    is_admin = user_role in ['admin', 'it-admin', 'superuser']
    is_hr = _is_hr(user)
    is_finance_exec = _is_finance_executive(user)
    is_finance_head = _is_finance_head(user)
    is_finance = is_finance_exec or is_finance_head

    is_hr_approval_step = (
        (isinstance(obj, Trip) and HRIntimation.objects.filter(trip=obj, is_approval=True, is_read=False).exists()) or
        (isinstance(obj, TravelClaim) and HRIntimation.objects.filter(claim=obj, is_approval=True, is_read=False).exists()) or
        (isinstance(obj, TravelAdvance) and HRIntimation.objects.filter(advance=obj, is_approval=True, is_read=False).exists()) or
        (isinstance(obj, BulkActivityBatch) and obj.trip and HRIntimation.objects.filter(trip=obj.trip, is_approval=True, is_read=False).exists())
    )
    is_finance_approval_step = (
        (isinstance(obj, Trip) and FinanceIntimation.objects.filter(trip=obj, is_approval=True, is_read=False).exists()) or
        (isinstance(obj, TravelClaim) and FinanceIntimation.objects.filter(claim=obj, is_approval=True, is_read=False).exists()) or
        (isinstance(obj, TravelAdvance) and FinanceIntimation.objects.filter(advance=obj, is_approval=True, is_read=False).exists()) or
        (isinstance(obj, BulkActivityBatch) and obj.trip and FinanceIntimation.objects.filter(trip=obj.trip, is_approval=True, is_read=False).exists())
    ) if is_admin else (
        (isinstance(obj, Trip) and FinanceIntimation.objects.filter(finance_user=user, trip=obj, is_approval=True, is_read=False).exists()) or
        (isinstance(obj, TravelClaim) and FinanceIntimation.objects.filter(finance_user=user, claim=obj, is_approval=True, is_read=False).exists()) or
        (isinstance(obj, TravelAdvance) and FinanceIntimation.objects.filter(finance_user=user, advance=obj, is_approval=True, is_read=False).exists()) or
        (isinstance(obj, BulkActivityBatch) and obj.trip and FinanceIntimation.objects.filter(finance_user=user, trip=obj.trip, is_approval=True, is_read=False).exists())
    )
    # 2. Authorization Check
    if not is_admin:
        if is_hr:
            user_has_intimation = (
                (isinstance(obj, Trip) and HRIntimation.objects.filter(hr_user=user, trip=obj, is_approval=True, is_read=False).exists()) or
                (isinstance(obj, TravelClaim) and HRIntimation.objects.filter(hr_user=user, claim=obj, is_approval=True, is_read=False).exists()) or
                (isinstance(obj, TravelAdvance) and HRIntimation.objects.filter(hr_user=user, advance=obj, is_approval=True, is_read=False).exists()) or
                (isinstance(obj, BulkActivityBatch) and obj.trip and HRIntimation.objects.filter(hr_user=user, trip=obj.trip, is_approval=True, is_read=False).exists())
            )
            if not user_has_intimation:
                raise Exception("Unauthorized: Missing required HRIntimation record for this workflow action.")

        authorized = False
        # A. Position-Centric Authority
        if obj.approver_position and (str(obj.approver_position) == str(user.active_position_id) or str(obj.approver_position) in user.get_active_position_identifiers()):
            authorized = True
        # B. Fallback User Authority
        elif not obj.approver_position and obj.current_approver == user:
            authorized = True
        # C. Functional Role Authority
        elif is_hr and obj.status in ['Manager Approved', 'PENDING_HR']:
            authorized = True
        elif is_hr and isinstance(obj, Trip) and HRIntimation.objects.filter(hr_user=user, trip=obj, is_approval=True, is_read=False).exists():
            authorized = True
        elif is_hr and isinstance(obj, TravelClaim) and HRIntimation.objects.filter(hr_user=user, claim=obj, is_approval=True, is_read=False).exists():
            authorized = True
        elif is_hr and isinstance(obj, TravelAdvance) and HRIntimation.objects.filter(hr_user=user, advance=obj, is_approval=True, is_read=False).exists():
            authorized = True
        elif is_hr and isinstance(obj, BulkActivityBatch) and obj.trip and HRIntimation.objects.filter(hr_user=user, trip=obj.trip, is_approval=True, is_read=False).exists():
            authorized = True
        elif is_finance and obj.status in ['PENDING_EXECUTIVE', 'PENDING_HEAD', 'PENDING_FINAL_RELEASE', 'HR Approved', 'Approved', 'Under Process', 'REJECTED_BY_HEAD', 'PENDING_FINANCE']:
            authorized = True
        elif is_finance and is_finance_approval_step:
            authorized = True
        # D. Submission Authority (for mobile)
        elif action == 'Submit' and hasattr(obj, 'user') and obj.user == user:
            authorized = True
            
        if not authorized:
            raise Exception(f"Unauthorized: You are not the current approver for this {request_type} in this profile.")

    from core.models import AuditLog # type: ignore
    
    # 3. Action: REJECT
    if action in ['Reject', 'RejectByFinance']:
        if is_finance_head and obj.status == 'PENDING_HEAD':
            obj.status = 'REJECTED_BY_HEAD'
            obj.current_approver = obj.sent_by_executive
        else:
            obj.status = 'Rejected'
            obj.current_approver = None
            obj.approver_position = None
        
        obj.save()

        if isinstance(obj, Trip):
            for batch in obj.activity_batches.exclude(status='Rejected'):
                batch.status = 'Rejected'
                batch.current_approver = None
                batch.approver_position = None
                batch.save()
        elif isinstance(obj, BulkActivityBatch) and obj.trip:
            obj.trip.status = 'Rejected'
            obj.trip.current_approver = None
            obj.trip.approver_position = None
            obj.trip.save()
        
        # If it was an HR approval step, mark HR intimations as read
        if isinstance(obj, Trip):
            HRIntimation.objects.filter(trip=obj).update(is_read=True, read_at=timezone.now())
        elif isinstance(obj, TravelClaim):
            HRIntimation.objects.filter(claim=obj).update(is_read=True, read_at=timezone.now())
        elif isinstance(obj, TravelAdvance):
            HRIntimation.objects.filter(advance=obj).update(is_read=True, read_at=timezone.now())
        elif isinstance(obj, BulkActivityBatch) and obj.trip:
            HRIntimation.objects.filter(trip=obj.trip).update(is_read=True, read_at=timezone.now())

        reason = (data.get('remarks') or data.get('finance_remarks') or 'No reason provided') if data else 'No reason provided'
        AuditLog.objects.create(
            user=user, action='REJECT', model_name=obj.__class__.__name__,
            object_id=str(obj.pk), object_repr=str(obj), details={'reason': reason}
        )
        Notification.objects.create(
            user=requester, title=f"{request_type} Rejected",
            message=f"Your {request_type} has been rejected by {user.name}. Reason: {reason}",
            type='error'
        )
        return Response({"message": "Rejected successfully"})

    # 4. Action: UNDER PROCESS
    if action == 'UnderProcess':
        obj.status = 'Under Process'
        obj.save()
        Notification.objects.create(
            user=requester, title="Status Updated",
            message=f"Your {request_type} is now Under Process by {user.name}.",
            type='info'
        )
        return Response({"message": "Status updated to Under Process"})

    # 5. Action: APPROVE / PAY / SUBMIT / CONFIRM
    if action in ['Approve', 'Transfer', 'Pay', 'Submit', 'Confirm']:
        # LOGGING
        AuditLog.objects.create(
            user=user, action=action.upper(),
            model_name=obj.__class__.__name__, object_id=str(obj.pk), object_repr=str(obj)
        )
        
        # --- PHASE A: TRIP SUBMISSION (From Mobile) ---
        if action == 'Submit' and isinstance(obj, Trip):
            total_sum = obj.expenses.aggregate(s=Sum('amount'))['s'] or 0
            if total_sum <= 0: raise Exception("Cannot submit claim with zero expenses.")
            
            # Use standard robust approver resolution to support parallel top-level routing
            current_approver, h_level, rm, sm, hod, pos_id = resolve_approver(requester)
            is_top_level = (current_approver is None)
            
            was_rejected = TravelClaim.objects.filter(trip=obj, status='Rejected').exists()
            claim, _ = TravelClaim.objects.update_or_create(
                trip=obj,
                defaults={
                    'total_amount': total_sum,
                    'approved_amount': total_sum,
                    'status': 'Resubmitted' if was_rejected else 'Submitted',
                    'requester_position': requester.active_position_id,
                    'current_approver': current_approver,
                    'approver_position': pos_id,
                    'hierarchy_level': 0 if is_top_level else h_level,
                    'submitted_at': timezone.now(),
                    'user_name': requester.name
                }
            )
            
            # Notify Approver with context
            msg = f"{requester.name} has resubmitted a previously rejected claim for Trip {obj.trip_id}." if was_rejected else f"{requester.name} has submitted an expense claim for Trip {obj.trip_id}."
            if current_approver:
                Notification.objects.create(
                    user=current_approver, target_position=pos_id,
                    title="Expense Claim Resubmitted" if was_rejected else "New Expense Claim",
                    message=msg, type='info'
                )
            
            # Notify Requester
            Notification.objects.create(
                user=requester,
                title="Claim Resubmitted" if was_rejected else "Claim Submitted",
                message=f"Your claim for Trip {obj.trip_id} has been {'resubmitted' if was_rejected else 'sent'} to {current_approver.name if current_approver else 'Finance'} for approval.",
                type='success'
            )
            update_trip_lifecycle(obj, "Settlement", "Claim submitted for review.")
            
            if is_top_level:
                # Top-level employees instantly release their claims to Finance Configuration pipelines
                dispatch_res = trigger_parallel_dispatch(claim, user)
                return Response(dispatch_res)
            else:
                if current_approver:
                    Notification.objects.create(
                        user=current_approver, target_position=pos_id,
                        title="New Expense Claim",
                        message=f"{requester.name} has submitted an expense claim for review.",
                        type='info'
                    )
                return Response({"message": "Claim submitted successfully and routed to Manager."})

        # --- PHASE B: WORKFLOW ROUTING ---
        assigned_approver = obj.current_approver
        assigned_is_hr = _is_hr(assigned_approver) if assigned_approver else False
        assigned_is_fin = (_is_finance_executive(assigned_approver) or _is_finance_head(assigned_approver)) if assigned_approver else False
        is_functional_stage = assigned_is_hr or assigned_is_fin or obj.status in ['Manager Approved', 'PENDING_EXECUTIVE', 'PENDING_HEAD', 'HR Approved', 'PENDING_HR'] or is_hr_approval_step

        # --- STAGE 1: Management Hierarchy ---
        if not is_functional_stage and not is_hr and not is_finance:
            if action != 'Approve':
                return Response({"error": f"{action} is not supported at this stage"}, status=400)
            trip = obj if isinstance(obj, Trip) else getattr(obj, 'trip', None)
            if trip:
                if isinstance(obj, TravelClaim):
                    approved_sum = obj.trip.expenses.exclude(status='Rejected').aggregate(s=Sum('amount'))['s'] or 0
                    obj.approved_amount = approved_sum
                    obj.save()
                    update_trip_lifecycle(trip, "Management Approval", f"Approved by {user.name}. Net Approved: ₹{approved_sum}.")
                else:
                    update_trip_lifecycle(trip, "Management Approval", f"Approved by {user.name}.")
            
            next_approver = None
            next_pos_id = None
            
            # 1. Smart Next Approver Lookup using Stored Snapshot Chain (Strict Position-to-Position)
            chain = getattr(obj, 'approval_chain', []) or (obj.trip.approval_chain if hasattr(obj, 'trip') and obj.trip else [])
            if isinstance(chain, list) and len(chain) > 0:
                # Filter only Managerial steps (functional pipelines like HR/Finance are handled later)
                mgr_steps = [step for step in chain if step.get('role') == 'Manager' or not step.get('role')]
                
                # current index corresponds to hierarchy_level - 1 (i.e. Level 1 is index 0)
                # Thus next index is exactly hierarchy_level (i.e. Level 2 is index 1)
                next_idx = obj.hierarchy_level
                
                if next_idx < len(mgr_steps):
                    next_step = mgr_steps[next_idx]
                    next_emp_id = next_step.get('employee_id')
                    next_pos_id = next_step.get('position_id')
                    
                    if next_emp_id:
                        from core.models import User # type: ignore
                        next_approver = User.objects.filter(employee_id=next_emp_id).first()
            
            # 2. Legacy Fallback: Compute Next Approver on-the-fly using dynamic User-to-User hierarchy
            if not next_approver:
                assigned_approver = obj.current_approver or user
                potential_manager = assigned_approver.reporting_manager
                
                if potential_manager and potential_manager != requester and potential_manager.employee_id != assigned_approver.employee_id:
                    # SPECIAL CASE: Bypassing COO in legacy hierarchy resolution
                    pm_pos = potential_manager.get_current_position()
                    pm_pos_name = pm_pos.get('name') if pm_pos else None
                    pm_pos_id = pm_pos.get('id') if pm_pos else None
                    
                    is_pm_coo = _is_coo_position(pm_pos_name, potential_manager.designation, employee_id=potential_manager.employee_id)
                    
                    # If it's a COO, check if approval is enabled for this project
                    coo_enabled = False
                    if is_pm_coo and pm_pos_id:
                        from travel.models import COOProjectSetting
                        proj_code = requester.project_code or 'General'
                        coo_enabled = COOProjectSetting.objects.filter(project_code=proj_code, coo_position_id=str(pm_pos_id), enable_coo_approval=True).exists()
                        
                    if not is_pm_coo or coo_enabled:
                        next_approver = potential_manager
                
                if next_approver and not next_pos_id:
                    # Resolve next position ID properly from requester hierarchy snapshots
                    if next_approver == requester.reporting_manager:
                        next_pos_id = requester.reporting_manager_position
                    elif next_approver == requester.senior_manager:
                        next_pos_id = requester.senior_manager_position
                    elif next_approver == requester.hod_director:
                        next_pos_id = requester.hod_director_position
                    
                    # Stepwise fallback from the current approver's reporting structure
                    if not next_pos_id:
                        aa_pos = assigned_approver.get_current_position()
                        if aa_pos:
                            for mgr in aa_pos.get('reporting_to', []):
                                if isinstance(mgr, dict) and mgr.get('employee_code') == next_approver.employee_id:
                                    next_pos_id = mgr.get('id') or mgr.get('position_id')
                                    break

                    # Final safe lookup for managers with active_position_id profiles
                    if not next_pos_id and next_approver:
                        next_pos_id = getattr(next_approver, 'active_position_id', None)
                        if not next_pos_id:
                            next_pos_obj = next_approver.get_current_position()
                            next_pos_id = next_pos_obj.get('id') if next_pos_obj else None

            # 3. SMART WORKFLOW ADVANCEMENT (Un-nested Unified Block)
            if next_approver:
                obj.current_approver = next_approver
                obj.approver_position = str(next_pos_id) if next_pos_id else None
                obj.hierarchy_level += 1
                obj.save()
                
                if isinstance(obj, Trip):
                    for batch in obj.activity_batches.exclude(status='Rejected'):
                        batch.current_approver = next_approver
                        batch.approver_position = next_pos_id
                        batch.hierarchy_level = obj.hierarchy_level
                        if obj.status in ['Pending', 'Submitted', 'Forwarded', 'Resubmitted']:
                            batch.status = obj.status
                        batch.save()
                elif isinstance(obj, BulkActivityBatch) and obj.trip:
                    obj.trip.current_approver = next_approver
                    obj.trip.approver_position = next_pos_id
                    obj.trip.hierarchy_level = obj.hierarchy_level
                    if obj.status in ['Pending', 'Submitted', 'Forwarded', 'Resubmitted']:
                        obj.trip.status = obj.status
                    obj.trip.save()
                
                Notification.objects.create(
                    user=next_approver, target_position=next_pos_id,
                    title=f"Pending Approval: {request_type}",
                    message=f"{requester.name}'s {request_type} requires your review (Forwarded by {user.name}).",
                    type='info'
                )
                
                # Notify Requester about who approved and who it was sent to
                Notification.objects.create(
                    user=requester,
                    title=f"{request_type} Approved & Forwarded",
                    message=f"Your {request_type} was approved by {user.name} and forwarded to {next_approver.name} for the next stage.",
                    type='success'
                )
                return Response({"message": f"Approved and forwarded to {next_approver.name}"})
            else:
                # Complete Management level. Trigger parallel dispatch to HR Configuration (Mark as Read)
                # and route (Advances/Claims) to the Finance configuration pipeline.
                dispatch_result = trigger_parallel_dispatch(obj, user)
                
                # Also update status logic for connected Batch objects if applicable
                if isinstance(obj, Trip):
                    for batch in obj.activity_batches.exclude(status='Rejected'):
                        batch.status = obj.status
                        batch.current_approver = None
                        batch.approver_position = None
                        batch.save()
                elif isinstance(obj, BulkActivityBatch) and obj.trip:
                    obj.trip.status = obj.status
                    obj.trip.current_approver = None
                    obj.trip.approver_position = None
                    obj.trip.save()
                    
                return Response(dispatch_result)

        # --- STAGE 2: HR Audit ---
        is_in_hr_stage = is_hr_approval_step or assigned_is_hr or obj.status in ['Manager Approved', 'PENDING_HR']
        if (is_hr and (
            obj.status in ['Manager Approved', 'PENDING_HR'] or 
            obj.current_approver == user or 
            (isinstance(obj, Trip) and HRIntimation.objects.filter(hr_user=user, trip=obj, is_approval=True, is_read=False).exists()) or
            (isinstance(obj, TravelClaim) and HRIntimation.objects.filter(hr_user=user, claim=obj, is_approval=True, is_read=False).exists()) or
            (isinstance(obj, TravelAdvance) and HRIntimation.objects.filter(hr_user=user, advance=obj, is_approval=True, is_read=False).exists()) or
            (isinstance(obj, BulkActivityBatch) and obj.trip and HRIntimation.objects.filter(hr_user=user, trip=obj.trip, is_approval=True, is_read=False).exists())
        )) or (is_admin and is_in_hr_stage):
            # 1. Mark current HR position's intimations as read/approved
            curr_pos_id = obj.approver_position
            if not curr_pos_id and (is_admin or is_hr):
                # Fallback to resolver
                project_code = 'General'
                if isinstance(obj, Trip):
                    project_code = obj.project_code or 'General'
                elif hasattr(obj, 'trip') and obj.trip:
                    project_code = obj.trip.project_code or 'General'
                elif isinstance(obj, BulkActivityBatch) and obj.trip:
                    project_code = obj.trip.project_code or 'General'
                requester_user = obj.user if hasattr(obj, 'user') else (obj.trip.user if hasattr(obj, 'trip') and obj.trip else None)
                if (not project_code or project_code in ['General', 'N/A']) and requester_user:
                    if hasattr(requester_user, 'project_code') and requester_user.project_code and requester_user.project_code != 'N/A':
                        project_code = requester_user.project_code
                first_cfg = HRPositionConfig.objects.filter(is_active=True, project_code=project_code).order_by('sequence_order').first()
                if not first_cfg:
                    first_cfg = HRPositionConfig.objects.filter(is_active=True, project_code='General').order_by('sequence_order').first()
                curr_pos_id = first_cfg.position_id if first_cfg else None

            if isinstance(obj, Trip):
                HRIntimation.objects.filter(trip=obj, hr_position=curr_pos_id).update(is_read=True, read_at=timezone.now())
            elif isinstance(obj, TravelClaim):
                HRIntimation.objects.filter(claim=obj, hr_position=curr_pos_id).update(is_read=True, read_at=timezone.now())
            elif isinstance(obj, TravelAdvance):
                HRIntimation.objects.filter(advance=obj, hr_position=curr_pos_id).update(is_read=True, read_at=timezone.now())
            elif isinstance(obj, BulkActivityBatch) and obj.trip:
                HRIntimation.objects.filter(trip=obj.trip, hr_position=curr_pos_id).update(is_read=True, read_at=timezone.now())

            # 2. Find next HR position in sequence requiring approval
            project_code = 'General'
            if isinstance(obj, Trip):
                project_code = obj.project_code or 'General'
            elif hasattr(obj, 'trip') and obj.trip:
                project_code = obj.trip.project_code or 'General'
            elif isinstance(obj, BulkActivityBatch) and obj.trip:
                project_code = obj.trip.project_code or 'General'

            # Fallback to requester's project_code if general/empty
            requester = obj.user if hasattr(obj, 'user') else (obj.trip.user if hasattr(obj, 'trip') and obj.trip else None)
            if (not project_code or project_code in ['General', 'N/A']) and requester:
                if hasattr(requester, 'project_code') and requester.project_code and requester.project_code != 'N/A':
                    project_code = requester.project_code

            hr_positions = HRPositionConfig.objects.filter(is_active=True, project_code=project_code).order_by('sequence_order')
            if not hr_positions.exists():
                hr_positions = HRPositionConfig.objects.filter(is_active=True, project_code='General').order_by('sequence_order')

            # Save amount if HR has edit permission (or is admin)
            pos_ids = user.get_active_position_identifiers() if hasattr(user, 'get_active_position_identifiers') else [user.active_position_id]
            configs = hr_positions.filter(position_id__in=pos_ids)
            if is_admin or configs.filter(edit_claims='CAN_EDIT').exists():
                exec_amount = data.get('approved_amount') or data.get('executive_approved_amount') if data else None
                if exec_amount is not None:
                    exec_amount_cleaned = _clean_decimal(exec_amount)
                    if isinstance(obj, TravelClaim):
                        all_exps = obj.trip.expenses.filter(is_deleted=False)
                        # Only recalculate from per-expense hr_selected_amount if HR has actually
                        # used hr-decide to set line-level decisions.
                        # If none are set, respect the total HR typed in the UI.
                        has_per_expense_decisions = all_exps.filter(hr_selected_amount__isnull=False).exists()
                        if has_per_expense_decisions:
                            from decimal import Decimal as _Decimal
                            calc_sum = sum(
                                (e.finance_selected_amount if e.finance_selected_amount is not None else (e.hr_selected_amount if e.hr_selected_amount is not None else e.amount))
                                for e in all_exps if e.status != 'Rejected'
                            )
                            exec_amount_cleaned = _Decimal(str(calc_sum))
                        # else: keep exec_amount_cleaned from frontend
                    if hasattr(obj, 'hr_approved_amount'):
                        obj.hr_approved_amount = exec_amount_cleaned
                    if hasattr(obj, 'executive_approved_amount'):
                        obj.executive_approved_amount = exec_amount_cleaned

                    trip = obj if isinstance(obj, Trip) else getattr(obj, 'trip', None)
                    if isinstance(obj, BulkActivityBatch) and obj.trip:
                        trip = obj.trip

                    if trip and hasattr(trip, 'executive_approved_amount'):
                        trip.executive_approved_amount = exec_amount_cleaned
                        trip.save()
                    obj.save()

            # Check if parallel flow is enabled for this project
            from travel.models import HRWorkflowSetting
            setting = HRWorkflowSetting.objects.filter(project_code=project_code).first()
            if not setting and project_code != 'General':
                setting = HRWorkflowSetting.objects.filter(project_code='General').first()
            is_parallel_flow = setting.is_parallel if setting else False
            enable_two_level_flow = setting.enable_two_level_flow if setting else False

            is_two_level_forwarding = False
            manager_configs = None
            if is_parallel_flow and enable_two_level_flow:
                current_config = hr_positions.filter(position_id=curr_pos_id).first()
                if not current_config:
                    user_positions = user.get_active_position_identifiers() if hasattr(user, 'get_active_position_identifiers') else [user.active_position_id]
                    current_config = hr_positions.filter(position_id__in=user_positions).first()
                
                if current_config and current_config.hr_level_type == 'assistant_manager':
                    manager_configs = hr_positions.filter(hr_level_type='manager')
                    if manager_configs.exists():
                        is_two_level_forwarding = True

            if is_two_level_forwarding and manager_configs:
                # 1. Mark all Assistant Manager intimations as read/approved
                am_pos_ids = list(hr_positions.filter(hr_level_type='assistant_manager').values_list('position_id', flat=True))
                am_filter = {'hr_position__in': am_pos_ids}
                if isinstance(obj, Trip): am_filter['trip'] = obj
                elif isinstance(obj, TravelClaim): am_filter['claim'] = obj
                elif isinstance(obj, TravelAdvance): am_filter['advance'] = obj
                elif isinstance(obj, BulkActivityBatch): am_filter['trip'] = obj.trip
                HRIntimation.objects.filter(**am_filter).update(is_read=True, read_at=timezone.now())

                # 2. Forward to Manager positions
                first_manager_config = manager_configs.first()
                obj.approver_position = first_manager_config.position_id
                
                manager_users_created = []
                for m_config in manager_configs:
                    m_users = get_users_by_position(m_config.position_id)
                    for mu in m_users:
                        intimation_filter = {'hr_user': mu, 'hr_position': m_config.position_id}
                        if isinstance(obj, Trip): intimation_filter['trip'] = obj
                        elif isinstance(obj, TravelClaim): intimation_filter['claim'] = obj
                        elif isinstance(obj, TravelAdvance): intimation_filter['advance'] = obj
                        elif isinstance(obj, BulkActivityBatch): intimation_filter['trip'] = obj.trip

                        is_appr = False
                        if isinstance(obj, Trip) and m_config.trips_approval == 'APPROVAL':
                            is_appr = True
                        elif isinstance(obj, BulkActivityBatch) and m_config.bulk_approval == 'APPROVAL':
                            is_appr = True
                        elif isinstance(obj, (TravelClaim, TravelAdvance)) and m_config.claims_approval == 'APPROVAL':
                            is_appr = True

                        is_read_only = False
                        if isinstance(obj, Trip) and m_config.trips_approval == 'MARK_READ':
                            is_read_only = True
                        elif isinstance(obj, BulkActivityBatch) and m_config.bulk_approval == 'MARK_READ':
                            is_read_only = True
                        elif isinstance(obj, (TravelClaim, TravelAdvance)) and m_config.claims_approval == 'MARK_READ':
                            is_read_only = True

                        if is_appr or is_read_only:
                            if not HRIntimation.objects.filter(**intimation_filter).exists():
                                HRIntimation.objects.create(is_approval=is_appr, **intimation_filter)
                            if mu not in manager_users_created:
                                manager_users_created.append(mu)

                if manager_users_created:
                    obj.current_approver = manager_users_created[0]
                else:
                    obj.current_approver = None
                obj.save()

                if isinstance(obj, Trip) and hasattr(obj, 'activity_batches'):
                    obj.activity_batches.exclude(status='Rejected').update(
                        current_approver=obj.current_approver, approver_position=obj.approver_position
                    )
                elif isinstance(obj, BulkActivityBatch) and obj.trip:
                    obj.trip.current_approver = obj.current_approver
                    obj.trip.approver_position = obj.approver_position
                    obj.trip.save()

                for mu in manager_users_created:
                    Notification.objects.create(
                        user=mu, target_position=obj.approver_position,
                        title="Pending HR Manager Approval",
                        message=f"{requester.name}'s {request_type} requires Manager HR approval.",
                        type='info'
                    )

                return Response({"message": f"HR Assistant Manager approved. Request forwarded to HR Managers."})

            next_approval_pos = None
            if not is_parallel_flow:
                current_config = hr_positions.filter(position_id=curr_pos_id).first()
                if current_config:
                    next_configs = hr_positions.filter(sequence_order__gt=current_config.sequence_order)
                    for hr_pos in next_configs:
                        is_appr_required = False
                        if isinstance(obj, Trip) and hr_pos.trips_approval == 'APPROVAL':
                            is_appr_required = True
                        elif isinstance(obj, BulkActivityBatch) and hr_pos.bulk_approval == 'APPROVAL':
                            is_appr_required = True
                        elif isinstance(obj, (TravelClaim, TravelAdvance)) and hr_pos.claims_approval == 'APPROVAL':
                            is_appr_required = True
                        
                        if is_appr_required:
                            next_approval_pos = hr_pos
                            break

            if next_approval_pos:
                # Forward to next HR position
                obj.approver_position = next_approval_pos.position_id
                next_users = get_users_by_position(next_approval_pos.position_id)
                if next_users:
                    obj.current_approver = next_users[0]
                else:
                    obj.current_approver = None
                obj.save()

                if isinstance(obj, Trip) and hasattr(obj, 'activity_batches'):
                    obj.activity_batches.exclude(status='Rejected').update(
                        current_approver=obj.current_approver, approver_position=obj.approver_position
                    )
                elif isinstance(obj, BulkActivityBatch) and obj.trip:
                    obj.trip.current_approver = obj.current_approver
                    obj.trip.approver_position = obj.approver_position
                    obj.trip.save()

                # Create intimations for next position
                for n_user in next_users:
                    intimation_filter = {'hr_user': n_user, 'hr_position': next_approval_pos.position_id}
                    if isinstance(obj, Trip): intimation_filter['trip'] = obj
                    elif isinstance(obj, TravelClaim): intimation_filter['claim'] = obj
                    elif isinstance(obj, TravelAdvance): intimation_filter['advance'] = obj
                    elif isinstance(obj, BulkActivityBatch): intimation_filter['trip'] = obj.trip

                    if not HRIntimation.objects.filter(**intimation_filter).exists():
                        HRIntimation.objects.create(is_approval=True, **intimation_filter)

                    Notification.objects.create(
                        user=n_user, target_position=next_approval_pos.position_id,
                        title="Pending HR Approval",
                        message=f"{requester.name}'s {request_type} requires your approval (Forwarded).",
                        type='info'
                    )

                return Response({"message": f"HR Approval registered. Forwarded to {next_approval_pos.position_name}."})

            # If no next approval position, finalize HR audit stage
            if isinstance(obj, BulkActivityBatch):
                obj.status = 'Approved'
                obj.current_approver = None
                obj.save()
                if obj.trip:
                    dispatch_result = trigger_finance_workflow(obj.trip, user)
                    HRIntimation.objects.filter(
                        Q(trip=obj.trip) & (Q(hr_user=user) | Q(is_approval=True))
                    ).update(is_read=True, read_at=timezone.now())
                    msg = f"HR Approval completed. {dispatch_result['message']}"
                else:
                    msg = "HR Approval completed."
                notify_hr(f"{request_type} Approved", f"{requester.name}'s {request_type} has been approved by HR.")
                return Response({"message": msg})

            if isinstance(obj, Trip):
                # TRIPS route to Finance
                dispatch_result = trigger_finance_workflow(obj, user)
                
                # Mark HR intimations as read/approved (preserving acknowledgement-only intimations for other HRs)
                HRIntimation.objects.filter(
                    Q(trip=obj) & (Q(hr_user=user) | Q(is_approval=True))
                ).update(is_read=True, read_at=timezone.now())
                
                notify_hr(f"Trip Approved", f"{requester.name}'s trip has been approved by HR.")
                return Response({"message": f"HR Approval completed. {dispatch_result['message']}"})

            # Claims and Advances proceed to Finance Workflow
            obj.status = 'HR Approved'
            
            # Mark HR intimations as read/approved for Claims and Advances (preserving other HRs' acknowledgement-only intimations)
            if isinstance(obj, TravelClaim):
                HRIntimation.objects.filter(
                    Q(claim=obj) & (Q(hr_user=user) | Q(is_approval=True))
                ).update(is_read=True, read_at=timezone.now())
            elif isinstance(obj, TravelAdvance):
                HRIntimation.objects.filter(
                    Q(advance=obj) & (Q(hr_user=user) | Q(is_approval=True))
                ).update(is_read=True, read_at=timezone.now())
            
            # Sequence handover to Finance Workflow
            dispatch_result = trigger_finance_workflow(obj, user)
            msg = "HR Audit completed successfully."
            if dispatch_result and 'message' in dispatch_result:
                msg = f"HR Audit completed successfully. {dispatch_result['message']}"
            
            if isinstance(obj, BulkActivityBatch) and obj.trip:
                 obj.trip.status = 'HR Approved'
                 obj.trip.current_approver = obj.current_approver
                 obj.trip.approver_position = obj.approver_position
                 obj.trip.save()
                 # SIDE EFFECT: Generate activities from this specific batch
                 _generate_expenses_from_batches(obj.trip)
                 # Update row statuses
                 updated_rows = []
                 for row in (obj.data_json or []):
                     if row.get('_status') != 'Rejected': row['_status'] = 'Approved'
                     updated_rows.append(row)
                 obj.data_json = updated_rows
                 obj.save()

            notify_hr(f"{request_type} HR Approved", f"{requester.name}'s {request_type} has been HR-Audited.")
            return Response({"message": msg})

        # --- STAGE 3: Finance Verification & Payout ---
        is_any_finance_approval_pending = (
            (isinstance(obj, Trip) and FinanceIntimation.objects.filter(trip=obj, is_approval=True, is_read=False).exists()) or
            (isinstance(obj, TravelClaim) and FinanceIntimation.objects.filter(claim=obj, is_approval=True, is_read=False).exists()) or
            (isinstance(obj, TravelAdvance) and FinanceIntimation.objects.filter(advance=obj, is_approval=True, is_read=False).exists()) or
            (isinstance(obj, BulkActivityBatch) and obj.trip and FinanceIntimation.objects.filter(trip=obj.trip, is_approval=True, is_read=False).exists())
        )
        if is_finance or assigned_is_fin or obj.status in ['PENDING_EXECUTIVE', 'PENDING_HEAD', 'PENDING_FINAL_RELEASE', 'HR Approved', 'REJECTED_BY_HEAD', 'PENDING_FINANCE'] or is_finance_approval_step or (is_admin and is_any_finance_approval_pending):
            trip = obj if isinstance(obj, Trip) else getattr(obj, 'trip', None)

            # 1. Action: REJECT
            if action in ['Reject', 'RejectByFinance']:
                obj.status = 'Rejected'
                obj.current_approver = None
                obj.approver_position = None
                obj.save()

                # Mark all intimations as read
                if isinstance(obj, Trip):
                    HRIntimation.objects.filter(trip=obj).update(is_read=True, read_at=timezone.now())
                    FinanceIntimation.objects.filter(trip=obj).update(is_read=True, read_at=timezone.now())
                    for batch in obj.activity_batches.exclude(status='Rejected'):
                        batch.status = 'Rejected'
                        batch.current_approver = None
                        batch.save()
                elif isinstance(obj, TravelClaim):
                    HRIntimation.objects.filter(claim=obj).update(is_read=True, read_at=timezone.now())
                    FinanceIntimation.objects.filter(claim=obj).update(is_read=True, read_at=timezone.now())
                elif isinstance(obj, TravelAdvance):
                    HRIntimation.objects.filter(advance=obj).update(is_read=True, read_at=timezone.now())
                    FinanceIntimation.objects.filter(advance=obj).update(is_read=True, read_at=timezone.now())
                elif isinstance(obj, BulkActivityBatch):
                    if obj.trip:
                        HRIntimation.objects.filter(trip=obj.trip).update(is_read=True, read_at=timezone.now())
                        FinanceIntimation.objects.filter(trip=obj.trip).update(is_read=True, read_at=timezone.now())

                notify_hr(f"{request_type} Rejected", f"{requester.name}'s {request_type} has been rejected by Finance.")
                return Response({"message": f"{request_type} rejected by Finance."})

            # 2. Action: TRANSFER / PAY
            if action in ['Transfer', 'Pay']:
                # Payout Logic
                payment_mode = data.get('payment_mode', 'BANK') if data else 'BANK'
                transaction_id = data.get('transaction_id', 'N/A') if data else 'N/A'
                from decimal import Decimal
                if isinstance(obj, TravelClaim):
                    all_exps = obj.trip.expenses.filter(is_deleted=False)
                    calc_sum = sum(
                        (e.finance_selected_amount if e.finance_selected_amount is not None else (e.hr_selected_amount if e.hr_selected_amount is not None else e.amount))
                        for e in all_exps if e.status != 'Rejected'
                    )
                    raw_amt = calc_sum
                    obj.executive_approved_amount = Decimal(str(raw_amt))
                    obj.approved_amount = Decimal(str(raw_amt))
                else:
                    raw_amt = getattr(obj, 'executive_approved_amount', getattr(obj, 'approved_amount', getattr(obj, 'requested_amount', 0)))
                amount = Decimal(str(raw_amt or 0))
                
                net_payable = amount
                if isinstance(obj, TravelClaim) and trip and trip.user:
                    adv_sum = trip.advances.filter(status='COMPLETED').aggregate(s=Sum('executive_approved_amount'))['s'] or 0
                    total_advances = Decimal(str(adv_sum))
                    
                    net_payable = amount - total_advances
                    user_profile = trip.user
                    if net_payable < Decimal('0'):
                        surplus = abs(net_payable)
                        current_cf = Decimal(str(user_profile.carry_forward_balance or 0))
                        user_profile.carry_forward_balance = current_cf + surplus
                        user_profile.save(update_fields=['carry_forward_balance'])
                        net_payable = Decimal('0')
                        payment_mode = 'Internal Adjustment'
                        transaction_id = f"RECON-{timezone.now().strftime('%Y%m%d%H%M')}"
                
                obj.status = 'COMPLETED' if not isinstance(obj, TravelClaim) else 'Paid'
                obj.payment_date = timezone.now()
                obj.payment_mode = payment_mode
                obj.transaction_id = transaction_id
                obj.paid_amount = net_payable
                obj.processed_by = user
                obj.save()
                
                # Mark associated HR/Finance intimations as read upon successful payout
                if isinstance(obj, TravelClaim):
                    HRIntimation.objects.filter(claim=obj).update(is_read=True, read_at=timezone.now())
                    FinanceIntimation.objects.filter(claim=obj).update(is_read=True, read_at=timezone.now())
                elif isinstance(obj, TravelAdvance):
                    HRIntimation.objects.filter(advance=obj).update(is_read=True, read_at=timezone.now())
                    FinanceIntimation.objects.filter(advance=obj).update(is_read=True, read_at=timezone.now())
                
                if trip:
                    if isinstance(obj, TravelClaim):
                        trip.status = 'Settled'
                        trip.save()
                        update_trip_lifecycle(trip, "Payment Processed", f"Claim payment of ₹{net_payable} processed.")
                    elif isinstance(obj, TravelAdvance):
                        update_trip_lifecycle(trip, "Advance Processed", f"Advance payment of ₹{net_payable} processed.")

                Notification.objects.create(
                    user=requester, title="Payment Successful",
                    message=f"Your {request_type} payment of ₹{net_payable} has been processed.",
                    type='success'
                )
                return Response({"message": "Payment processed successfully"})

            # 3. Action: APPROVE / CONFIRM
            if action in ['Approve', 'HRApprove', 'Confirm']:
                # Resolve steps and settings
                fin_steps, is_parallel_flow, enable_two_level_flow, project_code = _get_finance_steps_and_settings(obj)

                # Get user's current position configs in Finance
                pos_ids = user.get_active_position_identifiers() if hasattr(user, 'get_active_position_identifiers') else [user.active_position_id]
                curr_pos_id = obj.approver_position

                # ---------------------------------------------------------------
                # Find the active finance step config for the current approver.
                # Priority:
                #   1. Match by obj.approver_position (most reliable when set)
                #   2. Match by the user's intimation that was just read (best for
                #      sequential flows where approver_position may lag or be null)
                #   3. Fallback: match any step by the user's position IDs
                # ---------------------------------------------------------------
                current_config = None
                if curr_pos_id:
                    current_config = fin_steps.filter(position_id=curr_pos_id).first()
                if not current_config and obj.current_approver:
                    current_config = fin_steps.filter(user=obj.current_approver).first()

                if not current_config:
                    # Look up the intimation that this user just approved/read
                    _inti_base = {}
                    if isinstance(obj, Trip): _inti_base['trip'] = obj
                    elif isinstance(obj, TravelClaim): _inti_base['claim'] = obj
                    elif isinstance(obj, TravelAdvance): _inti_base['advance'] = obj
                    elif isinstance(obj, BulkActivityBatch): _inti_base['trip'] = obj.trip

                    just_read = FinanceIntimation.objects.filter(
                        finance_user=user, is_read=True, **_inti_base
                    ).order_by('-read_at').first()

                    if just_read:
                        if just_read.finance_position:
                            current_config = fin_steps.filter(position_id=just_read.finance_position).first()
                        else:
                            current_config = fin_steps.filter(user=just_read.finance_user).first()

                if not current_config:
                    # Last resort: match by any of the user's active position IDs
                    current_config = fin_steps.filter(position_id__in=pos_ids).first()
                if not current_config:
                    current_config = fin_steps.filter(user=user).first()

                    
                is_finance_hub = False
                if current_config and current_config.visibility_type in ['FINANCE_HUB', 'BOTH']:
                    is_finance_hub = True
                else:
                    matching_steps = fin_steps.filter(Q(position_id__in=pos_ids) | Q(user=user))
                    if matching_steps.filter(visibility_type__in=['FINANCE_HUB', 'BOTH']).exists():
                        is_finance_hub = True

                # A. Mark current Finance position's intimations as read/approved
                # If it is a finance hub step, mark all finance hub intimations as read/approved!
                if is_finance_hub:
                    hub_pos_ids = list(fin_steps.filter(visibility_type__in=['FINANCE_HUB', 'BOTH']).values_list('position_id', flat=True))
                    inti_filter = models.Q(finance_position__in=hub_pos_ids)
                    hub_users = list(fin_steps.filter(visibility_type__in=['FINANCE_HUB', 'BOTH'], user__isnull=False).values_list('user_id', flat=True))
                    if hub_users:
                        inti_filter |= models.Q(finance_user_id__in=hub_users)
                        
                    if isinstance(obj, Trip):
                        FinanceIntimation.objects.filter(inti_filter, trip=obj).update(is_read=True, read_at=timezone.now())
                    elif isinstance(obj, TravelClaim):
                        FinanceIntimation.objects.filter(inti_filter, claim=obj).update(is_read=True, read_at=timezone.now())
                    elif isinstance(obj, TravelAdvance):
                        FinanceIntimation.objects.filter(inti_filter, advance=obj).update(is_read=True, read_at=timezone.now())
                    elif isinstance(obj, BulkActivityBatch) and obj.trip:
                        FinanceIntimation.objects.filter(inti_filter, trip=obj.trip).update(is_read=True, read_at=timezone.now())
                else:
                    if is_admin:
                        if isinstance(obj, Trip):
                            FinanceIntimation.objects.filter(trip=obj, is_read=False).update(is_read=True, read_at=timezone.now())
                        elif isinstance(obj, TravelClaim):
                            FinanceIntimation.objects.filter(claim=obj, is_read=False).update(is_read=True, read_at=timezone.now())
                        elif isinstance(obj, TravelAdvance):
                            FinanceIntimation.objects.filter(advance=obj, is_read=False).update(is_read=True, read_at=timezone.now())
                        elif isinstance(obj, BulkActivityBatch) and obj.trip:
                            FinanceIntimation.objects.filter(trip=obj.trip, is_read=False).update(is_read=True, read_at=timezone.now())
                    else:
                        if isinstance(obj, Trip):
                            FinanceIntimation.objects.filter(finance_user=user, trip=obj).update(is_read=True, read_at=timezone.now())
                        elif isinstance(obj, TravelClaim):
                            FinanceIntimation.objects.filter(finance_user=user, claim=obj).update(is_read=True, read_at=timezone.now())
                        elif isinstance(obj, TravelAdvance):
                            FinanceIntimation.objects.filter(finance_user=user, advance=obj).update(is_read=True, read_at=timezone.now())
                        elif isinstance(obj, BulkActivityBatch) and obj.trip:
                            FinanceIntimation.objects.filter(finance_user=user, trip=obj.trip).update(is_read=True, read_at=timezone.now())

                # B. Edit amount if allowed
                if current_config and current_config.can_edit_amount:
                    exec_amount = data.get('approved_amount') or data.get('executive_approved_amount')
                    if exec_amount is not None:
                        exec_amount_cleaned = _clean_decimal(exec_amount)
                        if isinstance(obj, TravelClaim):
                            all_exps = obj.trip.expenses.filter(is_deleted=False)
                            has_fin_decisions = all_exps.filter(finance_selected_amount__isnull=False).exists()
                            if has_fin_decisions:
                                from decimal import Decimal as _Decimal
                                calc_sum = sum(
                                    (e.finance_selected_amount if e.finance_selected_amount is not None else (e.hr_selected_amount if e.hr_selected_amount is not None else e.amount))
                                    for e in all_exps if e.status != 'Rejected'
                                )
                                exec_amount_cleaned = _Decimal(str(calc_sum))
                        if hasattr(obj, 'executive_approved_amount'): obj.executive_approved_amount = exec_amount_cleaned
                        if trip: trip.executive_approved_amount = exec_amount_cleaned; trip.save()

                # C. Check two level forwarding AM -> Manager
                is_two_level_forwarding = False
                manager_steps = None
                if is_parallel_flow and enable_two_level_flow:
                    if current_config and current_config.finance_level_type == 'assistant_manager' and not is_finance_hub:
                        manager_steps = fin_steps.filter(finance_level_type='manager')
                        if manager_steps.exists():
                            is_two_level_forwarding = True

                if is_two_level_forwarding and manager_steps:
                    # Mark all AM-level intimations as read
                    am_pos_ids = list(fin_steps.filter(finance_level_type='assistant_manager').values_list('position_id', flat=True))
                    am_filter = {'finance_position__in': am_pos_ids}
                    if isinstance(obj, Trip): am_filter['trip'] = obj
                    elif isinstance(obj, TravelClaim): am_filter['claim'] = obj
                    elif isinstance(obj, TravelAdvance): am_filter['advance'] = obj
                    elif isinstance(obj, BulkActivityBatch): am_filter['trip'] = obj.trip
                    FinanceIntimation.objects.filter(**am_filter).update(is_read=True, read_at=timezone.now())

                    first_manager_step = manager_steps.filter(visibility_type='INBOX').first() or manager_steps.first()
                    obj.approver_position = first_manager_step.position_id

                    # Dispatch all Manager-level INBOX steps (NOT Finance Hub yet)
                    manager_users_created = []
                    for m_step in manager_steps.filter(visibility_type='INBOX'):
                        m_users = []
                        if m_step.position_id:
                            m_users = get_users_by_position(m_step.position_id)
                        elif m_step.user:
                            m_users = [m_step.user]

                        for mu in m_users:
                            intimation_filter = {'finance_user': mu, 'finance_position': m_step.position_id or ''}
                            if isinstance(obj, Trip): intimation_filter['trip'] = obj
                            elif isinstance(obj, TravelClaim): intimation_filter['claim'] = obj
                            elif isinstance(obj, TravelAdvance): intimation_filter['advance'] = obj
                            elif isinstance(obj, BulkActivityBatch): intimation_filter['trip'] = obj.trip

                            is_appr = (m_step.trip_control == 'APPROVAL') if isinstance(obj, (Trip, BulkActivityBatch)) else True

                            if not FinanceIntimation.objects.filter(**intimation_filter).exists():
                                FinanceIntimation.objects.create(is_approval=is_appr, is_read=False, **intimation_filter)
                            if mu not in manager_users_created:
                                manager_users_created.append(mu)

                    if manager_users_created:
                        obj.current_approver = manager_users_created[0]
                    else:
                        obj.current_approver = None
                    obj.save()

                    if isinstance(obj, Trip) and hasattr(obj, 'activity_batches'):
                        obj.activity_batches.exclude(status='Rejected').update(
                            current_approver=obj.current_approver, approver_position=obj.approver_position
                        )
                    elif isinstance(obj, BulkActivityBatch) and obj.trip:
                        obj.trip.current_approver = obj.current_approver
                        obj.trip.approver_position = obj.approver_position
                        obj.trip.save()

                    for mu in manager_users_created:
                        Notification.objects.create(
                            user=mu, target_position=obj.approver_position,
                            title="Pending Finance Manager Approval",
                            message=f"{requester.name}'s {request_type} requires Manager Finance approval.",
                            type='info'
                        )

                    return Response({"message": f"Finance Assistant Manager approved. Request forwarded to Finance Managers."})

                # D-2. Parallel-only: after any inbox approval, dispatch Finance Hub steps
                if is_parallel_flow and not enable_two_level_flow and not is_finance_hub:
                    # Mark all inbox intimations as read (this person approved, clear others)
                    inbox_pos_ids = list(fin_steps.filter(visibility_type='INBOX').values_list('position_id', flat=True))
                    inbox_user_ids = list(fin_steps.filter(visibility_type='INBOX', user__isnull=False).values_list('user_id', flat=True))
                    inbox_inti_filter = {}
                    if isinstance(obj, Trip): inbox_inti_filter['trip'] = obj
                    elif isinstance(obj, TravelClaim): inbox_inti_filter['claim'] = obj
                    elif isinstance(obj, TravelAdvance): inbox_inti_filter['advance'] = obj
                    elif isinstance(obj, BulkActivityBatch): inbox_inti_filter['trip'] = obj.trip
                    FinanceIntimation.objects.filter(
                        models.Q(finance_position__in=inbox_pos_ids) | models.Q(finance_user_id__in=inbox_user_ids),
                        **inbox_inti_filter
                    ).update(is_read=True, read_at=timezone.now())

                    # Now dispatch Finance Hub steps
                    hub_steps = fin_steps.filter(visibility_type__in=['FINANCE_HUB', 'BOTH'])
                    hub_users_dispatched = []
                    for h_step in hub_steps:
                        h_users = get_users_by_position(h_step.position_id) if h_step.position_id else ([h_step.user] if h_step.user else [])
                        for hu in h_users:
                            inti_filter = {'finance_user': hu, 'finance_position': h_step.position_id or ''}
                            if isinstance(obj, Trip): inti_filter['trip'] = obj
                            elif isinstance(obj, TravelClaim): inti_filter['claim'] = obj
                            elif isinstance(obj, TravelAdvance): inti_filter['advance'] = obj
                            elif isinstance(obj, BulkActivityBatch): inti_filter['trip'] = obj.trip
                            if not FinanceIntimation.objects.filter(**inti_filter).exists():
                                FinanceIntimation.objects.create(is_approval=True, is_read=False, **inti_filter)
                                hub_users_dispatched.append(hu)
                                Notification.objects.create(
                                    user=hu, target_position=h_step.position_id,
                                    title=f"Finance Hub: {request_type} Pending Release",
                                    message=f"{requester.name}'s {request_type} has cleared inbox approval and needs your Finance Hub action.",
                                    type='info'
                                )

                    if hub_users_dispatched:
                        obj.approver_position = hub_steps.first().position_id if hub_steps.exists() else obj.approver_position
                        obj.current_approver = hub_users_dispatched[0]
                        obj.status = 'PENDING_FINAL_RELEASE'
                        obj.save()
                        return Response({"message": "Finance Inbox approved. Forwarded to Finance Hub for final release."})
                    else:
                        # No Finance Hub steps — fall through to finalize below
                        pass

                # D-3. 2-Level: Manager approved — dispatch Finance Hub steps
                if is_parallel_flow and enable_two_level_flow and not is_finance_hub:
                    if current_config and current_config.finance_level_type == 'manager':
                        # Mark all manager intimations as read
                        mgr_pos_ids = list(fin_steps.filter(finance_level_type='manager').values_list('position_id', flat=True))
                        mgr_filter = {'finance_position__in': mgr_pos_ids}
                        if isinstance(obj, Trip): mgr_filter['trip'] = obj
                        elif isinstance(obj, TravelClaim): mgr_filter['claim'] = obj
                        elif isinstance(obj, TravelAdvance): mgr_filter['advance'] = obj
                        elif isinstance(obj, BulkActivityBatch): mgr_filter['trip'] = obj.trip
                        FinanceIntimation.objects.filter(**mgr_filter).update(is_read=True, read_at=timezone.now())

                        hub_steps = fin_steps.filter(visibility_type__in=['FINANCE_HUB', 'BOTH'])
                        hub_users_dispatched = []
                        for h_step in hub_steps:
                            h_users = get_users_by_position(h_step.position_id) if h_step.position_id else ([h_step.user] if h_step.user else [])
                            for hu in h_users:
                                inti_filter = {'finance_user': hu, 'finance_position': h_step.position_id or ''}
                                if isinstance(obj, Trip): inti_filter['trip'] = obj
                                elif isinstance(obj, TravelClaim): inti_filter['claim'] = obj
                                elif isinstance(obj, TravelAdvance): inti_filter['advance'] = obj
                                elif isinstance(obj, BulkActivityBatch): inti_filter['trip'] = obj.trip
                                if not FinanceIntimation.objects.filter(**inti_filter).exists():
                                    FinanceIntimation.objects.create(is_approval=True, is_read=False, **inti_filter)
                                    hub_users_dispatched.append(hu)
                                    Notification.objects.create(
                                        user=hu, target_position=h_step.position_id,
                                        title=f"Finance Hub: {request_type} Pending Release",
                                        message=f"{requester.name}'s {request_type} has cleared Manager approval and needs Finance Hub action.",
                                        type='info'
                                    )

                        if hub_users_dispatched:
                            obj.approver_position = hub_steps.first().position_id if hub_steps.exists() else obj.approver_position
                            obj.current_approver = hub_users_dispatched[0]
                            obj.status = 'PENDING_FINAL_RELEASE'
                            obj.save()
                            return Response({"message": "Finance Manager approved. Forwarded to Finance Hub for final release."})

                # D. Next steps (sequential flow)
                next_blocking_step = None
                passed_read_steps = []
                if not is_parallel_flow:
                    is_claim_or_advance = isinstance(obj, (TravelClaim, TravelAdvance))
                    if current_config:
                        subsequent_steps = fin_steps.filter(sequence_order__gt=current_config.sequence_order).order_by('sequence_order')
                    else:
                        # Fallback: if we couldn't resolve current_config directly,
                        # look at already read intimations to find the highest completed step
                        _inti_base = {}
                        if isinstance(obj, Trip): _inti_base['trip'] = obj
                        elif isinstance(obj, TravelClaim): _inti_base['claim'] = obj
                        elif isinstance(obj, TravelAdvance): _inti_base['advance'] = obj
                        elif isinstance(obj, BulkActivityBatch): _inti_base['trip'] = obj.trip

                        read_intimations = FinanceIntimation.objects.filter(is_read=True, **_inti_base)
                        max_sequence = 0
                        for ri in read_intimations:
                            step = None
                            if ri.finance_position:
                                step = fin_steps.filter(position_id=ri.finance_position).first()
                            else:
                                step = fin_steps.filter(user=ri.finance_user).first()
                            if step and step.sequence_order > max_sequence:
                                max_sequence = step.sequence_order

                        if max_sequence > 0:
                            subsequent_steps = fin_steps.filter(sequence_order__gt=max_sequence).order_by('sequence_order')
                        else:
                            subsequent_steps = fin_steps.order_by('sequence_order')

                    for step in subsequent_steps:
                        is_appr_step = (step.trip_control == 'APPROVAL') if not is_claim_or_advance else True
                        if is_appr_step:
                            next_blocking_step = step
                            break
                        else:
                            passed_read_steps.append(step)

                # Dispatch any bypassed read steps
                for step in passed_read_steps:
                    step_users = get_users_by_position(step.position_id) if step.position_id else ([step.user] if step.user else [])
                    for su in step_users:
                        intimation_filter = {'finance_user': su, 'finance_position': step.position_id or ''}
                        if isinstance(obj, Trip): intimation_filter['trip'] = obj
                        elif isinstance(obj, TravelClaim): intimation_filter['claim'] = obj
                        elif isinstance(obj, TravelAdvance): intimation_filter['advance'] = obj
                        elif isinstance(obj, BulkActivityBatch): intimation_filter['trip'] = obj.trip

                        if not FinanceIntimation.objects.filter(**intimation_filter).exists():
                            FinanceIntimation.objects.create(is_approval=False, is_read=False, **intimation_filter)
                            
                            Notification.objects.create(
                                user=su, target_position=step.position_id,
                                title="Pending Finance Action (Notification)",
                                message=f"{requester.name}'s {request_type} requires your finance review (Forwarded).",
                                type='info'
                            )

                if not is_parallel_flow and next_blocking_step:
                    obj.approver_position = next_blocking_step.position_id
                    if isinstance(obj, (Trip, BulkActivityBatch)):
                        obj.status = 'PENDING_FINAL_RELEASE' if next_blocking_step.visibility_type == 'FINANCE_HUB' else 'PENDING_FINANCE'
                    else:
                        obj.status = 'PENDING_FINAL_RELEASE' if next_blocking_step.visibility_type == 'FINANCE_HUB' else 'PENDING_EXECUTIVE'
                    
                    next_users = []
                    if next_blocking_step.position_id:
                        next_users = get_users_by_position(next_blocking_step.position_id)
                    elif next_blocking_step.user:
                        next_users = [next_blocking_step.user]

                    if next_users:
                        obj.current_approver = next_users[0]
                    else:
                        obj.current_approver = None
                    obj.save()

                    if isinstance(obj, Trip) and hasattr(obj, 'activity_batches'):
                        obj.activity_batches.exclude(status='Rejected').update(
                            current_approver=obj.current_approver, approver_position=obj.approver_position, status=obj.status
                        )
                    elif isinstance(obj, BulkActivityBatch) and obj.trip:
                        obj.trip.current_approver = obj.current_approver
                        obj.trip.approver_position = obj.approver_position
                        obj.trip.status = obj.status
                        obj.trip.save()

                    for n_user in next_users:
                        intimation_filter = {'finance_user': n_user, 'finance_position': next_blocking_step.position_id or ''}
                        if isinstance(obj, Trip): intimation_filter['trip'] = obj
                        elif isinstance(obj, TravelClaim): intimation_filter['claim'] = obj
                        elif isinstance(obj, TravelAdvance): intimation_filter['advance'] = obj
                        elif isinstance(obj, BulkActivityBatch): intimation_filter['trip'] = obj.trip

                        if not FinanceIntimation.objects.filter(**intimation_filter).exists():
                            FinanceIntimation.objects.create(is_approval=True, is_read=False, **intimation_filter)
                            
                            Notification.objects.create(
                                user=n_user, target_position=next_blocking_step.position_id,
                                title="Pending Finance Approval",
                                message=f"{requester.name}'s {request_type} requires your finance approval (Forwarded).",
                                type='info'
                            )

                    next_name = next_blocking_step.position_name or (obj.current_approver.name if obj.current_approver else f"Position {obj.approver_position}")
                    return Response({"message": f"Finance Action registered. Forwarded to {next_name}."})

                # E. Finalize workflow
                unread_filter = {'is_approval': True, 'is_read': False}
                if isinstance(obj, Trip): unread_filter['trip'] = obj
                elif isinstance(obj, TravelClaim): unread_filter['claim'] = obj
                elif isinstance(obj, TravelAdvance): unread_filter['advance'] = obj
                elif isinstance(obj, BulkActivityBatch): unread_filter['trip'] = obj.trip

                unread_approvals = FinanceIntimation.objects.filter(**unread_filter).exists()
                if unread_approvals:
                    return Response({"message": "Finance Approval registered. Awaiting other finance approvals."})

                if isinstance(obj, (Trip, BulkActivityBatch)):
                    active_trip = obj if isinstance(obj, Trip) else obj.trip
                    finalize_trip_approval(active_trip)
                    FinanceIntimation.objects.filter(trip=active_trip).update(is_read=True, read_at=timezone.now())
                    HRIntimation.objects.filter(trip=active_trip).update(is_read=True, read_at=timezone.now())
                    notify_hr("Trip Approved", f"{requester.name}'s trip has been finalized by Finance.")
                    return Response({"message": "Finance Approval completed. Trip finalized."})
                else:
                    obj.status = 'PENDING_FINAL_RELEASE'
                    obj.final_executive = user
                    obj.current_approver = None
                    obj.approver_position = None
                    obj.save()
                    
                    if isinstance(obj, TravelClaim):
                        FinanceIntimation.objects.filter(claim=obj).update(is_read=True, read_at=timezone.now())
                    elif isinstance(obj, TravelAdvance):
                        FinanceIntimation.objects.filter(advance=obj).update(is_read=True, read_at=timezone.now())

                    return Response({"message": "Finance Approval completed. Pending final payout release."})

    return Response({"error": f"Action '{action}' is not allowed for this {request_type} in its current status ({obj.status})."}, status=400)


class ExpenseViewSet(viewsets.ModelViewSet):
    queryset = Expense.objects.all()
    serializer_class = ExpenseSerializer
    permission_classes = [IsCustomAuthenticated]
    http_method_names = ['get', 'post', 'patch', 'put', 'delete', 'head', 'options']

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return Expense.objects.none()
            
        role_name = user.active_role.lower() if hasattr(user, 'active_role') else ''
        is_admin = role_name in ['admin', 'superuser']
        is_finance = 'finance' in role_name or role_name == 'cfo'
        is_manager = role_name == 'reporting_authority'
        
        if is_admin or is_finance:
            queryset = self.queryset
        elif is_manager:
            # Allow managers to see their own expenses OR expenses for trips they are/were responsible for
            # We use snapshot names for efficient filtering as dynamic hierarchy lookup is too slow for list views
            queryset = self.queryset.filter(
                Q(trip__user=user) | 
                Q(trip__current_approver=user) |
                Q(trip__reporting_manager_name=user.name) |
                Q(trip__senior_manager_name=user.name) |
                Q(trip__hod_director_name=user.name)
            )
        else:
            queryset = self.queryset.filter(trip__user=user)
            
        trip_id = self.request.query_params.get('trip_id')
        if trip_id:
            trip_id = decode_id(trip_id)
            queryset = queryset.filter(trip__trip_id=trip_id)
        return queryset

    def perform_create(self, serializer):
        user = getattr(self.request, 'custom_user', None)
        trip = serializer.validated_data.get('trip')
        if trip and trip.user != user:
            raise serializers.ValidationError("Unauthorized trip association")
        serializer.save()

class PolicyDocumentViewSet(viewsets.ModelViewSet):
    queryset = PolicyDocument.objects.all()
    serializer_class = PolicyDocumentSerializer
    permission_classes = [IsCustomAuthenticated]

    def get_serializer_class(self):
        if self.action == 'retrieve':
            return PolicyDocumentDetailSerializer
        return PolicyDocumentSerializer

    def get_queryset(self):
        # Always return a fresh queryset; returning the class-level queryset
        # directly can reuse a cached empty result across requests.
        return PolicyDocument.objects.all()

    def perform_create(self, serializer):
        user = getattr(self.request, 'custom_user', None)
        serializer.save(uploaded_by=user)

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            from core.permissions import IsAdmin # type: ignore
            return [IsAdmin()]
        return [IsCustomAuthenticated()]

class TravelClaimViewSet(viewsets.ModelViewSet):
    queryset = TravelClaim.objects.all()
    serializer_class = TravelClaimSerializer
    permission_classes = [IsCustomAuthenticated]

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return TravelClaim.objects.none()
            
        is_admin = hasattr(user, 'role') and user.role.name.lower() in ['admin', 'superuser']
        is_hr = _is_hr(user)
        is_finance = _is_finance_executive(user) or _is_finance_head(user)
        
        if is_admin or is_hr or is_finance:
            queryset = self.queryset
        else:
            q = (
                (Q(trip__user=user) & Q(requester_position=user.active_position_id)) |
                Q(current_approver=user) |
                Q(approver_position=user.active_position_id) |
                Q(trip__current_approver=user) |
                Q(trip__approver_position=user.active_position_id)
            )
            queryset = self.queryset.filter(q)
            
        trip_id = self.request.query_params.get('trip_id')
        if trip_id:
            trip_id = decode_id(trip_id)
            queryset = queryset.filter(trip__trip_id=trip_id)
        return queryset

    def perform_create(self, serializer):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            raise serializers.ValidationError("Authentication failed. Please log in again.")
            
        trip = serializer.validated_data.get('trip')
        if trip and trip.user != user:
            raise serializers.ValidationError("Unauthorized trip association")
        
        current_approver, h_level, rm, sm, hod, pos_id = resolve_approver(user)
        is_top_level = (current_approver is None)
        total_expense_sum = trip.expenses.aggregate(s=Sum('amount'))['s'] or 0

        claim = serializer.save(
            status='Submitted',
            requester_position=user.active_position_id,
            current_approver=current_approver,
            approver_position=pos_id,
            hierarchy_level=0 if is_top_level else h_level,
            submitted_at=timezone.now(),
            total_amount=total_expense_sum,
            approved_amount=total_expense_sum,
            # Populate snapshots for resilience
            user_name=user.name,
            user_designation=user.designation,
            user_department=user.department,
            reporting_manager_name=rm.name if rm else None,
            senior_manager_name=sm.name if sm else None,
            hod_director_name=hod.name if hod else None
        )
        
        # Update trip status as requested by user
        if trip:
            trip.status = 'Claim Submitted'
            trip.save(update_fields=['status'])
        
        if is_top_level:
            # Instantly trigger parallel HR intimation + Finance Workflow routing for top-level employees
            trigger_parallel_dispatch(claim, user)
            rm_pos = rm.get_current_position() if rm else None
            rm_pos_name = rm_pos.get('name') if rm_pos else None
            if rm and _is_coo_position(rm_pos_name, rm.designation, employee_id=rm.employee_id):
                update_trip_lifecycle(trip, "Settlement", "Claim submitted. Reporting COO bypassed and routed to Finance/HR.")
        else:
            if current_approver:
                is_resubmitted = (claim.status == 'Resubmitted')
                Notification.objects.create(
                    user=current_approver, target_position=pos_id,
                    title="Expense Claim Resubmitted" if is_resubmitted else "New Expense Claim",
                    message=f"{user.name} has {'resubmitted' if is_resubmitted else 'submitted'} an expense claim for Trip {claim.trip.trip_id}.",
                    type='info'
                )
                
                # Notify Requester
                Notification.objects.create(
                    user=user,
                    title="Claim Resubmitted" if is_resubmitted else "Claim Submitted",
                    message=f"Your claim for Trip {claim.trip.trip_id} has been sent to {current_approver.name} for approval.",
                    type='success'
                )
            
        # Notify HR
        notify_hr("New Expense Claim", f"{user.name} has submitted an expense claim for Trip {claim.trip.trip_id}.")
            
        self._update_trip_lifecycle(claim.trip)

    def perform_update(self, serializer):
        user = getattr(self.request, 'custom_user', None)
        claim = self.get_object()
        is_hr = _is_hr(user)
        is_finance = _is_finance_executive(user) or _is_finance_head(user)
        is_admin = hasattr(user, 'role') and user.role.name.lower() in ['admin', 'superuser']
        
        has_edit_claim_perm = False
        if user and hasattr(user, 'role') and user.role:
            from django.db.models import Q
            from core.models import Role
            role_from_api = getattr(user, 'role_from_api', None)
            designation = getattr(user, 'designation', None)
            matching_role = None
            if role_from_api or designation:
                q_obj = Q()
                if role_from_api:
                    q_obj |= Q(name__iexact=role_from_api)
                if designation:
                    q_obj |= Q(name__iexact=designation)
                matching_role = Role.objects.filter(q_obj).first()
            user_role_obj = matching_role or user.role
            if user_role_obj and isinstance(user_role_obj.permissions, dict):
                has_edit_claim_perm = (
                    user_role_obj.permissions.get('can_edit_submitted_claim', False) or
                    user_role_obj.permissions.get('can_edit_claims', False)
                )

        if is_admin or is_finance or has_edit_claim_perm:
            pass
        elif is_hr:
            from .models import HRPositionConfig
            project_code = claim.trip.project_code or 'General' if claim.trip else 'General'
            # Fallback to requester's project_code if general/empty
            requester = claim.trip.user if claim.trip else None
            if (not project_code or project_code in ['General', 'N/A']) and requester:
                if hasattr(requester, 'project_code') and requester.project_code and requester.project_code != 'N/A':
                    project_code = requester.project_code

            configs = _get_matching_hr_configs(user, project_code=project_code)
            if not configs.exists():
                configs = _get_matching_hr_configs(user, project_code='General')
            
            can_edit = configs.filter(edit_claims='CAN_EDIT').exists()
            if not can_edit:
                raise serializers.ValidationError("Permission Denied: Your HR position does not have permission to edit claims.")
        elif claim.trip.user != user:
            raise serializers.ValidationError("Permission Denied: Unauthorized edit attempt.")
            
        claim = serializer.save()
        total_amount = claim.trip.expenses.aggregate(s=Sum('amount'))['s'] or 0
        claim.total_amount = total_amount
        claim.approved_amount = total_amount
        claim.save()
        
        if claim.status == 'Submitted':
            self._update_trip_lifecycle(claim.trip)

    def _update_trip_lifecycle(self, trip):
        update_trip_lifecycle(trip, "Settlement", "Travel reimbursement claim submitted for review.")

    @action(detail=True, methods=['get'], url_path='compute-allowance', permission_classes=[IsCustomAuthenticated])
    def compute_allowance(self, request, pk=None):
        """
        GET /api/claims/{id}/compute-allowance/
        Returns per-expense entitlement comparison for HR approval.
        Called when HR opens a claim to see Claimed vs Allowed amounts.
        """
        try:
            claim = self.get_object()
        except Exception:
            return Response({'error': 'Claim not found.'}, status=status.HTTP_404_NOT_FOUND)

        from travel_masters.eligibility import compute_allowance_for_claim
        result = compute_allowance_for_claim(claim)

        # Persist resolved city_type and allowed_amount on each expense line
        # so HR can see them even without re-calling this endpoint
        if 'expense_allowances' in result:
            for ea in result['expense_allowances']:
                exp_id = ea.get('expense_id')
                if exp_id:
                    try:
                        from .models import Expense as ExpenseModel
                        exp = ExpenseModel.objects.get(pk=exp_id)
                        exp.city_type_resolved = ea.get('city_type')
                        if ea.get('allowed_amount') is not None:
                            exp.allowed_amount = ea['allowed_amount']
                        exp.policy_note = ea.get('policy_note', '')[:255]
                        exp.save(update_fields=['city_type_resolved', 'allowed_amount', 'policy_note'])
                    except Exception:
                        pass

        return Response(result)

    @action(detail=True, methods=['patch'], url_path='hr-decide', permission_classes=[IsCustomAuthenticated])
    def hr_decide(self, request, pk=None):
        """
        PATCH /api/claims/{id}/hr-decide/
        Records HR's per-expense decisions (use claimed / use allowed / manual amount).
        Body: { "expense_decisions": [ { expense_id, hr_selected_amount, source } ] }
        Constraint: hr_selected_amount must not exceed the original claimed amount.
        """
        from .models import Expense as ExpenseModel
        user = getattr(request, 'custom_user', None)
        
        has_edit_claim_perm = False
        if user and hasattr(user, 'role') and user.role:
            from django.db.models import Q
            from core.models import Role
            role_from_api = getattr(user, 'role_from_api', None)
            designation = getattr(user, 'designation', None)
            matching_role = None
            if role_from_api or designation:
                q_obj = Q()
                if role_from_api:
                    q_obj |= Q(name__iexact=role_from_api)
                if designation:
                    q_obj |= Q(name__iexact=designation)
                matching_role = Role.objects.filter(q_obj).first()
            user_role_obj = matching_role or user.role
            if user_role_obj and isinstance(user_role_obj.permissions, dict):
                has_edit_claim_perm = (
                    user_role_obj.permissions.get('can_edit_submitted_claim', False) or
                    user_role_obj.permissions.get('can_edit_claims', False)
                )

        if not (_is_hr(user) or (hasattr(user, 'role') and user.role.name.lower() in ['admin', 'superuser']) or has_edit_claim_perm):
            return Response({'error': 'Only HR, Admin or authorized roles can record claim decisions.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            claim = self.get_object()
        except Exception:
            return Response({'error': 'Claim not found.'}, status=status.HTTP_404_NOT_FOUND)

        decisions = request.data.get('expense_decisions', [])
        if not decisions:
            return Response({'error': 'expense_decisions list is required.'}, status=status.HTTP_400_BAD_REQUEST)

        errors = []
        updated = 0
        final_total = 0.0

        for dec in decisions:
            exp_id = dec.get('expense_id')
            hr_amt = dec.get('hr_selected_amount')
            source = dec.get('source')  # 'claimed' | 'allowed' | 'manual'

            if not exp_id or hr_amt is None:
                errors.append({'expense_id': exp_id, 'error': 'expense_id and hr_selected_amount are required'})
                continue

            try:
                exp = ExpenseModel.objects.get(pk=exp_id, trip=claim.trip)
            except ExpenseModel.DoesNotExist:
                errors.append({'expense_id': exp_id, 'error': 'Expense not found for this claim'})
                continue

            try:
                hr_amt_dec = float(hr_amt)
            except (ValueError, TypeError):
                errors.append({'expense_id': exp_id, 'error': 'Invalid amount'})
                continue

            # Enforce cap: hr_selected_amount must not exceed original claimed amount
            claimed = float(exp.amount)
            if hr_amt_dec > claimed:
                errors.append({
                    'expense_id': exp_id,
                    'error': f'hr_selected_amount ({hr_amt_dec}) cannot exceed claimed amount ({claimed})'
                })
                continue

            if hr_amt_dec < 0:
                errors.append({
                    'expense_id': exp_id,
                    'error': 'Amount cannot be negative'
                })
                continue

            allowed_amount = exp.allowed_amount
            if allowed_amount is None:
                from travel_masters.eligibility import compute_allowance_for_claim
                allowance_data = compute_allowance_for_claim(claim)
                for ea in allowance_data.get('expense_allowances', []):
                    if ea.get('expense_id') == exp.id:
                        allowed_amount = ea.get('allowed_amount')
                        exp.allowed_amount = allowed_amount
                        exp.city_type_resolved = ea.get('city_type')
                        exp.policy_note = ea.get('policy_note', '')[:255]
                        exp.save(update_fields=['allowed_amount', 'city_type_resolved', 'policy_note'])
                        break

            if allowed_amount is not None:
                allowed = float(allowed_amount)
                if claimed > allowed:
                    if hr_amt_dec < allowed:
                        errors.append({
                            'expense_id': exp_id,
                            'error': f'For over-limit claims, approved amount must be at least the policy limit (₹{allowed})'
                        })
                        continue
                    policy_note_val = dec.get('policy_note') or dec.get('note') or ''
                    if not policy_note_val or not str(policy_note_val).strip():
                        errors.append({
                            'expense_id': exp_id,
                            'error': 'Policy note / justification is required for over-limit claims'
                        })
                        continue

            role_name = 'HR'
            if _is_hr(user):
                role_name = 'HR'
            elif user and hasattr(user, 'role') and user.role and user.role.name.lower() in ['admin', 'superuser']:
                role_name = 'Admin'
            elif user:
                claim_user = claim.trip.user if claim.trip else None
                if claim_user:
                    chain = claim_user.get_effective_reporting_chain()
                    approver_emp_code = str(user.employee_id or '').strip().lower()
                    approver_positions = []
                    if hasattr(user, 'get_active_position_identifiers'):
                        approver_positions = [str(pid).strip() for pid in user.get_active_position_identifiers() if pid]
                    if user.active_position_id:
                        approver_positions.append(str(user.active_position_id).strip())

                    matched_role = None
                    for idx, candidate_role in enumerate(['Reporting Manager', 'Senior Manager', 'HOD']):
                        if len(chain) > idx:
                            mgr_info = chain[idx]
                            if mgr_info:
                                if isinstance(mgr_info, dict):
                                    pos_id = str(mgr_info.get('id') or mgr_info.get('position_id') or '').strip()
                                    emp_code = str(mgr_info.get('employee_code') or mgr_info.get('employee_id') or '').strip().lower()
                                    if pos_id and pos_id in approver_positions:
                                        matched_role = candidate_role
                                        break
                                    if emp_code and emp_code == approver_emp_code:
                                        matched_role = candidate_role
                                        break
                                else:
                                    if str(mgr_info).strip().lower() == approver_emp_code:
                                        matched_role = candidate_role
                                        break
                    if matched_role:
                        role_name = matched_role
                    else:
                        role_name = 'Reporting Manager'
                else:
                    role_name = 'Manager'

            exp.hr_selected_amount = hr_amt_dec
            exp.hr_amount_source = source or 'manual'
            exp.hr_selected_by_role = role_name
            if dec.get('policy_note') is not None:
                exp.policy_note = dec.get('policy_note')[:255]
            exp.save(update_fields=['hr_selected_amount', 'hr_amount_source', 'hr_selected_by_role', 'policy_note'])
            final_total += hr_amt_dec
            updated += 1

        if errors and updated == 0:
            return Response({'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

        # Update claim's approved_amount to the sum of HR-decided line items
        if updated > 0:
            # Recalculate: sum hr_selected_amount where set, else use expense.amount (exclude Rejected)
            all_exps = claim.trip.expenses.filter(is_deleted=False)
            total = sum(
                float(e.hr_selected_amount) if e.hr_selected_amount is not None else float(e.amount)
                for e in all_exps if e.status != 'Rejected'
            )
            from decimal import Decimal as _Dec
            total_dec = _Dec(str(total))
            claim.approved_amount = total_dec
            claim.hr_approved_amount = total_dec
            # Also propagate to executive_approved_amount so Finance inbox reflects HR edits
            claim.executive_approved_amount = total_dec
            claim.save(update_fields=['approved_amount', 'hr_approved_amount', 'executive_approved_amount'])
            # Mirror on the trip so trip-level queries are consistent
            trip = claim.trip
            if trip and hasattr(trip, 'executive_approved_amount'):
                trip.executive_approved_amount = total_dec
                trip.save(update_fields=['executive_approved_amount'])

        return Response({
            'updated': updated,
            'final_approved_total': round(total, 2) if updated > 0 else None,
            'errors': errors
        })

    @action(detail=True, methods=['patch'], url_path='finance-decide', permission_classes=[IsCustomAuthenticated])
    def finance_decide(self, request, pk=None):
        """
        PATCH /api/claims/{id}/finance-decide/
        Records Finance's per-expense decisions (use claimed / use allowed / manual amount).
        Body: { "expense_decisions": [ { expense_id, finance_selected_amount, source, note } ] }
        """
        from .models import Expense as ExpenseModel
        user = getattr(request, 'custom_user', None)
        is_fin = _is_finance_executive(user) or _is_finance_head(user)
        is_admin = hasattr(user, 'role') and user.role.name.lower() in ['admin', 'it-admin', 'superuser']
        
        has_edit_claim_perm = False
        if user and hasattr(user, 'role') and user.role:
            from django.db.models import Q
            from core.models import Role
            role_from_api = getattr(user, 'role_from_api', None)
            designation = getattr(user, 'designation', None)
            matching_role = None
            if role_from_api or designation:
                q_obj = Q()
                if role_from_api:
                    q_obj |= Q(name__iexact=role_from_api)
                if designation:
                    q_obj |= Q(name__iexact=designation)
                matching_role = Role.objects.filter(q_obj).first()
            user_role_obj = matching_role or user.role
            if user_role_obj and isinstance(user_role_obj.permissions, dict):
                has_edit_claim_perm = (
                    user_role_obj.permissions.get('can_edit_submitted_claim', False) or
                    user_role_obj.permissions.get('can_edit_claims', False)
                )

        if not (is_fin or is_admin or has_edit_claim_perm):
            return Response({'error': 'Only Finance, Admin or authorized roles can record finance claim decisions.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            claim = self.get_object()
        except Exception:
            return Response({'error': 'Claim not found.'}, status=status.HTTP_404_NOT_FOUND)

        decisions = request.data.get('expense_decisions', [])
        if not decisions:
            return Response({'error': 'expense_decisions list is required.'}, status=status.HTTP_400_BAD_REQUEST)

        errors = []
        updated = 0
        final_total = 0.0

        for dec in decisions:
            exp_id = dec.get('expense_id')
            fin_amt = dec.get('finance_selected_amount') if dec.get('finance_selected_amount') is not None else dec.get('hr_selected_amount')
            source = dec.get('source')  # 'claimed' | 'allowed' | 'manual'
            remarks = dec.get('note') or dec.get('remarks') or dec.get('policy_note') or ''

            if not exp_id or fin_amt is None:
                errors.append({'expense_id': exp_id, 'error': 'expense_id and finance_selected_amount are required'})
                continue

            try:
                exp = ExpenseModel.objects.get(pk=exp_id, trip=claim.trip)
            except ExpenseModel.DoesNotExist:
                errors.append({'expense_id': exp_id, 'error': 'Expense not found for this claim'})
                continue

            try:
                fin_amt_dec = float(fin_amt)
            except (ValueError, TypeError):
                errors.append({'expense_id': exp_id, 'error': 'Invalid amount'})
                continue

            # Enforce cap: finance_selected_amount must not exceed original claimed amount
            claimed = float(exp.amount)
            if fin_amt_dec > claimed:
                errors.append({
                    'expense_id': exp_id,
                    'error': f'finance_selected_amount ({fin_amt_dec}) cannot exceed claimed amount ({claimed})'
                })
                continue

            if fin_amt_dec < 0:
                errors.append({
                    'expense_id': exp_id,
                    'error': 'Amount cannot be negative'
                })
                continue

            allowed_amount = exp.allowed_amount
            if allowed_amount is None:
                from travel_masters.eligibility import compute_allowance_for_claim
                allowance_data = compute_allowance_for_claim(claim)
                for ea in allowance_data.get('expense_allowances', []):
                    if ea.get('expense_id') == exp.id:
                        allowed_amount = ea.get('allowed_amount')
                        exp.allowed_amount = allowed_amount
                        exp.city_type_resolved = ea.get('city_type')
                        exp.policy_note = ea.get('policy_note', '')[:255]
                        exp.save(update_fields=['allowed_amount', 'city_type_resolved', 'policy_note'])
                        break

            if allowed_amount is not None:
                allowed = float(allowed_amount)
                if claimed > allowed:
                    if fin_amt_dec < allowed:
                        errors.append({
                            'expense_id': exp_id,
                            'error': f'For over-limit claims, approved amount must be at least the policy limit (₹{allowed})'
                        })
                        continue
                    if not remarks or not str(remarks).strip():
                        errors.append({
                            'expense_id': exp_id,
                            'error': 'Policy note / justification is required for over-limit claims'
                        })
                        continue

            exp.finance_selected_amount = fin_amt_dec
            exp.finance_amount_source = source or 'manual'
            if remarks:
                exp.finance_remarks = remarks
                # Mirror to policy_note if we want consistency
                exp.policy_note = remarks[:255]
            exp.save(update_fields=['finance_selected_amount', 'finance_amount_source', 'finance_remarks', 'policy_note'])
            final_total += fin_amt_dec
            updated += 1

        if errors and updated == 0:
            return Response({'errors': errors}, status=status.HTTP_400_BAD_REQUEST)

        # Update claim's approved_amount to the sum of finance/HR decided line items
        if updated > 0:
            all_exps = claim.trip.expenses.filter(is_deleted=False)
            total = sum(
                float(e.finance_selected_amount) if e.finance_selected_amount is not None else (float(e.hr_selected_amount) if e.hr_selected_amount is not None else float(e.amount))
                for e in all_exps
            )
            claim.approved_amount = total
            claim.executive_approved_amount = total
            claim.save(update_fields=['approved_amount', 'executive_approved_amount'])

        return Response({
            'updated': updated,
            'final_approved_total': round(total, 2) if updated > 0 else None,
            'errors': errors
        })

class TravelAdvanceViewSet(viewsets.ModelViewSet):
    queryset = TravelAdvance.objects.all()
    serializer_class = TravelAdvanceSerializer

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return TravelAdvance.objects.none()
            
        is_admin = hasattr(user, 'role') and user.role.name.lower() in ['admin', 'superuser']
        is_hr = _is_hr(user)
        is_finance = _is_finance_executive(user) or _is_finance_head(user)
        
        if is_admin or is_hr or is_finance:
            queryset = self.queryset
        else:
            q = (
                (Q(trip__user=user) & Q(requester_position=user.active_position_id)) |
                Q(current_approver=user) |
                Q(approver_position=user.active_position_id) |
                Q(trip__current_approver=user) |
                Q(trip__approver_position=user.active_position_id)
            )
            queryset = self.queryset.filter(q)
            
        trip_id = self.request.query_params.get('trip_id')
        if trip_id:
            trip_id = decode_id(trip_id)
            queryset = queryset.filter(trip__trip_id=trip_id)
        return queryset
    def perform_create(self, serializer):
        user = getattr(self.request, 'custom_user', None)
        if not user:
             from rest_framework.exceptions import AuthenticationFailed # type: ignore
             raise AuthenticationFailed("Action requires authentication.")

        trip = serializer.validated_data.get('trip')
        if trip and trip.user != user:
            raise serializers.ValidationError("Unauthorized trip association")

        current_approver, h_level, rm, sm, hod, pos_id = resolve_approver(user)
        is_top_level = (current_approver is None)

        advance = serializer.save(
            status='Submitted',
            requester_position=user.active_position_id,
            current_approver=current_approver,
            approver_position=pos_id,
            hierarchy_level=0 if is_top_level else h_level,
            submitted_at=timezone.now(),
            # Populate snapshots for resilience
            user_name=user.name,
            user_designation=user.designation,
            user_department=user.department,
            reporting_manager_name=rm.name if rm else None,
            senior_manager_name=sm.name if sm else None,
            hod_director_name=hod.name if hod else None
        )
        
        if is_top_level:
            # Instantly trigger parallel HR intimation + Finance Workflow routing for top-level employees
            trigger_parallel_dispatch(advance, user)
            rm_pos = rm.get_current_position() if rm else None
            rm_pos_name = rm_pos.get('name') if rm_pos else None
            if rm and _is_coo_position(rm_pos_name, rm.designation, employee_id=rm.employee_id):
                update_trip_lifecycle(trip, "Advance Requested", "Advance request submitted. Reporting COO bypassed and routed to Finance/HR.")
        else:
            if current_approver:
                is_resubmitted = (advance.status == 'Resubmitted')
                Notification.objects.create(
                    user=current_approver, target_position=pos_id,
                    title="Advance Request Resubmitted" if is_resubmitted else "New Advance Request",
                    message=f"{user.name} has {'resubmitted' if is_resubmitted else 'requested'} an advance of ₹{advance.requested_amount} for Trip {advance.trip.trip_id}.",
                    type='info'
                )
                
                # Notify Requester
                Notification.objects.create(
                    user=user,
                    title="Advance Resubmitted" if is_resubmitted else "Advance Requested",
                    message=f"Your advance request for Trip {advance.trip.trip_id} has been sent to {current_approver.name} for approval.",
                    type='success'
                )
            
        # Notify HR
        notify_hr("New Advance Request", f"{user.name} has requested an advance of ₹{advance.requested_amount} for Trip {advance.trip.trip_id}.")

        self._update_trip_lifecycle(advance)

    def perform_update(self, serializer):
        advance = serializer.save()
        if advance.status == 'Submitted':
            self._update_trip_lifecycle(advance)

    def _update_trip_lifecycle(self, advance):
        update_trip_lifecycle(advance.trip, "Advance Requested", f"Pre-trip advance of ₹{advance.requested_amount} requested for: {advance.purpose[:40]}...")



class DisputeViewSet(viewsets.ModelViewSet):
    queryset = Dispute.objects.all()
    serializer_class = DisputeSerializer
    permission_classes = [IsCustomAuthenticated]

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return Dispute.objects.none()
        
        if hasattr(user, 'role') and user.role.name.lower() in ['admin', 'superuser']:
            return Dispute.objects.all()
        
        return Dispute.objects.filter(raised_by=user)

    def perform_create(self, serializer):
        user = getattr(self.request, 'custom_user', None)
        serializer.save(raised_by=user)

class TripOdometerViewSet(viewsets.ModelViewSet):
    queryset = TripOdometer.objects.all()
    serializer_class = TripOdometerSerializer
    permission_classes = [IsCustomAuthenticated]

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return TripOdometer.objects.none()
            
        is_admin = hasattr(user, 'role') and user.role.name.lower() in ['admin', 'superuser']
        
        if is_admin:
            queryset = self.queryset
        else:
            queryset = self.queryset.filter(trip__user=user)
            
        trip_id = self.request.query_params.get('trip_id')
        if trip_id:
            trip_id = decode_id(trip_id)
            queryset = queryset.filter(trip__trip_id=trip_id)
        return queryset

    def perform_create(self, serializer):
        user = getattr(self.request, 'custom_user', None)
        trip = serializer.validated_data.get('trip')
        if trip and trip.user != user:
            raise serializers.ValidationError("Unauthorized trip association")
            
            
        odo = serializer.save()
        if odo.start_odo_reading:
            trip = odo.trip
            trip.status = 'On-Going'
            trip.save(update_fields=['status'])
            update_trip_lifecycle(trip, "Journey Started", f"Trip journey started with odometer reading {odo.start_odo_reading} KM.")

    def perform_update(self, serializer):
        odo = serializer.save()
        if odo.end_odo_reading:
            trip = odo.trip
            trip.status = 'Completed'
            trip.save(update_fields=['status'])
            update_trip_lifecycle(trip, "Journey Ended", f"Trip journey completed with final odometer reading {odo.end_odo_reading} KM.")

class DashboardStatsView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request):
        user = getattr(request, 'custom_user', None)
        if not user:
            return Response({"error": "User not found"}, status=401)
        
        is_admin = user.active_role.lower() == 'admin'
        is_gh_manager = user.active_role.lower() == 'guesthousemanager'
        is_fin_head = _is_finance_head(user)
        is_fin_exec = _is_finance_executive(user)
        is_finance = is_fin_head or is_fin_exec
        is_cfo = user.active_role.lower() == 'cfo'
        is_hr = _is_hr(user)

        if is_admin or is_gh_manager or is_finance or is_cfo or is_hr:
            trips = Trip.objects.all()
            advances = TravelAdvance.objects.all()
            claims = TravelClaim.objects.all()
            base_expenses = Expense.objects.all()
        else:
            trips = Trip.objects.filter(user=user, requester_position=user.active_position_id)
            advances = TravelAdvance.objects.filter(trip__user=user, requester_position=user.active_position_id)
            claims = TravelClaim.objects.filter(trip__user=user, requester_position=user.active_position_id)
            base_expenses = Expense.objects.filter(trip__user=user, trip__requester_position=user.active_position_id)

        total_trips = trips.filter(status__in=['Approved', 'Settled']).count()
        # Only count records that are actually waiting for approval
        in_review = trips.filter(status='Pending').count()
        
        # Count tasks awaiting this user's approval - MIRROR ApprovalCountView exactly
        counts = _get_actual_pending_tasks_count(user, 'all')
        pending_action = counts['total']

        if is_gh_manager:
            # For GH Manager, count trips that need room booking
            pending_action += Trip.objects.filter(
                accommodation_requests__contains='Request for Room',
                status__in=['Manager Approved', 'Approved']
            ).exclude(room_bookings__isnull=False).distinct().count()
        
        total_expenses = base_expenses.aggregate(Sum('amount'))['amount__sum'] or 0
        
        # compute total approved advances for the selected trips - used for wallet/advance balances
        # mirror TripSerializer.get_total_approved_advance logic (consider executive_approved_amount when >0)
        total_approved_advances = 0.0
        advances_qs = TravelAdvance.objects.filter(
            trip__in=trips,
            status__in=['Paid', 'Transferred', 'COMPLETED']
        )
        for adv in advances_qs:
            amt = float(adv.executive_approved_amount) if float(adv.executive_approved_amount) > 0 else float(adv.requested_amount)
            total_approved_advances += amt
        
        wallet_balance = float(user.carry_forward_balance or 0) + (float(total_approved_advances) - float(total_expenses))
        
        completed_statuses = ['Paid', 'COMPLETED', 'Completed', 'Settled', 'Transferred']
        approved_expenses_qs = base_expenses.filter(
            Q(trip__claim__status__in=completed_statuses) |
            Q(trip__status__in=completed_statuses, trip__consider_as_local=True)
        )
        approved_expenses = approved_expenses_qs.aggregate(Sum('amount'))['amount__sum'] or 0
        
        categories = base_expenses.values('category').annotate(total=Sum('amount'))
        
        recent_trips = trips.order_by('-created_at')[:5]
        recent_data = []
        for t in recent_trips:
            # If trip is approved/completed, show actual total instead of estimate
            if t.status in ['Approved', 'Completed', 'Settled']:
                actual_total = Expense.objects.filter(trip=t).aggregate(Sum('amount'))['amount__sum'] or 0
                display_amount = f"₹{actual_total:,.0f}" if actual_total > 0 else t.cost_estimate
            else:
                display_amount = t.cost_estimate

            recent_data.append({
                "id": t.trip_id,
                "title": f"{'Travel' if t.consider_as_local else 'Trip'} to {t.destination}",
                "subtitle": f"{t.user.name} - {t.purpose}" if (is_admin or is_gh_manager or is_hr) and t.user else t.purpose,
                "status": t.status,
                "amount": display_amount
            })

        kpis = [
            { "title": 'Total Trips', "value": str(total_trips), "label": 'Managed trips' if is_admin or is_finance or is_gh_manager or is_hr else 'Your trips', "icon": 'Briefcase', "color": 'orange' },
            { "title": 'Approved Expenses', "value": f"₹{approved_expenses:,.0f}", "label": 'Finalized' if is_finance else 'Confirmed', "icon": 'CreditCard', "color": 'red' },
            { "title": 'Total Spend' if not is_finance else 'Total Disbursements', "value": f"₹{total_expenses:,.0f}", "label": 'Recorded', "icon": 'TrendingUp', "color": 'magenta' },
            { 
                "title": 'Action Required', 
                "value": str(pending_action), 
                "label": 'Pending your action', 
                "icon": 'Clock', 
                "color": 'yellow' 
            }
        ]

        if is_finance:
            # Shift to vibrant functional colors (strictly within supported index.css classes)
            kpis[0]['color'] = 'orange'
            kpis[1]['color'] = 'magenta'
            kpis[1]['title'] = 'Net Payouts'
            kpis[2]['color'] = 'red'

        return Response({
            "kpis": kpis,
            "recent_activity": recent_data,
            "is_finance_hub": is_finance,
            "expenditure_mix": [
                { 
                    "type": dict(Expense.CATEGORY_CHOICES).get(cat['category'], cat['category']), 
                    "amount": float(cat['total']), 
                    "percentage": (float(cat['total']) / float(total_expenses) * 100) if total_expenses > 0 else 0 
                }
                for cat in categories
            ],
            "total_spend": total_expenses,
            # wallet_balance represents approved advances minus expenses across the user/trips
            "wallet_balance": wallet_balance,
            # advance_balance is simply the sum of approved advances
            "advance_balance": float(total_approved_advances)
        })

class TripSettlementView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request):
        user = getattr(request, 'custom_user', None)
        trip_id_enc = request.query_params.get('trip_id')
        
        if not trip_id_enc:
            # If no trip_id provided, return all trips that are ready for settlement
            # (i.e. trips with approved claims that aren't PAID/COMPLETED yet)
            # For simplicity, returning a list of all trips the user can see for now
            # but we filter for finance if they are finance
            is_finance = _is_finance_executive(user) or _is_finance_head(user)
            if is_finance:
                trips = Trip.objects.filter(status__in=['Claim Submitted', 'Manager Approved', 'Approved'])
            else:
                trips = Trip.objects.filter(user=user)
            
            data = []
            for t in trips:
                advances = TravelAdvance.objects.filter(trip=t, status='COMPLETED').aggregate(Sum('requested_amount'))['requested_amount__sum'] or 0
                claim_amt = 0
                if hasattr(t, 'claim'):
                    claim_amt = t.claim.total_amount
                
                data.append({
                    "trip_id": t.trip_id,
                    "destination": t.destination,
                    "employee": t.user.name if t.user else "Unknown",
                    "advance": float(advances),
                    "claim": float(claim_amt),
                    "balance": float(claim_amt - advances),
                    "status": t.status
                })
            return Response(data)

        trip_id = decode_id(trip_id_enc)
        try:
            trip = Trip.objects.get(trip_id=trip_id)
        except Trip.DoesNotExist:
            return Response({"error": "Trip not found"}, status=404)

        advances_list = TravelAdvance.objects.filter(trip=trip).order_by('created_at')
        total_advance = sum(a.requested_amount for a in advances_list if a.status == 'COMPLETED')
        
        claim_amt = 0
        claim_id = None
        claim_date = None
        if hasattr(trip, 'claim'):
            claim_amt = trip.claim.total_amount
            claim_id = f"CLAIM-{trip.claim.id}"
            claim_date = trip.claim.submitted_at.strftime("%B %d, %Y") if trip.claim.submitted_at else None

        breakdown = []
        for adv in advances_list:
            if adv.status == 'COMPLETED':
                breakdown.append({
                    "id": f"ADV-{adv.id}",
                    "type": "Advance",
                    "description": f"Advance via {adv.payment_mode or 'Bank Transfer'}",
                    "date": adv.payment_date.strftime("%B %d, %Y") if adv.payment_date else adv.created_at.strftime("%B %d, %Y"),
                    "amount": -float(adv.requested_amount),
                    "is_negative": True
                })

        if claim_amt > 0:
            breakdown.append({
                "id": claim_id,
                "type": "Claim",
                "description": f"Expense Claim: {trip.trip_id}",
                "date": claim_date or trip.updated_at.strftime("%B %d, %Y"),
                "amount": float(claim_amt),
                "is_negative": False
            })

        balance = float(claim_amt - total_advance)
        
        return Response({
            "summary": {
                "advance": float(total_advance),
                "claimTotal": float(claim_amt),
                "balance": balance,
                "type": "Payable" if balance > 0 else "Recoverable",
                "status": trip.status
            },
            "breakdown": breakdown,
            "trip": {
                "id": trip.trip_id,
                "destination": trip.destination,
                "employee": trip.user.name if trip.user else "Unknown"
            }
        })

    def post(self, request):
        # Handle Finalize & Settle action
        user = getattr(request, 'custom_user', None)
        if not user:
            return Response({"error": "Authentication required"}, status=401)
            
        trip_id_enc = request.data.get('trip_id')
        if not trip_id_enc:
            return Response({"error": "Trip ID required"}, status=400)
            
        trip_id = decode_id(trip_id_enc)
        try:
            trip = Trip.objects.get(trip_id=trip_id)
        except Trip.DoesNotExist:
            return Response({"error": "Trip not found"}, status=404)

        # Logic to finalize
        trip.status = 'Settled'
        trip.save()
        
        if hasattr(trip, 'claim'):
            trip.claim.status = 'Paid'
            trip.claim.save()

        update_trip_lifecycle(trip, "Final Settlement", f"Accounts settled and closed by {user.name}.")
        
        return Response({"message": "Trip account successfully settled."})

class CFOWarRoomView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request):
        now = timezone.now()
        first_day_of_current_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        # 1. Monthly Total Spend
        monthly_spend = Expense.objects.filter(date__gte=first_day_of_current_month.date()).aggregate(Sum('amount'))['amount__sum'] or 0
        
        # Comparison with last month
        last_month_end = first_day_of_current_month - timezone.timedelta(seconds=1)
        last_month_start = last_month_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        last_month_spend = Expense.objects.filter(date__gte=last_month_start.date(), date__lte=last_month_end.date()).aggregate(Sum('amount'))['amount__sum'] or 0
        
        percent_change_spend = 0
        if last_month_spend > 0:
            percent_change_spend = ((float(monthly_spend) - float(last_month_spend)) / float(last_month_spend)) * 100
        
        # 2. Avg Cost Per Trip
        total_claims = TravelClaim.objects.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        total_trips = Trip.objects.count()
        avg_cost = (float(total_claims) / total_trips) if total_trips > 0 else 0
        
        # 3. Guest House Occupancy
        from guest_house.models import Room, RoomBooking # type: ignore
        total_rooms = Room.objects.count()
        # Active bookings today
        active_bookings = RoomBooking.objects.filter(start_date__lte=now, end_date__gte=now).count()
        occupancy = (active_bookings / total_rooms * 100) if total_rooms > 0 else 0
        
        # 4. Policy Overrides/Disputes
        # Using Category 'Policy' in Dispute as a proxy for overrides
        policy_issues = Dispute.objects.filter(category='Policy').count()
        policy_rate = (policy_issues / total_trips * 100) if total_trips > 0 else 0

        # KPI Stats Construction
        stats = [
            { 
                "title": 'Total Spend (Monthly)', 
                "value": f"₹{monthly_spend:,.0f}", 
                "change": f"{abs(float(percent_change_spend)):.1f}%", # type: ignore
                "trend": 'up' if percent_change_spend >= 0 else 'down',
                "icon": 'IndianRupee'
            },
            { 
                "title": 'Avg Cost per Trip', 
                "value": f"₹{avg_cost:,.0f}", 
                "change": '+2.4%', # Placeholder for complexity
                "trend": 'up', 
                "icon": 'TrendingUp' 
            },
            { 
                "title": 'Guest House Occupancy', 
                "value": f"{occupancy:.1f}%", 
                "change": '+1.8%', 
                "trend": 'up', 
                "icon": 'Building2' 
            },
            { 
                "title": 'Policy Overrides', 
                "value": f"{policy_rate:.1f}%", 
                "change": '-0.2%', 
                "trend": 'down', 
                "icon": 'AlertCircle' 
            },
        ]

        # 5. Spend by Department
        dept_raw = Expense.objects.values('trip__user_department').annotate(total=Sum('amount')).order_by('-total')
        spend_by_dept = [
            { "name": d['trip__user_department'] or "Unknown", "value": float(d['total']) }
            for d in dept_raw
        ]

        # 6. Advance Aging
        advances = TravelAdvance.objects.exclude(status='COMPLETED')
        aging_buckets = { "0-30": 0.0, "31-60": 0.0, "60+": 0.0 }
        for adv in advances:
            days = (now - adv.created_at).days
            amt = float(adv.requested_amount)
            if days <= 30: aging_buckets["0-30"] += amt
            elif days <= 60: aging_buckets["31-60"] += amt
            else: aging_buckets["60+"] += amt
            
        aging_data = [
            { "range": '0-30 Days', "amount": aging_buckets["0-30"], "color": 'success' },
            { "range": '31-60 Days', "amount": aging_buckets["31-60"], "color": 'warning' },
            { "range": '60+ Days', "amount": aging_buckets["60+"], "color": 'danger' },
        ]

        # 7. Critical Anomalies
        anomalies = [
            { "entity": "Logistics Surge", "reason": "Local Conveyance in Metro areas up by 40%", "impact": "High", "action": "Review" },
            { "entity": "Dept Override", "reason": "Bulk travel booking policy bypassed in IT dept", "impact": "Medium", "action": "Investigate" }
        ]

        return Response({
            "stats": stats,
            "report_month": now.strftime("%B %Y"),
            "spend_by_dept": spend_by_dept,
            "aging": aging_data,
            "anomalies": anomalies
        })

import json

class BulkActivityBatchViewSet(viewsets.ModelViewSet):
    queryset = BulkActivityBatch.objects.all()
    serializer_class = BulkActivityBatchSerializer
    permission_classes = [IsCustomAuthenticated]

    def get_object(self):
        pk = self.kwargs.get('pk')
        if isinstance(pk, str) and pk.startswith('BATCH-'):
            self.kwargs['pk'] = pk.replace('BATCH-', '')
        return super().get_object()

    def get_queryset(self):
        user = getattr(self.request, 'custom_user', None)
        if not user:
            return BulkActivityBatch.objects.none()

        role_name = (user.role.name if user.role else '').lower()
        
        # Start with base queryset
        queryset = BulkActivityBatch.objects.all().order_by('-created_at')

        # Filter by trip_id if provided in query params
        trip_id = self.request.query_params.get('trip_id')
        if trip_id:
            queryset = queryset.filter(trip_id=trip_id)

        # Role-based restriction: Admins/Finance/COO see all, others only see their own/involved batches
        privileged_keywords = ['admin', 'finance', 'cfo', 'coo']
        is_privileged = any(kw in role_name for kw in privileged_keywords)
        
        # EXCEPTION: Users defined in Finance Workflow should NOT see global bulk activity batches
        is_in_finance_workflow = FinanceWorkflowStep.objects.filter(user=user, is_active=True).exists()
        if is_in_finance_workflow and not _is_admin(user):
            # Finance staff only see batches they personally submitted
            return queryset.filter(user=user)
        
        if not is_privileged:
            if _is_hr(user):
                # HR sees anything management-approved (next in line) OR assigned to them specifically in this role
                # Exclude 'HR Approved' as that signifies completion for the HR stage
                manager_pos_ids = user.get_active_position_identifiers()
                hr_q = Q(approver_position__in=manager_pos_ids) | Q(current_approver=user, approver_position__isnull=True) | Q(status__in=['Pending', 'Submitted', 'Manager Approved', 'Resubmitted', 'Forwarded'])
                queryset = queryset.filter(hr_q).exclude(status='HR Approved')
            else:
                # Strict Role Filtering: Only see batches submitted in this role OR assigned for approval in this role
                manager_pos_ids = user.get_active_position_identifiers()
                q_requester = Q(user=user) & Q(requester_position__in=manager_pos_ids)
                q_approver = Q(approver_position__in=manager_pos_ids) | Q(current_approver=user, approver_position__isnull=True)
                queryset = queryset.filter(q_requester | q_approver)

        return queryset

    @action(detail=False, methods=['get'], url_path='history')
    def history(self, request):
        """Returns the history of bulk activity uploads for a specific trip."""
        # Use existing logic from get_queryset and list
        queryset = self.filter_queryset(self.get_queryset())
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def template(self, request):
        import openpyxl # type: ignore
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side # type: ignore
        from openpyxl.utils import get_column_letter # type: ignore
        from openpyxl.worksheet.datavalidation import DataValidation # type: ignore
        from openpyxl.styles.differential import DifferentialStyle # type: ignore
        from openpyxl.formatting.rule import FormulaRule # type: ignore
        import datetime
        from django.utils import timezone # type: ignore

        wb = openpyxl.Workbook()

        # ── Hidden locations sheet ───────────────────────────────────────────
        loc_sheet = wb.create_sheet("_Locations")
        loc_sheet.sheet_state = 'hidden'

        # 1. Try to get locations of employees reporting to the current manager
        from api_management.services import get_manager_reports_locations # type: ignore
        user = getattr(request, 'custom_user', None)
        manager_code = user.employee_id if user else None
        
        # Gather clusters from all employees in the reporting chain
        team_locs = set()
        if manager_code:
            for loc in get_manager_reports_locations(manager_code):
                if loc:
                    team_locs.add(loc)

        # Add the manager's own cluster/district (using geo_location priority)
        if user:
            own_loc = (user.office_location or '').strip()
            if own_loc:
                team_locs.add(own_loc)

        # Build final sorted list – strictly team clusters only, no trip history
        if team_locs:
            locations = sorted(team_locs)
        else:
            locations = ["Head Office", "Field Office", "Client Site"]

        for i, loc in enumerate(locations, start=1):
            loc_sheet.cell(row=i, column=1, value=loc)

        # Create a named range for the locations
        loc_range = f"_Locations!$A$1:$A${len(locations)}"
        wb.defined_names['LocationList'] = openpyxl.workbook.defined_name.DefinedName(
            'LocationList', attr_text=loc_range
        )

        # ── Main data sheet ─────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Monthly Activities"

        # ── Styles ──────────────────────────────────────────────────────────
        HEADER_FILL   = PatternFill("solid", fgColor="BB0633")   # brand red
        NOTE_FILL     = PatternFill("solid", fgColor="FFF3CD")   # warm yellow
        HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        NOTE_FONT   = Font(name="Calibri", italic=True, color="856404", size=9)
        DATA_FONT     = Font(name="Calibri", size=10)
        CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
        LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin          = Side(style="thin", color="CCCCCC")
        BORDER        = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Column layout ────────────────────────────────────────────────────
        # A=Date  B=From Location  C=Start Time  D=To Location  E=Reach Time  F=Purpose
        columns = [
            ("Date",          16, CENTER),
            ("From Location", 28, LEFT),
            ("Start Time",    14, CENTER),
            ("To Location",   28, LEFT),
            ("Reach Time",    14, CENTER),
            ("Purpose",       40, LEFT),
        ]

        # Row 1 – Headers
        ws.row_dimensions[1].height = 30
        for col_idx, (label, width, align) in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font   = HEADER_FONT
            cell.fill   = HEADER_FILL
            cell.alignment = CENTER
            cell.border    = BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Row 2 – Instructional note (merged across all columns)
        ws.merge_cells("A2:F2")
        note_cell = ws.cell(row=2, column=1,
            value="📋  Instructions: Date must be ≥ today  |  Start Time & Reach Time must be HH:MM (24h)  |"
                  " Reach Time must be > Start Time | "
                  "Choose From/To Location from the dropdown  |  Purpose is free text")
        note_cell.font      = NOTE_FONT
        note_cell.fill      = NOTE_FILL
        note_cell.alignment = LEFT
        ws.row_dimensions[2].height = 22

        # Row 3 – Sample data row
        # Values for pre-filled row 3
        # Explicitly convert to IST (UTC+5:30) — server TIME_ZONE is UTC
        from zoneinfo import ZoneInfo
        ist_tz    = ZoneInfo('Asia/Kolkata')
        now_ist   = datetime.datetime.now(ist_tz)
        
        # Round time to nearest 5 minutes
        minute_rounded = (now_ist.minute // 5) * 5
        now_rounded = now_ist.replace(minute=minute_rounded, second=0, microsecond=0)

        # Use actual date/time objects so validation formulas can compare them
        today_obj = datetime.date.today()
        # Ensure Row 3 sample data is treated as date/time
        sample_data = [
            today_obj,
            locations[0] if locations else "Office",
            now_rounded.time(), 
            locations[1] if len(locations) > 1 else "Client Site",
            (now_rounded + datetime.timedelta(hours=1)).time(),
            "Site Inspection / Field Visit"
        ]
        ws.row_dimensions[3].height = 20
        for col_idx, val in enumerate(sample_data, start=1):
            cell = ws.cell(row=3, column=col_idx, value=val)
            cell.font      = DATA_FONT
            cell.alignment = columns[col_idx - 1][2]
            cell.border    = BORDER
            # Set explicit formats
            if col_idx == 1: cell.number_format = 'YYYY-MM-DD'
            if col_idx == 2: cell.number_format = 'HH:MM'

        # Rows 4-103 – Empty data rows (100 rows)
        for row in range(4, 104):
            ws.row_dimensions[row].height = 20
            for col_idx in range(1, 7):
                cell = ws.cell(row=row, column=col_idx)
                cell.font      = DATA_FONT
                cell.alignment = columns[col_idx - 1][2]
                cell.border    = BORDER
                # Column A = Date, Column C = Start Time, Column E = Reach Time
                if col_idx == 1: cell.number_format = 'YYYY-MM-DD'
                if col_idx in [3, 5]: cell.number_format = 'HH:MM'

        DATA_ROWS = "3:103"   # applies to sample + 100 blank rows

        # ── Validation 1: Date >= today ──────────────────────────────────────
        today_serial = (datetime.date.today() - datetime.date(1899, 12, 30)).days
        dv_date = DataValidation(
            type="date",
            operator="greaterThanOrEqual",
            formula1=str(today_serial),
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Invalid Date",
            error=f"Date must be on or after today ({datetime.date.today().strftime('%d-%b-%Y')}). "
                  f"Please enter a valid date.",
            showInputMessage=True,
            promptTitle="Date",
            prompt=f"Enter date ≥ {datetime.date.today().strftime('%d-%b-%Y')} (YYYY-MM-DD format)"
        )
        ws.add_data_validation(dv_date)
        dv_date.sqref = "A3:A103"   # Column A

        # -- Updated formula: 5-min increments AND (if Date=Today, Time>Now; else Date>Today) --
        # Now column C
        start_time_formula = '=AND(MOD(ROUND(C3*1440,0),5)=0, IF(A3=TODAY(), C3>MOD(NOW(),1), A3>TODAY()))'
        
        dv_start_time = DataValidation(
            type="custom",
            formula1=start_time_formula,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Invalid Start Time",
            error="Start Time must be a multiple of 5 minutes. If selected date is today, time must be > current time.",
            showInputMessage=True,
            promptTitle="Start Time (HH:MM)",
            prompt="Enter time in HH:MM (24h). If date is today, it must be > current time. Must be a multiple of 5 minutes."
        )
        ws.add_data_validation(dv_start_time)
        dv_start_time.sqref = "C3:C103"   # Column C

        # -- Reach Time Validation: Multiple of 5 mins AND Reach Time > Start Time --
        # Now column E, compared with C
        reach_time_formula = '=AND(MOD(ROUND(E3*1440,0),5)=0, E3>C3)'
        dv_reach_time = DataValidation(
            type="custom",
            formula1=reach_time_formula,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Invalid Reach Time",
            error="Reach Time must be a multiple of 5 minutes and must be GREATER than Start Time.",
            showInputMessage=True,
            promptTitle="Reach Time (HH:MM)",
            prompt="Enter time in HH:MM (24h). Must be > Start Time and a multiple of 5 minutes."
        )
        ws.add_data_validation(dv_reach_time)
        dv_reach_time.sqref = "E3:E103"   # Column E

        # ── Validation 3: From Location dropdown ─────────────────────────────
        # Now column B
        dv_from = DataValidation(
            type="list",
            formula1=f'_Locations!$A$1:$A${len(locations)}',
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Invalid Location",
            error="Please select a location from the dropdown list.",
            showInputMessage=True,
            promptTitle="From Location",
            prompt="Select the starting location from the dropdown."
        )
        ws.add_data_validation(dv_from)
        dv_from.sqref = "B3:B103"   # Column B

        # ── Validation 4: To Location dropdown ───────────────────────────────
        # Now column D
        dv_to = DataValidation(
            type="list",
            formula1=f'_Locations!$A$1:$A${len(locations)}',
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Invalid Location",
            error="Please select a location from the dropdown list.",
            showInputMessage=True,
            promptTitle="To Location",
            prompt="Select the destination location from the dropdown."
        )
        ws.add_data_validation(dv_to)
        dv_to.sqref = "D3:D103"   # Column D

        # Highlight Red if From == To (B and D)
        from openpyxl.styles import PatternFill # type: ignore
        red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
        ws.conditional_formatting.add("B3:D103",
            openpyxl.formatting.rule.FormulaRule(formula=["=$B3=$D3"], stopIfTrue=True, fill=red_fill))


        # Column E (Purpose) – no list validation, free text; just an input hint
        dv_purpose = DataValidation(
            type="textLength",
            operator="greaterThan",
            formula1="0",
            showErrorMessage=True,
            errorTitle="Purpose Required",
            error="Please describe the purpose of your visit.",
            showInputMessage=True,
            promptTitle="Purpose",
            prompt="Briefly describe the reason for this visit (e.g. Site Inspection, Client Meeting)."
        )
        ws.add_data_validation(dv_purpose)
        dv_purpose.sqref = "F3:F103"   # Column F

        # ── Freeze panes below header + note rows ────────────────────────────
        ws.freeze_panes = "A3"

        # ── Output ───────────────────────────────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="bulk_activity_template.xlsx"'
        return response


    @action(detail=False, methods=['post'])
    def upload(self, request):
        user = getattr(request, 'custom_user', None)
        if not user:
            return Response({"error": "Authentication required"}, status=401)

        file = request.FILES.get('file')
        json_data = request.data.get('data_json') # For resubmitting/editing rejected rows
        trip_id = request.data.get('trip_id') # Selected by user in UI
        parent_batch_id = request.data.get('parent_batch_id')
        
        if not file and not json_data:
            return Response({"error": "No file or data provided"}, status=400)
            
        rows = []
        file_name = "resubmission.json"
        
        try:
            if file:
                file_name = file.name
                import pandas as pd
                df = pd.read_excel(file)
                # drop the instruction row that sits just below the header
                import datetime as dt_mod

                def _extract_excel_time(val, default_time):
                    if pd.isna(val) or str(val).strip() in ('', 'nan', 'NaT'):
                        return default_time
                    
                    # 1. Handle actual datetime/time objects (most robust)
                    if isinstance(val, dt_mod.time):
                        return val.strftime('%H:%M')
                    if isinstance(val, dt_mod.datetime):
                        return val.astimezone(timezone.get_current_timezone()).strftime('%H:%M') if timezone.is_aware(val) else val.strftime('%H:%M')
                    
                    s = str(val).strip()
                    if not s: return default_time

                    # 2. Handle numeric/float representation (Excel serial time)
                    try:
                        f = float(s)
                        if 0 <= f < 1:
                            total_secs = int(f * 86400)
                            return f"{total_secs // 3600:02d}:{(total_secs % 3600) // 60:02d}"
                    except (ValueError, TypeError):
                        pass

                    # 3. Handle string formats (HH:MM, HH:MM:SS, or "YYYY-MM-DD HH:MM:SS")
                    if ' ' in s: # Common for datetime strings
                        s = s.split(' ')[-1]
                    
                    parts = s.split(':')
                    if len(parts) >= 2:
                        # Extract the HH and MM, but ensure we don't pick up the date part if it survived the split
                        hour_part = parts[0][-2:].zfill(2)
                        min_part = parts[1][:2].zfill(2)
                        return f"{hour_part}:{min_part}"
                    
                    return default_time

                for index, row in df.iterrows():
                    if row.dropna().empty: continue
                    
                    # Date extraction
                    date_raw = row.get('Date', row.get('Date (YYYY-MM-DD)', ''))
                    if pd.isna(date_raw): continue
                    
                    if isinstance(date_raw, (dt_mod.date, dt_mod.datetime)):
                        date_val = date_raw.strftime('%Y-%m-%d')
                    else:
                        date_val = str(date_raw).strip()
                        if 'instruc' in date_val.lower(): continue
                        # If string contains time, strip it
                        if ' ' in date_val: date_val = date_val.split(' ')[0]
                        if len(str(date_val)) > 10: date_val = str(date_val)[:10]

                    # Time extraction
                    start_time_raw = row.get('Start Time', row.get('Time', ''))
                    start_time_str = _extract_excel_time(start_time_raw, '09:00')

                    reach_time_raw = row.get('Reach Time', row.get('End Time', ''))
                    reach_time_str = _extract_excel_time(reach_time_raw, '18:00')

                    origin = str(row.get('From Location', row.get('from location', ''))).strip()
                    destination = str(row.get('To Location', row.get('to location', ''))).strip()
                    if origin and destination and origin.lower() == destination.lower():
                        raise Exception(f"Row {index + 3}: From Location and To Location cannot be the same.")

                    rows.append({
                        "date": date_val, "start_time": start_time_str, "reach_time": reach_time_str, "mode": "Bike",
                        "origin_route": origin, "destination_route": destination,
                        "vehicle": "Own Bike", "visit_intent": str(row.get('Purpose', row.get('purpose', ''))),
                        "daily_allowance": None
                    })
            else:
                 # Direct JSON submission (resubmitting rejected rows)
                raw_rows = json_data if isinstance(json_data, list) else []
                rows = raw_rows # Preserve all fields including odometer/subtype
                from datetime import datetime
                file_name = f"resubmitted_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
            
            if not rows:
                return Response({"error": "No valid activity rows found"}, status=400)

            if not user:
                return Response({"error": "User not found in request"}, status=400)

            # Link to trip object if possible
            trip_obj = None
            if trip_id:
                trip_obj = Trip.objects.filter(trip_id=trip_id).first()

            # Linking to correct approver (Priority: Trip's current approver, then manager chain)
            rm, sm, hod = user.reporting_manager, user.senior_manager, user.hod_director
            
            def is_mgmt(u):
                return u and hasattr(u, 'role') and u.role and u.role.name.lower() in ['admin', 'superuser', 'it admin']

            # Check if Employee API is unreachable (to prevent accidental auto-approval for standard users)
            from core.middleware import should_skip_external_api
            user_role = (user.role.name.lower() if user.role else '')
            if not user._get_api_data() and not should_skip_external_api() and user_role not in ['admin', 'superuser', 'it-admin']:
                raise ValidationError({
                    "detail": "Failed to resolve employee profile from the HCM API. "
                              "Please contact support or ensure the staging server has network access to the API."
                })

            # ALWAYS start from the beginning of the management chain for Bulk Uploads and Resubmissions
            # (Following the flow of bulk upload records only, as requested)
            current_approver, h_level, rm, sm, hod, pos_id = resolve_approver(user)
            is_top_level = (current_approver is None)
            
            batch_status = 'Approved' if is_top_level else ('Resubmitted' if parent_batch_id else 'Submitted')
            
            batch = BulkActivityBatch.objects.create(
                user=user, requester_position=user.active_position_id,
                trip=trip_obj, trip_id=trip_id, file_name=file_name,
                data_json=rows, status=batch_status,
                current_approver=current_approver, approver_position=pos_id,
                hierarchy_level=0 if is_top_level else h_level
            )
            
            if is_top_level:
                # Instantly trigger parallel self-approval and dispatch HR Intimation for top-level employees
                trigger_parallel_dispatch(batch, user)
            else:
                if trip_obj:
                    from .utils import build_approval_chain
                    trip_obj.approval_chain = build_approval_chain(user)
                    trip_obj.status = batch_status
                    trip_obj.current_approver = current_approver
                    trip_obj.approver_position = pos_id
                    trip_obj.hierarchy_level = batch.hierarchy_level
                    trip_obj.save()
            
            # If this is a resubmission, mark the parent batch as Resolved
            if parent_batch_id:
                try:
                    parent_batch = BulkActivityBatch.objects.get(id=parent_batch_id)
                    parent_batch.status = 'Resolved'
                    parent_batch.save()
                    
                    # Notify user about successful resubmission
                    Notification.objects.create(
                        user=user,
                        target_position=batch.requester_position,
                        title="Resubmission",
                        message="The rejected record has been resubmitted",
                        type='success'
                    )
                except (BulkActivityBatch.DoesNotExist, Trip.DoesNotExist):
                    pass
            
            if not is_top_level and current_approver:
                is_resub = 'resubmitted' if parent_batch_id else 'submitted'
                Notification.objects.create(
                    user=current_approver, 
                    target_position=batch.approver_position,
                    title=f"{is_resub.capitalize()} Bulk Activity Batch",
                    message=f"{user.name} {is_resub} a bulk travel log for approval.", 
                    type='info'
                )
            
            return Response(BulkActivityBatchSerializer(batch).data)
        except Exception as e:
            return Response({"error": f"Failed to process submission: {str(e)}"}, status=400)

    @action(detail=True, methods=['post'], url_path='approve')
    def approve_batch(self, request, pk=None):
        batch = self.get_object()
        user = getattr(request, 'custom_user', None)
        if not user:
             return Response({"error": "Authentication required"}, status=401)
             
        # Sync any row-level rejections/remarks from frontend before processing workflow
        if 'data_json' in request.data and request.data['data_json']:
            batch.data_json = request.data['data_json']
            batch.save()

        # Route to centralized workflow
        return handle_workflow_action(batch, 'Approve', user, request.data)

    @action(detail=True, methods=['post'], url_path='reject')
    def reject_batch(self, request, pk=None):
        batch = self.get_object()
        user = getattr(request, 'custom_user', None)
        if not user:
             return Response({"error": "Authentication required"}, status=401)
             
        # Sync any row-level rejections/remarks from frontend before processing workflow
        if 'data_json' in request.data and request.data['data_json']:
            batch.data_json = request.data['data_json']
            batch.save()

        # Route to centralized workflow
        return handle_workflow_action(batch, 'Reject', user, request.data)

from django.db import models
from rest_framework.decorators import action

# --- MASTER VIEWSETS ---

class MasterActionMixin:
    """Mixin to provide restore and include_deleted functionality to Master ViewSets."""
    pagination_class = None
    def get_queryset(self):
        include_deleted = self.request.query_params.get('include_deleted') == 'true'
        model = self.queryset.model
        if include_deleted:
            if hasattr(model, 'all_objects'):
                return model.all_objects.all().order_by('id')
            return model.objects.all().order_by('id')
        
        # Default view (active only)
        if hasattr(model, 'all_objects'):
            qs = model.objects.all()
        else:
            qs = model.objects.all()
            
        if hasattr(model, 'status'):
            qs = qs.filter(status=True)
            
        return qs.order_by('id')

    def perform_create(self, serializer):
        # Master unique name check: if deleted record exists, restore it instead
        model = self.queryset.model
        unique_field = None
        
        # Identify the unique character field (usually mode_name, class_name, etc.)
        for field in model._meta.fields:
            if isinstance(field, (models.CharField, models.TextField)) and field.unique:
                unique_field = field.name
                break
        
        if unique_field:
            val = serializer.validated_data.get(unique_field)
            if val:
                existing = model.all_objects.filter(**{unique_field: val}, is_deleted=True).first()
                if existing:
                    existing.is_deleted = False
                    existing.deleted_at = None
                    if hasattr(existing, 'status'):
                        existing.status = True
                    # Update other fields to the new values provided
                    for k, v in serializer.validated_data.items():
                        setattr(existing, k, v)
                    existing.save()
                    # Return the restored object
                    serializer.instance = existing
                    return
        
        serializer.save()

    @action(detail=True, methods=['post'])
    def restore(self, request, pk=None):
        instance = self.queryset.model.all_objects.get(pk=pk)
        instance.is_deleted = False
        instance.deleted_at = None
        instance.deleted_by = None
        if hasattr(instance, 'status'):
            instance.status = True
        instance.save()
        return Response({"status": "restored"})

class TravelModeMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = TravelModeMaster.objects.all()
    serializer_class = TravelModeMasterSerializer

class BookingTypeMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = BookingTypeMaster.objects.all()
    serializer_class = BookingTypeMasterSerializer

class OperatorMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = OperatorMaster.objects.all()
    serializer_class = OperatorMasterSerializer

class TravelClassMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = TravelClassMaster.objects.all()
    serializer_class = TravelClassMasterSerializer

class VehicleMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = VehicleMaster.objects.all()
    serializer_class = VehicleMasterSerializer

class ProviderMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = ProviderMaster.objects.all()
    serializer_class = ProviderMasterSerializer

class TicketStatusMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = TicketStatusMaster.objects.all()
    serializer_class = TicketStatusMasterSerializer

class QuotaTypeMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = QuotaTypeMaster.objects.all()
    serializer_class = QuotaTypeMasterSerializer

class LocalTravelModeMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = LocalTravelModeMaster.objects.all()
    serializer_class = LocalTravelModeMasterSerializer

class LocalProviderMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = LocalProviderMaster.objects.all()
    serializer_class = LocalProviderMasterSerializer

class LocalSubTypeMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = LocalSubTypeMaster.objects.all()
    serializer_class = LocalSubTypeMasterSerializer

class StayTypeMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = StayTypeMaster.objects.all()
    serializer_class = StayTypeMasterSerializer

class RoomTypeMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = RoomTypeMaster.objects.all()
    serializer_class = RoomTypeMasterSerializer

class StayBookingTypeMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = StayBookingTypeMaster.objects.all()
    serializer_class = StayBookingTypeMasterSerializer

class StayBookingSourceMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = StayBookingSourceMaster.objects.all()
    serializer_class = StayBookingSourceMasterSerializer

class MealCategoryMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = MealCategoryMaster.objects.all()
    serializer_class = MealCategoryMasterSerializer

class MealTypeMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = MealTypeMaster.objects.all()
    serializer_class = MealTypeMasterSerializer

class MealSourceMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = MealSourceMaster.objects.all()
    serializer_class = MealSourceMasterSerializer

class MealProviderMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = MealProviderMaster.objects.all()
    serializer_class = MealProviderMasterSerializer

class JobReportViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = JobReport.objects.all()
    serializer_class = JobReportSerializer

    def get_queryset(self):
        trip_id = self.request.query_params.get('trip_id')
        if trip_id:
            return self.queryset.filter(trip_id=trip_id)
        return self.queryset

    def perform_create(self, serializer):
        serializer.save(user=self.request.custom_user)

class IncidentalTypeMasterViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = IncidentalTypeMaster.objects.all()
    serializer_class = IncidentalTypeMasterSerializer

class MasterModuleViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = MasterModule.objects.all()
    serializer_class = MasterModuleSerializer

    @action(detail=False, methods=['get'])
    @permission_classes_decorator([AllowAny])
    def seed_metadata(self, request):
        # 0. Clean Sweep (Ensure fresh definitions)
        CustomMasterDefinition.objects.all().delete()
        MasterModule.objects.all().delete()

        # 1. Modules
        modules_data = [
            {'name': 'Travel', 'display_order': 1},
            {'name': 'Local Conveyance', 'display_order': 2},
            {'name': 'Stay', 'display_order': 3},
            {'name': 'Food', 'display_order': 4},
            {'name': 'Incidental', 'display_order': 5},
        ]
        
        modules = {}
        for m_data in modules_data:
            mod, _ = MasterModule.objects.update_or_create(
                name=m_data['name'],
                defaults={'display_order': m_data['display_order']}
            )
            modules[m_data['name']] = mod

        # 2. Definitions
        definitions_data = [
            {'module': 'Travel', 'table_name': 'Travel Modes', 'endpoint': 'travel-mode-masters', 'fields': 'mode_name,status'},
            {'module': 'Travel', 'table_name': 'Booking Types', 'endpoint': 'booking-type-masters', 'fields': 'booking_type,status'},
            {'module': 'Travel', 'table_name': 'Operators (Flight/Train/Bus)', 'endpoint': 'operator-masters', 'fields': 'operator_name,is_flight,is_train,is_bus,status'},
            {'module': 'Travel', 'table_name': 'Travel Classes', 'endpoint': 'travel-class-masters', 'fields': 'class_name,is_flight,is_train,is_bus,status'},
            {'module': 'Travel', 'table_name': 'Vehicles', 'endpoint': 'vehicle-masters', 'fields': 'vehicle_name,is_bus,is_intercity_cab,status'},
            {'module': 'Travel', 'table_name': 'Providers', 'endpoint': 'provider-masters', 'fields': 'provider_name,is_flight,is_train,is_bus,is_intercity_cab,status'},
            {'module': 'Travel', 'table_name': 'Ticket Statuses', 'endpoint': 'ticket-status-masters', 'fields': 'status_name,is_flight,is_train,is_bus,is_intercity_cab,status'},
            {'module': 'Travel', 'table_name': 'Quota Types', 'endpoint': 'quota-type-masters', 'fields': 'quota_name,status'},
            
            {'module': 'Local Conveyance', 'table_name': 'Local Modes', 'endpoint': 'local-travel-mode-masters', 'fields': 'mode_name,status'},
            {'module': 'Local Conveyance', 'table_name': 'Local Providers', 'endpoint': 'local-provider-masters', 'fields': 'provider_name,is_car,is_bike,is_auto,is_bus,is_metro,status'},
            {'module': 'Local Conveyance', 'table_name': 'Local Sub-Types', 'endpoint': 'local-sub-type-masters', 'fields': 'sub_type,is_car,is_bike,is_auto,status'},
            
            {'module': 'Stay', 'table_name': 'Stay Types', 'endpoint': 'stay-type-masters', 'fields': 'stay_type,status'},
            {'module': 'Stay', 'table_name': 'Room Types', 'endpoint': 'room-type-masters', 'fields': 'room_type,status'},
            {'module': 'Stay', 'table_name': 'Stay Booking Types', 'endpoint': 'stay-booking-type-masters', 'fields': 'booking_type,status'},
            {'module': 'Stay', 'table_name': 'Stay Booking Sources', 'endpoint': 'stay-booking-source-masters', 'fields': 'source_name,status'},
            
            {'module': 'Food', 'table_name': 'Meal Categories', 'endpoint': 'meal-category-masters', 'fields': 'category_name,status'},
            {'module': 'Food', 'table_name': 'Meal Types', 'endpoint': 'meal-type-masters', 'fields': 'meal_type,status'},
            {'module': 'Food', 'table_name': 'Meal Sources', 'endpoint': 'meal-source-masters', 'fields': 'source_name,status'},
            {'module': 'Food', 'table_name': 'Meal Providers', 'endpoint': 'meal-provider-masters', 'fields': 'provider_name,status'},
            
            {'module': 'Incidental', 'table_name': 'Incidental Types', 'endpoint': 'incidental-type-masters', 'fields': 'expense_type,category,status'},
        ]
        
        for d_data in definitions_data:
            CustomMasterDefinition.objects.update_or_create(
                table_name=d_data['table_name'],
                defaults={
                    'module_ref': modules[d_data['module']],
                    'api_endpoint': d_data['endpoint'],
                    'fields_list': d_data['fields'],
                    'is_system': True
                }
            )

        # 3. Seed Initial Records for popular masters
        # Stay Booking Types
        for btype in ['Self Booking', 'Company Booking', 'Online Booking']:
            StayBookingTypeMaster.objects.get_or_create(booking_type=btype, defaults={'status': True})
        
        # Room Types
        for rtype in ['Standard', 'Deluxe', 'Executive', 'Suite']:
            RoomTypeMaster.objects.get_or_create(room_type=rtype, defaults={'status': True})

        # Stay Types
        for stype in ['Hotel Stay', 'Guest House', 'Self Stay']:
            StayTypeMaster.objects.get_or_create(stay_type=stype, defaults={'status': True})

        # Local Modes
        for lmode in ['Car', 'Bike', 'Auto', 'Public Transport']:
            LocalTravelModeMaster.objects.get_or_create(mode_name=lmode, defaults={'status': True})

        # Meal Categories
        for mcat in ['Self Meal', 'Business Lunch', 'Dinner with Client']:
            MealCategoryMaster.objects.get_or_create(category_name=mcat, defaults={'status': True})

        return Response({
            "status": "Master Metadata and Initial Records Seeded Successfully", 
            "modules": len(modules_data), 
            "definitions": len(definitions_data),
            "records_note": "Added initial values for Stay Booking Types, Room Types, Stay Types, Local Modes, and Meal Categories."
        })

class CustomMasterDefinitionViewSet(viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = CustomMasterDefinition.objects.all()
    serializer_class = CustomMasterDefinitionSerializer

class CustomMasterValueViewSet(MasterActionMixin, viewsets.ModelViewSet):
    permission_classes = [IsCustomAuthenticated]
    queryset = CustomMasterValue.objects.all()
    serializer_class = CustomMasterValueSerializer
    filterset_fields = ['definition']

from core.models import Role
from django.db import models as django_models

MASTER_GROUPS_CONFIG = {
    'travel': {
        'filename': 'Long_Distance_Masters',
        'sheets': [
            {'sheet_name': 'Travel Modes', 'model': TravelModeMaster, 'fields': ['mode_name', 'status']},
            {'sheet_name': 'Providers', 'model': ProviderMaster, 'fields': ['provider_name', 'is_flight', 'is_train', 'is_bus', 'is_intercity_cab', 'status']},
            {'sheet_name': 'Operators', 'model': OperatorMaster, 'fields': ['operator_name', 'is_flight', 'is_train', 'is_bus', 'status']},
            {'sheet_name': 'Travel Classes', 'model': TravelClassMaster, 'fields': ['class_name', 'is_flight', 'is_train', 'is_bus', 'status']},
            {'sheet_name': 'Vehicles', 'model': VehicleMaster, 'fields': ['vehicle_name', 'is_bus', 'is_intercity_cab', 'status']},
            {'sheet_name': 'Booking Types', 'model': BookingTypeMaster, 'fields': ['booking_type', 'status']},
            {'sheet_name': 'Ticket Statuses', 'model': TicketStatusMaster, 'fields': ['status_name', 'is_flight', 'is_train', 'is_bus', 'is_intercity_cab', 'status']},
            {'sheet_name': 'Quota Types', 'model': QuotaTypeMaster, 'fields': ['quota_name', 'status']},
        ]
    },
    'local': {
        'filename': 'Local_Conveyance_Masters',
        'sheets': [
            {'sheet_name': 'Travel Modes', 'model': LocalTravelModeMaster, 'fields': ['mode_name', 'status']},
            {'sheet_name': 'Providers', 'model': LocalProviderMaster, 'fields': ['provider_name', 'is_car', 'is_bike', 'is_auto', 'is_bus', 'is_metro', 'status']},
            {'sheet_name': 'Sub Types', 'model': LocalSubTypeMaster, 'fields': ['sub_type', 'is_car', 'is_bike', 'is_auto', 'status']},
        ]
    },
    'stay': {
        'filename': 'Stay_Lodging_Masters',
        'sheets': [
            {'sheet_name': 'Stay Types', 'model': StayTypeMaster, 'fields': ['stay_type', 'status']},
            {'sheet_name': 'Room Types', 'model': RoomTypeMaster, 'fields': ['room_type', 'status']},
            {'sheet_name': 'Booking Types', 'model': StayBookingTypeMaster, 'fields': ['booking_type', 'status']},
            {'sheet_name': 'Booking Sources', 'model': StayBookingSourceMaster, 'fields': ['source_name', 'status']},
        ]
    },
    'food': {
        'filename': 'Food_Refreshments_Masters',
        'sheets': [
            {'sheet_name': 'Meal Categories', 'model': MealCategoryMaster, 'fields': ['category_name', 'status']},
            {'sheet_name': 'Meal Types', 'model': MealTypeMaster, 'fields': ['meal_type', 'status']},
            {'sheet_name': 'Meal Sources', 'model': MealSourceMaster, 'fields': ['source_name', 'status']},
            {'sheet_name': 'Meal Providers', 'model': MealProviderMaster, 'fields': ['provider_name', 'status']},
        ]
    },
    'incidental': {
        'filename': 'Incidental_Expenses_Masters',
        'sheets': [
            {'sheet_name': 'Incidental Types', 'model': IncidentalTypeMaster, 'fields': ['expense_type', 'category', 'status']},
        ]
    },
    'access': {
        'filename': 'Access_Control_Masters',
        'sheets': [
            {'sheet_name': 'Role Permissions', 'model': Role, 'fields': ['name', 'description', 'permissions']},
        ]
    },
    'entitlement': {
        'filename': 'Entitlement_Policy_Masters',
        'sheets': [
            {'sheet_name': 'Cadres', 'model': Cadre, 'fields': ['name', 'description', 'designation_keywords']},
            {'sheet_name': 'Eligibility Rules', 'model': EligibilityRule, 'fields': [
                'cadre', 'air_allowed', 'air_class', 'train_allowed', 'train_class', 
                'bus_allowed', 'bus_class', 'car_allowed', 'car_notes', 'local_conveyance_allowed', 
                'local_conveyance_type', 'company_guest_house_status', 
                'accommodation_state_hq', 'accommodation_districts', 'accommodation_others', 
                'daily_allowance_amount', 'monthly_tour_daily_allowance_amount', 'own_stay_state_hq_pct', 'own_stay_districts_pct', 
                'own_stay_others_pct'
            ]}
        ]
    }
}

class MasterBulkExportView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def get(self, request):
        user = getattr(request, 'custom_user', None)
        if not _is_admin(user):
            return Response({"error": "Unauthorized"}, status=403)

        group_id = request.query_params.get('group')
        if not group_id or group_id not in MASTER_GROUPS_CONFIG:
            return Response({"error": "Invalid or missing group parameter"}, status=400)

        group_config = MASTER_GROUPS_CONFIG[group_id]
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_info in group_config['sheets']:
                model = sheet_info['model']
                fields = sheet_info['fields']
                
                if hasattr(model, 'all_objects'):
                    queryset = model.all_objects.all().order_by('id')
                else:
                    queryset = model.objects.all().order_by('id')

                rows = []
                for obj in queryset:
                    row_data = {}
                    for field in fields:
                        val = getattr(obj, field)
                        if isinstance(val, (dict, list)):
                            val = json.dumps(val)
                        elif field == 'status' and isinstance(val, bool):
                            val = 'ACTIVE' if val else 'INACTIVE'
                        elif val is not None and not isinstance(val, (int, float, bool, str)):
                            # For related fields, store their string representation
                            val = str(val)
                        row_data[field.upper()] = val
                    rows.append(row_data)

                if rows:
                    df = pd.DataFrame(rows)
                else:
                    df = pd.DataFrame(columns=[f.upper() for f in fields])

                df.to_excel(writer, index=False, sheet_name=sheet_info['sheet_name'])

        output.seek(0)
        
        # Add Data Validation for Dropdowns using openpyxl
        from openpyxl import load_workbook
        from openpyxl.worksheet.datavalidation import DataValidation
        
        wb = load_workbook(output)
        for sheet_info in group_config['sheets']:
            sheet_name = sheet_info['sheet_name']
            model = sheet_info['model']
            fields = sheet_info['fields']
            
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Check each column
                for col_idx, field_name in enumerate(fields, start=1):
                    # We get the column letter
                    col_letter = ws.cell(row=1, column=col_idx).column_letter
                    
                    try:
                        field = model._meta.get_field(field_name)
                    except:
                        field = None
                        
                    dv = None
                    if field_name == 'status':
                        dv = DataValidation(type="list", formula1='"ACTIVE,INACTIVE"', allow_blank=True)
                    elif field and isinstance(field, django_models.BooleanField):
                        dv = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
                    elif field and getattr(field, 'choices', None):
                        choices_list = [str(c[0]) for c in field.choices]
                        dv = DataValidation(type="list", formula1=f'"{",".join(choices_list)}"', allow_blank=True)
                        
                    if dv:
                        ws.add_data_validation(dv)
                        dv.add(f"{col_letter}2:{col_letter}1000")
                        
        # Save back to a new output BytesIO
        new_output = io.BytesIO()
        wb.save(new_output)
        new_output.seek(0)

        filename = f"{group_config['filename']}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        response = HttpResponse(
            new_output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

class MasterBulkImportView(APIView):
    permission_classes = [IsCustomAuthenticated]

    def post(self, request):
        user = getattr(request, 'custom_user', None)
        if not _is_admin(user):
            return Response({"error": "Unauthorized"}, status=403)

        group_id = request.data.get('group')
        if not group_id or group_id not in MASTER_GROUPS_CONFIG:
            return Response({"error": "Invalid or missing group parameter"}, status=400)

        if 'file' not in request.FILES:
            return Response({"error": "No file uploaded"}, status=400)

        import_file = request.FILES['file']
        group_config = MASTER_GROUPS_CONFIG[group_id]

        report = {}
        total_created = 0
        total_updated = 0
        total_errors = []

        try:
            xls = pd.ExcelFile(import_file)
            sheet_names = xls.sheet_names

            for sheet_info in group_config['sheets']:
                sheet_name = sheet_info['sheet_name']
                model = sheet_info['model']
                fields = sheet_info['fields']

                matched_sheet = None
                for sn in sheet_names:
                    if sn.strip().lower() == sheet_name.strip().lower():
                        matched_sheet = sn
                        break

                if not matched_sheet:
                    continue

                df = pd.read_excel(import_file, sheet_name=matched_sheet)
                df.columns = [str(c).strip().upper() for c in df.columns]

                unique_field = None
                for field_obj in model._meta.fields:
                    if isinstance(field_obj, (django_models.CharField, django_models.TextField)) and field_obj.unique:
                        unique_field = field_obj.name
                        break

                unique_together_fields = None
                if not unique_field and model._meta.unique_together:
                    unique_together_fields = model._meta.unique_together[0]

                sheet_created = 0
                sheet_updated = 0
                sheet_errors = 0

                for idx, row in df.iterrows():
                    try:
                        record_id = row.get('ID')
                        if pd.isna(record_id) or str(record_id).strip() == '':
                            record_id = None
                        else:
                            try:
                                record_id = int(float(record_id))
                            except ValueError:
                                record_id = None

                        obj = None
                        if record_id:
                            if hasattr(model, 'all_objects'):
                                obj = model.all_objects.filter(id=record_id).first()
                            else:
                                obj = model.objects.filter(id=record_id).first()

                        if not obj and unique_field:
                            uf_upper = unique_field.upper()
                            unique_val = row.get(uf_upper)
                            if not pd.isna(unique_val) and str(unique_val).strip() != '':
                                unique_val_str = str(unique_val).strip()
                                if hasattr(model, 'all_objects'):
                                    obj = model.all_objects.filter(**{unique_field: unique_val_str}).first()
                                else:
                                    obj = model.objects.filter(**{unique_field: unique_val_str}).first()

                        if not obj and unique_together_fields:
                            lookup = {}
                            all_fields_present = True
                            for f in unique_together_fields:
                                f_upper = f.upper()
                                val = row.get(f_upper)
                                if not pd.isna(val) and str(val).strip() != '':
                                    lookup[f] = str(val).strip()
                                else:
                                    all_fields_present = False
                                    break
                            if all_fields_present:
                                if hasattr(model, 'all_objects'):
                                    obj = model.all_objects.filter(**lookup).first()
                                else:
                                    obj = model.objects.filter(**lookup).first()

                        update_data = {}
                        for f in fields:
                            if f == 'id':
                                continue
                            f_upper = f.upper()
                            if f_upper in df.columns:
                                val = row.get(f_upper)
                                
                                if f == 'status':
                                    if pd.isna(val) or str(val).strip() == '':
                                        val = True
                                    else:
                                        val_str = str(val).strip().lower()
                                        val = val_str in ['true', '1', 'active', 'yes']
                                elif f in ['permissions', 'designation_keywords']:
                                    if pd.isna(val) or str(val).strip() == '':
                                        val = [] if f == 'designation_keywords' else {}
                                    else:
                                        try:
                                            val = json.loads(str(val))
                                        except:
                                            if f == 'designation_keywords':
                                                val = [k.strip() for k in str(val).split(',') if k.strip()]
                                            else:
                                                val = {}
                                elif model._meta.get_field(f).is_relation:
                                    field_obj = model._meta.get_field(f)
                                    related_model = field_obj.related_model
                                    if pd.isna(val) or str(val).strip() == '':
                                        val = None
                                    else:
                                        val_str = str(val).strip()
                                        # Match by name or other unique text fields
                                        related_unique_field = None
                                        for rf in related_model._meta.fields:
                                            if rf.unique and isinstance(rf, (django_models.CharField, django_models.TextField)):
                                                related_unique_field = rf.name
                                                break
                                        if not related_unique_field:
                                            related_unique_field = 'name'
                                        val = related_model.objects.filter(**{related_unique_field: val_str}).first()
                                elif isinstance(model._meta.get_field(f), django_models.BooleanField):
                                    if pd.isna(val) or str(val).strip() == '':
                                        val = False
                                    else:
                                        val_str = str(val).strip().lower()
                                        val = val_str in ['true', '1', 'active', 'yes']
                                else:
                                    if pd.isna(val):
                                        val = None
                                    else:
                                        val = str(val).strip()

                                update_data[f] = val

                        if obj:
                            if hasattr(obj, 'is_deleted') and obj.is_deleted:
                                obj.is_deleted = False
                                obj.deleted_at = None
                            
                            for k, v in update_data.items():
                                setattr(obj, k, v)
                            obj.save()
                            sheet_updated += 1
                            total_updated += 1
                        else:
                            model.objects.create(**update_data)
                            sheet_created += 1
                            total_created += 1

                    except Exception as row_err:
                        sheet_errors += 1
                        total_errors.append(f"Sheet '{sheet_name}', Row {idx + 2} error: {str(row_err)}")

                report[sheet_name] = {
                    'created': sheet_created,
                    'updated': sheet_updated,
                    'errors': sheet_errors
                }

            return Response({
                "message": "Bulk import completed",
                "total_created": total_created,
                "total_updated": total_updated,
                "details": report,
                "errors": total_errors
            })

        except Exception as e:
            return Response({"error": f"Failed to process import: {str(e)}"}, status=500)
